"""
엑셀 파일 처리 서비스
간호사 정보 엑셀 업로드/다운로드 관련 기능 제공
"""
import pandas as pd
import uuid
import tempfile
import os
import re
from typing import List, Dict, Any, Tuple, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from datetime import datetime

from db.models import Nurse as NurseModel, Group as GroupModel, Office as OfficeModel
from schemas.auth_schema import User as UserSchema
from db.client2 import msdb_manager
from datalayer.member import Member
from datalayer.setting import Setting
from utils.security import create_access_token
import requests


def create_nurse_template() -> str:
    """간호사 정보 엑셀 템플릿 생성"""
    
    # 템플릿 데이터 구조
    template_data = {
        '병동명': ['ICU', 'ICU', '응급실', '(입력 가이드)', ''],
        '식별코드': ['UUID자동생성', 'UUID자동생성', 'UUID자동생성', '(UUID는 자동생성됨)', ''],
        '계정 ID': ['nurse001', 'nurse002', 'nurse003', '(영문숫자조합)', ''],
        '이름': ['김간호', '이수간', '박일반', '(한글이름)', ''],
        '경력': [5, 10, 3, '(1이상정수)', ''],
        '직군': ['간호사', '간호사', '간호사', '(간호사)', ''],
        '직책': ['주임', '수간호사', '일반', '(주임/수간호사/일반)', ''],
        '수간호사여부': ['N', 'Y', 'N', '(Y/N)', '']
    }
    
    # DataFrame 생성
    df = pd.DataFrame(template_data)
    
    # 임시 파일 생성
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        template_path = tmp_file.name
    
    # 엑셀 파일로 저장
    with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='간호사정보', index=False)
        
        # 워크시트 스타일링
        worksheet = writer.sheets['간호사정보']
        
        # 헤더 스타일 적용
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(fgColor="CCCCCC")
        
        # 가이드 행 스타일 적용 (4번째 행)
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=4, column=col)
            cell.font = cell.font.copy(italic=True, color="666666")
    
    return template_path
def create_nurse_template2() -> str:
    """엑셀 템플릿2: 계정ID/이름 두 컬럼만 포함."""
    template_data = {
        '사번(필수)': ['1001', '1002', '1003', '(영문숫자조합)'],
        '계정 ID(필수)': ['nurse001', 'nurse002', 'nurse003', '(영문숫자조합)'],
        '직원명(필수)': ['김수간', '이간호', '최간호', '(한글이름)'],
        '직무(필수)': ['HN', 'AN', 'RN', ('직무코드')],
        '경력(필수)': [25, 15, 1, ('경력년수')],
        '수간호사여부(필수)': ['Y', 'N', 'N', ('수간호사여부 정보')],
        '입사일(선택)': ['2025-01-03', '', '', ('입사일 정보')],
        # '적용해제일(선택)': ['2025-01-25', '', '', ('근무 표 적용해제일 정보')],
        '생년월일(필수)': ['1999-01-01', '', '', ('생년월일 정보')],
        '연락처(필수)': ['010-0000-0000', '', '', ('연락처 정보')],
        '성별(필수)': ['남', '', '', ('성별 정보')],
        '이메일(선택)': ['nurse001@hospital.com', '', '', ('이메일 정보')],

    }
    df = pd.DataFrame(template_data)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        template_path = tmp_file.name
    with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='업로드2', index=False)
        ws = writer.sheets['업로드2']
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
    return template_path



def get_or_create_group(group_name: str, user: UserSchema, db: Session) -> Tuple[Optional[str], bool, List[str]]:
    """
    병동명을 기반으로 group_id를 찾거나 새로 생성
    Returns: (group_id, is_new_group, warnings)
    """
    warnings = []
    
    # 현재 사용자의 office_id로 해당 office의 모든 그룹 조회
    existing_group = db.query(GroupModel).filter(
        GroupModel.office_id == user.office_id,
        GroupModel.group_name == group_name
    ).first()
    
    if existing_group:
        warnings.append(f"기존 '{group_name}' 그룹에 추가됩니다.")
        return existing_group.group_id, False, warnings
    
    # 기존에 없는 그룹명인 경우 새 그룹 ID 생성
    new_group_id = generate_new_group_id(user.office_id, db)
    warnings.append(f"새로운 그룹 '{group_name}'이 생성됩니다.")
    return new_group_id, True, warnings


def generate_new_group_id(office_id: str, db: Session) -> str:
    """office_id + 001, 002, 003... 형태로 새 그룹 ID 생성"""
    
    # 해당 office의 기존 그룹들 중 최대 번호 찾기
    existing_groups = db.query(GroupModel).filter(
        GroupModel.office_id == office_id,
        GroupModel.group_id.like(f"{office_id}%")
    ).all()
    
    max_number = 0
    for group in existing_groups:
        try:
            # office_id 뒤의 숫자 부분 추출
            number_part = group.group_id.replace(office_id, "")
            if number_part.isdigit():
                max_number = max(max_number, int(number_part))
        except:
            continue
    
    # 다음 번호로 새 그룹 ID 생성
    new_number = max_number + 1
    return f"{office_id}{new_number:03d}"


def create_new_group(group_name: str, group_id: str, user: UserSchema, db: Session) -> str:
    """새로운 그룹 생성"""
    
    new_group = GroupModel(
        group_id=group_id,
        office_id=user.office_id,
        group_name=group_name
    )
    
    db.add(new_group)
    db.flush()  # DB에 즉시 반영하되 커밋은 나중에
    
    return group_id


def get_next_sequence(group_id: str, active_status: int, db: Session, role: str = "RN") -> int:
    """해당 그룹의 특정 active 상태 + role 그룹에서 다음 sequence 번호 반환"""
    from services.nurse_service import get_role_group, _role_group_filter

    role_group = get_role_group(role)
    max_sequence = db.query(func.max(NurseModel.sequence)).filter(
        NurseModel.group_id == group_id,
        NurseModel.active == active_status,
        _role_group_filter(role_group)
    ).scalar()

    return (max_sequence or 0) + 1


# def process_excel_upload(file_path: str, user: UserSchema, db: Session) -> Dict[str, Any]:
#     """엑셀 파일 업로드 처리 및 검증"""
    
#     try:
#         # 엑셀 파일 읽기
#         df = pd.read_excel(file_path, sheet_name=0)
        
#         # 빈 행 및 가이드 행 제거
#         df = df.dropna(how='all')  # 모든 컬럼이 비어있는 행 제거
#         df = df[~df.iloc[:, 0].astype(str).str.contains('입력 가이드', na=False)]  # 가이드 행 제거
        
#         # 최대 행 수 검증
#         if len(df) > 1000:
#             raise ValueError("최대 1000행까지만 업로드 가능합니다.")
#         # 컬럼 매핑
#         column_mapping = {
#             '병동명': 'group_name',
#             '식별코드': 'nurse_id', 
#             '계정 ID': 'account_id',
#             '이름': 'name',
#             '경력': 'experience',
#             '직군': 'role',
#             '직책': 'level_',
#             '수간호사여부': 'is_head_nurse',
#             '생년월일': 'birth_date',
#             '연락처': 'phone_number'
#         }
#         # 컬럼명 유연 매핑 (유사한 이름 인식)
#         flexible_mapping = {}
#         for excel_col in df.columns:
#             excel_col_clean = str(excel_col).strip()
#             for standard_col, db_field in column_mapping.items():
#                 if (excel_col_clean == standard_col or 
#                     excel_col_clean in ['병동', '부서'] and standard_col == '병동명' or
#                     excel_col_clean in ['ID', '아이디'] and standard_col == '계정 ID' or
#                     excel_col_clean in ['성명', '간호사명'] and standard_col == '이름' or
#                     excel_col_clean in ['년차', '경력년수'] and standard_col == '경력' or
#                     excel_col_clean in ['수간호사', '헤드너스'] and standard_col == '수간호사여부' or
#                     excel_col_clean in ['출생일', 'Birthday', '생일'] and standard_col == '생년월일' or
#                     excel_col_clean in ['전화번호', 'Phone', '휴대폰'] and standard_col == '연락처' or
#                     excel_col_clean in ['성별', 'Gender', '남/여'] and standard_col == '성별'):
#                     flexible_mapping[excel_col] = db_field
#                     break
#         # 필수 컬럼 확인
#         required_fields = ['group_name', 'account_id', 'name', 'experience', 'role', 'level_', 'is_head_nurse', 'birth_date', 'phone_number']
#         missing_fields = [field for field in required_fields if field not in flexible_mapping.values()]
#         if missing_fields:
#             missing_korean = []
#             field_korean_map = {
#                 'group_name': '병동명',
#                 'account_id': '계정 ID', 
#                 'name': '이름',
#                 'experience': '경력',
#                 'role': '직군',
#                 'level_': '직책',
#                 'is_head_nurse': '수간호사여부',
#                 'birth_date': '셍냔월일',
#                 'phone_number': '연락처'
#             }
#             for field in missing_fields:
#                 missing_korean.append(field_korean_map.get(field, field))
#             raise ValueError(f"필수 컬럼이 누락되었습니다: {', '.join(missing_korean)}")
            
#         # 병동명별 그룹 정보 수집
#         # 동일한 병동명이 여러 병원(office)에서 존재할 수 있으므로, 엑셀에 오피스/지점 컬럼이 존재하는 경우
#         # 먼저 현재 사용자 office_id와 일치하는 행으로 필터링한다.
#         office_col = None
#         for c in ['office_id', 'Office ID', '오피스ID', '병원ID', '병원코드', '기관ID', '지점ID']:
#             if c in df.columns:
#                 office_col = c
#                 break
#         if office_col:
#             df_filtered = df[df[office_col].astype(str).str.strip() == str(user.office_id)]
#         else:
#             df_filtered = df

#         unique_groups = df_filtered[get_excel_column_by_field('group_name', flexible_mapping)].dropna().unique()
#         group_info = {}
#         new_groups_needed = []
        
#         for group_name in unique_groups:
#             group_name = str(group_name).strip()
#             if not group_name:
#                 continue
                
#             group_id, is_new, warnings = get_or_create_group(group_name, user, db)
#             group_info[group_name] = {
#                 'group_id': group_id,
#                 'is_new': is_new,
#                 'warnings': warnings
#             }
            
#             if is_new:
#                 new_groups_needed.append(group_name)
        
#         # 그룹별 sequence 카운터 초기화 (활성 상태 기준)
#         group_sequence_counters = {}
#         for group_name, info in group_info.items():
#             group_id = info['group_id']
#             if info['is_new']:
#                 # 새 그룹인 경우 1부터 시작
#                 group_sequence_counters[group_id] = 1
#             else:
#                 # 기존 그룹인 경우 활성 상태(active=1)의 다음 sequence 가져오기
#                 group_sequence_counters[group_id] = get_next_sequence(group_id, 1, db)
        
#         # 데이터 변환
#         processed_data = []
#         validation_results = []
        
#         for idx, row in df.iterrows():
#             try:
#                 # 병동명 처리
#                 group_name = str(row[get_excel_column_by_field('group_name', flexible_mapping)]).strip()
#                 group_data = group_info.get(group_name, {})
#                 group_id = group_data.get('group_id')
                
#                 # sequence 할당
#                 sequence = group_sequence_counters.get(group_id, 0)
#                 group_sequence_counters[group_id] = sequence + 1
                
#                 # 기본 데이터 변환 (nurses.office_id 함께 저장)
#                 nurse_data = {
#                     # 'group_name': group_name,
#                     'group_id': group_id,
#                     'office_id': user.office_id,
#                     'nurse_id': str(uuid.uuid4()) if pd.isna(row.get('식별코드')) or str(row.get('식별코드')).strip() == 'UUID자동생성' else str(row.get('식별코드')),
#                     'account_id': str(row[get_excel_column_by_field('account_id', flexible_mapping)]).strip(),
#                     'name': str(row[get_excel_column_by_field('name', flexible_mapping)]).strip(),
#                     'experience': int(float(row[get_excel_column_by_field('experience', flexible_mapping)])),
#                     'role': str(row[get_excel_column_by_field('role', flexible_mapping)]).strip(),
#                     'level_': str(row[get_excel_column_by_field('level_', flexible_mapping)]).strip(),
#                     'is_head_nurse': parse_boolean(row[get_excel_column_by_field('is_head_nurse', flexible_mapping)]),
#                     # 'is_night_nurse': False,  # 기본값
#                     'is_night_nurse': [],  # 기본값
#                     'personal_off_adjustment': 0,  # 기본값
#                     'preceptor_id': None,  # 기본값
#                     'joining_date': None,  # 기본값
#                     'resignation_date': None,  # 기본값
#                     'sequence': sequence,
#                     'active': 1,  # 엑셀 업로드는 기본적으로 활성 상태
#                     'birth_date': str(row[get_excel_column_by_field('birth_date', flexible_mapping)]).strip() if get_excel_column_by_field('birth_date', flexible_mapping) in row and pd.notna(row[get_excel_column_by_field('birth_date', flexible_mapping)]) else None,  # 신규 컬럼: 생년월일
#                     'phone_number': str(row[get_excel_column_by_field('phone_number', flexible_mapping)]).strip() if get_excel_column_by_field('phone_number', flexible_mapping) in row and pd.notna(row[get_excel_column_by_field('phone_number', flexible_mapping)]) else None  # 신규 컬럼: 연락처
#                 }
                
#                 # 개별 행 검증
#                 row_validation = validate_single_row(group_name, nurse_data, user, db)
#                 print('row_validation!', row_validation)
                
#                 # 그룹 상태 정보 추가
#                 row_validation['warnings'].extend(group_data.get('warnings', []))
#                 row_validation['is_new_group'] = group_data.get('is_new', False)
#                 row_validation['group_name'] = group_name
                
#                 processed_data.append(nurse_data)
#                 validation_results.append(row_validation)
                
#             except Exception as e:
#                 # 행별 오류 처리
#                 error_data = {
#                     'row_index': idx + 2,  # 엑셀 행 번호 (헤더 포함)
#                     'error': str(e),
#                     'is_valid': False,
#                     'errors': [str(e)],
#                     'warnings': [],
#                     'is_new_group': False
#                 }
#                 validation_results.append(error_data)
#                 processed_data.append(None)
        
#         # 전체 검증 결과 요약
#         valid_count = sum(1 for result in validation_results if result.get('is_valid', False))
#         error_count = len(validation_results) - valid_count
#         overwrite_count = sum(1 for result in validation_results if result.get('is_overwrite', False))
#         new_group_count = len(new_groups_needed)
        
#         return {
#             'success': True,
#             'data': processed_data,
#             'validation_results': validation_results,
#             'new_groups_needed': new_groups_needed,
#             'summary': {
#                 'total': len(validation_results),
#                 'valid': valid_count,
#                 'error': error_count,
#                 'overwrite': overwrite_count,
#                 'new_groups': new_group_count
#             }
#         }
        
#     except Exception as e:
#         return {
#             'success': False,
#             'error': str(e),
#             'data': [],
#             'validation_results': [],
#             'new_groups_needed': [],
#             'summary': {'total': 0, 'valid': 0, 'error': 0, 'overwrite': 0, 'new_groups': 0}
#         }


def upload2_validate(file_path: str, user: UserSchema, db: Session, group_id: str) -> Dict[str, Any]:
    """업로드2: 파일을 검증만 수행하고, 정규화된 행과 오류를 반환한다.

    - 행별 오류: 포맷/타입/허용 계정/필수값 등
    - 전역 오류: 수간호사 최소 1명, DB 중복 계정 등
    """
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        df = df.dropna(how='all')
        if len(df) > 2000:
            raise ValueError("최대 2000행까지만 업로드 가능합니다.")

        def find_col(candidates: list[str]) -> str:
            for c in df.columns:
                cc = str(c).strip()
                if cc in candidates:
                    return c
            raise ValueError(f"필수 컬럼 누락: {candidates}")
        def find_col_optional(candidates: list[str]) -> Optional[str]:
            for c in df.columns:
                cc = str(c).strip()
                if cc in candidates:
                    return c
            return None

        col_empnum = find_col(['사번(필수)','사번','EmpNum','emp_num'])
        col_acc = find_col(['계정 ID(필수)','계정 ID','ID','아이디','account_id'])
        col_name = find_col(['직원명(필수)','이름','성명','name'])
        col_role = find_col(['직무(필수)','직무','role'])
        col_exp = find_col(['경력(필수)','경력','experience'])
        col_head = find_col(['수간호사여부(필수)','수간호사여부','is_head_nurse'])
        col_join = find_col_optional(['입사일(선택)','입사일','joining_date'])
        # col_resi = find_col_optional(['적용해제일(선택)','적용해제일','퇴사일','resignation_date'])
        col_birth = find_col(['생년월일(필수)', '생년월일', 'birth_date'])
        col_phone = find_col(['연락처(필수)', '연락처', 'phone_number'])
        col_gender = find_col(['성별(필수)', '성별', 'gender'])
        col_email = find_col_optional(['이메일(선택)', '이메일', 'email'])
        
        office_id = user.office_id
        rows_allowed = msdb_manager.fetch_all(Member.member_accounts_by_office(), params=(str(office_id),))
        allowed: dict[str, tuple[str, str | None]] = {}
        for r in rows_allowed or []:
            acc = str(r.get('account_id', '')).strip()
            nm = str(r.get('name', '')).strip()
            auth = r.get('EmpAuthGbn')
            nurse_id = str(r.get('nurse_id', uuid.uuid4())).strip()
            if acc:
                allowed[acc] = (nm, auth, nurse_id)
        normalized: list[dict] = []
        errors: list[dict] = []
        head_count = 0
        acc_in_file: set[str] = set()

        def parse_dt(v):
            try:
                if pd.isna(v) or str(v).strip() == '':
                    return None
                return pd.to_datetime(v, errors='coerce').to_pydatetime()
            except Exception:
                return None

        for i, row in df.iterrows():
            ridx = int(i) + 2
            row_errs: list[str] = []
            emp_num_val = row.get(col_empnum)
            emp_num = '' if pd.isna(emp_num_val) else str(emp_num_val).strip()
            account_id = str(row.get(col_acc, '')).strip()
            name = str(row.get(col_name, '')).strip()
            role_val = row.get(col_role)
            role = 'RN' if pd.isna(role_val) or not str(role_val).strip() else str(role_val).strip()
            birth_date = row.get(col_birth)
            phone_num = row.get(col_phone)
            gender = row.get(col_gender)
            email_val = str(row.get(col_email, '')).strip() if col_email else None
            email_val = email_val if email_val else None
            # experience: 비어있으면 None 허용, 값이 있으면 숫자만 허용
            exp_val = None
            raw_exp_val = row.get(col_exp, 1)
            # print('raw_exp_val', raw_exp_val)
            import math
            # NaN/None/빈문자/문자 'nan' 전부 필터링
            if raw_exp_val not in ['', None] and not (isinstance(raw_exp_val, float) and math.isnan(raw_exp_val)) and str(raw_exp_val).lower() != 'nan':
                try:
                    exp_val = int(float(str(raw_exp_val).strip()))
                except Exception:
                    # print('raw_exp_val', raw_exp_val)
                    row_errs.append("경력은 숫자여야 합니다. 예: 1, 3, 10")
            head_raw = str(row.get(col_head, '')).strip().upper()
            is_head = True if head_raw in ['Y','YES','1','TRUE','T'] else False
            if is_head:
                head_count += 1
            joining_raw = row.get(col_join) if col_join else None
            # resignation_raw = row.get(col_resi) if col_resi else None
            # 날짜: None 허용, 숫자 허용, 문자열인 경우에만 YYYY-MM-DD 형식 검증
            joining_dt = parse_dt(joining_raw) if col_join else None
            if isinstance(joining_raw, str) and joining_raw.strip() and not re.match(r'^\d{4}-\d{2}-\d{2}$', joining_raw.strip()):
                row_errs.append("입사일 형식이 올바르지 않습니다. 예: 2025-10-10")
            # resignation_dt = parse_dt(resignation_raw) if col_resi else None
            # if isinstance(resignation_raw, str) and resignation_raw.strip() and not re.match(r'^\d{4}-\d{2}-\d{2}$', resignation_raw.strip()):
            #     row_errs.append("적용해제일 형식이 올바르지 않습니다. 예: 2025-10-10")
            if not account_id:
                row_errs.append("계정 ID 누락")
            elif not re.match(r'^\S{1,50}$', account_id):
                row_errs.append("계정 ID 형식 오류: 공백 제외 최대 50자까지 허용됩니다.")
            # elif account_id in acc_in_file:
                # existing = db.query(NurseModel.account_id).filter(NurseModel.account_id.in_(list(acc_in_file))).all()
                # if existing:
                #     row_errs.append('이미 존재하는 계정 ID')
            elif account_id in acc_in_file:
                row_errs.append("엑셀 내 중복 계정 ID")
            elif account_id not in allowed:
                row_errs.append(f"계정이 등록되지 않았습니다.")
            if not name:
                # 원장에서 이름 보강
                name = (allowed.get(account_id, ("", None))[0] if account_id in allowed else "")
                if not name:
                    row_errs.append("직원명 누락")
            if not role:
                row_errs.append("직무 누락")
            else:
                acc_in_file.add(account_id)
            
            def is_invalid_value(v):
                """이름 등 필드가 비어있거나 nan인지 검사. float nan / pd.NA / 문자열 'nan' 포함."""
                if v is None:
                    return True
                if isinstance(v, float) and math.isnan(v):
                    return True
                try:
                    if pd.isna(v):
                        return True
                except Exception:
                    pass
                if isinstance(v, str):
                    return v.strip().lower() in ("", "nan", "none", "null")
                return False
            
            if pd.isna(emp_num_val):
                emp_num = '-'
            elif isinstance(emp_num_val, float) and emp_num_val.is_integer():
                emp_num = str(int(emp_num_val))
            else:
                emp_num = str(emp_num_val).strip()
            if is_invalid_value(name):
                row_errs.append("등록된 이름이 없습니다.")

            if account_id in allowed:
                normalized.append({
                    'row': ridx,
                    'emp_num': emp_num or None,
                    'account_id': account_id,
                    'name': name,
                    'role': role,
                    'experience': exp_val,
                    'is_head_nurse': is_head,
                    'joining_date': joining_dt.isoformat() if joining_dt else None,
                    # 'resignation_date': resignation_dt.isoformat() if resignation_dt else None,
                    'nurse_id': allowed[account_id][2],
                    'birth_date': birth_date,
                    'phone_number': phone_num,
                    'is_night_nurse': [],
                    'gender': gender,
                    'email': email_val,
                })
        
            if row_errs:
                errors.append({'row': ridx, 'reason': ' | '.join(row_errs)})

        existing_head_nurses = db.query(NurseModel).filter(
            NurseModel.office_id == user.office_id,
            NurseModel.group_id == group_id,
            NurseModel.is_head_nurse == 1
        ).count()
        # 전역 검증: 수간호사 최소 1명
        # if head_count == 0:
        if head_count == 0 and existing_head_nurses == 0:
            errors.append({'row': 0, 'reason': '수간호사는 최소 1명 이상이어야 합니다.'})
        # 전역 검증: DB 중복 계정

        return {
            'success': 0 if errors else len(normalized),
            'errors': errors,
            'rows': normalized,
            'summary': {
                'total': len(normalized),
                'head_nurses': head_count,
                'error_count': len(errors),
            }
        }
    except Exception as e:
        print('error', e)
        return {"success": 0, "errors": [{"row": 0, "reason": str(e)}], "rows": [], 'summary': {'total': 0, 'head_nurses': 0, 'error_count': 1}}


#

# def upload2_confirm(rows: List[Dict[str, Any]], user: UserSchema, db: Session, target_group_id: str) -> Dict[str, Any]:
#     """업로드2: 검증된 행을 저장한다. 오류 포함 행은 건너뜀."""
    

#     try:
#         if not target_group_id:
#             print("[ERROR] target_group_id가 없습니다!")
#             return {"success": 0, "errors": [{"row": 0, "reason": "group_id가 필요합니다."}]}

#         saved = 0
#         updated = 0
#         errors = [] # 추가

#         for idx, item in enumerate(rows, 1):
#             print(f"\n[행 {idx:2d}] 처리 시작 ───────────────────────────────────────")
            
#             account_id = item.get('account_id')
#             name = item.get('name')
#             role = item.get('role', 'RN')
#             exp_val = item.get('experience', 1)
#             nurse_id_raw = item.get('nurse_id')
#             nurse_id = str(nurse_id_raw).strip() if nurse_id_raw else None
#             is_head = bool(item.get('is_head_nurse', False))
#             emp_num = item.get('emp_num', '')
#             jd = item.get('joining_date')
#             birth_dt = item.get('birth_date')
#             phone_number = item.get('phone_number')
#             gender = item.get('gender')
#             is_night_nurse = item.get('is_night_nurse', [])
#             work_shifts_val = item.get('work_shifts', [])

#             print(f"   • account_id     : {account_id}")
#             print(f"   • nurse_id (원본) : {nurse_id_raw}")
#             print(f"   • nurse_id (처리후): {nurse_id}")
#             print(f"   • name           : {name}")
#             print(f"   • is_head_nurse  : {is_head}")
#             print(f"   • experience     : {exp_val}")
#             print(f"   • joining_date   : {jd}")
#             print(f"   • birth_date     : {birth_dt}")
#             print(f"   • phone_number   : {phone_number}")
#             print(f"   • gender         : {gender}")            # nurse_id 안전 처리
#             if not nurse_id or len(nurse_id) < 8:
#                 nurse_id = str(uuid.uuid4())
#                 print(f"   → nurse_id 이상 → 새 UUID 생성: {nurse_id}")

#             existing = db.query(NurseModel).filter(NurseModel.account_id == account_id).first()

#             if existing:
#                 print(f"   → 기존 레코드 발견! nurse_id={existing.nurse_id}, name={existing.name}")
#                 print(f"      현재 DB 값 - experience={existing.experience}, is_head_nurse={existing.is_head_nurse}")

#                 if name and existing.name != name:
#                     existing.name = name
#                     print("      → 이름 업데이트")

#                 existing.emp_num = emp_num if emp_num is not None else ''
#                 existing.role = role if role is not None else ''
                
#                 if isinstance(exp_val, int):
#                     existing.experience = exp_val
#                     print(f"      → experience 업데이트: {exp_val}")

#                 existing.is_head_nurse = is_head

#                 try:
#                     existing.office_id = user.office_id
#                     print("      → office_id 강제 업데이트")
#                 except Exception:
#                     print("      → office_id 업데이트 실패 (무시)")

#                 if jd:
#                     try:
#                         joining_dt = pd.to_datetime(jd).to_pydatetime()
#                         existing.joining_date = joining_dt
#                         print(f"      → joining_date 업데이트: {joining_dt}")
#                     except:
#                         print("      → joining_date 변환 실패")

#                 existing.birth_date = birth_dt
#                 existing.phone_number = phone_number
#                 existing.gender = gender
#                 existing.is_night_nurse = is_night_nurse
#                 existing.work_shifts = work_shifts_val
                
#                 updated += 1
#                 print(f"   → 업데이트 완료 (현재 updated 누적: {updated})")
#                 continue

#             print("   → 신규 등록 시작")
#             try:
#                 seq_next = get_next_sequence(target_group_id, 1, db)
#                 print(f"      • 다음 sequence 값: {seq_next}")

#                 new_nurse = NurseModel(
#                     nurse_id=nurse_id,
#                     group_id=target_group_id,
#                     office_id=user.office_id,
#                     emp_num=emp_num if emp_num is not None else '',
#                     account_id=account_id,
#                     name=name or account_id,
#                     experience=exp_val if isinstance(exp_val, int) else 1,
#                     role=role if role is not None else '',
#                     level_='일반',
#                     is_head_nurse=is_head,
#                     is_night_nurse=is_night_nurse,
#                     personal_off_adjustment=0,
#                     preceptor_id=None,
#                     joining_date=pd.to_datetime(jd).to_pydatetime() if jd else None,
#                     sequence=seq_next,
#                     active=1,
#                     birth_date=birth_dt,
#                     phone_number=phone_number,
#                     gender=gender,
#                     work_shifts=work_shifts_val,
#                 )

#                 print("      • new_nurse 객체 생성 완료")
#                 import pprint
#                 print("      • new_nurse 내용:")
#                 pprint.pprint(new_nurse.__dict__)

#                 db.add(new_nurse)
#                 print(f"      • db.add() 완료 (nurse_id={nurse_id})")
#                 saved += 1
#                 print(f"   → 신규 등록 완료 (현재 saved 누적: {saved})")

#             except Exception as inner_e:
#                 print(f"   ★★★ 신규 등록 중 오류 ★★★")
#                 print(f"      • 에러 타입: {type(inner_e).__name__}")
#                 print(f"      • 에러 메시지: {str(inner_e)}")
#                 import traceback
#                 traceback.print_exc()
#                 errors.append({"row": item.get('row', 0), "reason": str(inner_e)})

#         print("\n[upload2_confirm] COMMIT 직전 상태")
#         print(f"  • 최종 saved = {saved}")
#         print(f"  • 최종 updated = {updated}")
#         print(f"  • 누적 에러 개수 = {len(errors)}")

#         db.commit()
#         print("[upload2_confirm] ★★★ COMMIT 성공 ★★★")
#         print("====================================================")

#         return {"success": saved + updated, "saved": saved, "updated": updated, "errors": errors}

#     except Exception as e:
#         print("[upload2_confirm] ★★★ 전체 예외 발생 ★★★")
#         print(f"  • 에러 타입: {type(e).__name__}")
#         print(f"  • 에러 메시지: {str(e)}")
#         import traceback
#         traceback.print_exc()
#         db.rollback()
#         print("[upload2_confirm] ROLLBACK 완료")
#         print("====================================================")
        
#         return {"success": 0, "errors": [{"row": 0, "reason": f"저장 실패: {str(e)}"}]}


def get_excel_column_by_field(field: str, mapping: Dict[str, str]) -> str:
    """필드명으로 엑셀 컬럼명 찾기"""
    for excel_col, db_field in mapping.items():
        if db_field == field:
            return excel_col
    raise ValueError(f"필드 {field}에 해당하는 엑셀 컬럼을 찾을 수 없습니다.")


def parse_boolean(value: Any) -> bool:
    """다양한 형태의 불린 값 파싱"""
    if pd.isna(value):
        return False
    
    str_value = str(value).strip().upper()
    return str_value in ['Y', 'YES', 'TRUE', '1', 'T', '참', '예']


def validate_single_row(group_name: str, nurse_data: Dict[str, Any], user: UserSchema, db: Session) -> Dict[str, Any]:
    """개별 행 데이터 검증"""
    
    errors = []
    warnings = []
    is_overwrite = False
    
    # # 병동명 검증
    # if not nurse_data.get('group_name'):
    #     errors.append("병동명은 필수입니다.")
    # elif not nurse_data.get('group_id'):
    #     errors.append("유효하지 않은 병동명입니다.")
    
    # 필수 필드 검증
    if not nurse_data.get('account_id'):
        errors.append("계정 ID는 필수입니다.")
    elif not re.match(r'^\S{1,50}$', nurse_data['account_id']):
        errors.append("계정 ID 형식 오류: 공백 제외 최대 50자까지 허용됩니다.")
    
    if not nurse_data.get('name'):
        errors.append("이름은 필수입니다.")
    
    if nurse_data.get('experience', 0) < 1:
        errors.append("경력은 1년 이상이어야 합니다.")
    
    if not nurse_data.get('role'):
        errors.append("직군은 필수입니다.")
    
    if not nurse_data.get('level_'):
        errors.append("직책은 필수입니다.")
    
    # 중복 검사
    if nurse_data.get('group_id'):
        existing_nurse = db.query(NurseModel).filter(
            NurseModel.nurse_id == nurse_data['nurse_id']
        ).first()
        
        if existing_nurse:
            is_overwrite = True
            warnings.append(f"기존 간호사 '{existing_nurse.name}' 데이터를 덮어씁니다.")
        
        # 계정 ID 중복 검사 (다른 nurse_id와)
        existing_account = db.query(NurseModel).filter(
            NurseModel.account_id == nurse_data['account_id'],
            NurseModel.nurse_id != nurse_data['nurse_id']
        ).first()
        
        if existing_account:
            errors.append(f"계정 ID '{nurse_data['account_id']}'는 이미 사용 중입니다.")
    
    return {
        'is_valid': len(errors) == 0,
        'is_overwrite': is_overwrite,
        'errors': errors,
        'warnings': warnings,
        'nurse_data': nurse_data,
        'group_name': group_name
    }


def validate_excel_data(data: List[Dict[str, Any]], user: UserSchema, db: Session) -> Dict[str, Any]:
    """엑셀 데이터 전체 유효성 검증"""
    
    validation_results = []
    
    for nurse_data in data:
        if nurse_data is None:
            validation_results.append({'is_valid': False, 'errors': ['데이터 파싱 오류']})
            continue
            
        result = validate_single_row(nurse_data, user, db)
        validation_results.append(result)

    
    # 수간호사 최소 1명 검증
    head_nurses = [
        result for result in validation_results 
        if result.get('is_valid') and result.get('nurse_data', {}).get('is_head_nurse')
    ]
    
    if len(head_nurses) == 0:
        # 기존 수간호사가 있는지 확인
        existing_head_nurses = db.query(NurseModel).filter(
            NurseModel.group_id == user.group_id,
            NurseModel.is_head_nurse == True
        ).count()
        
        if existing_head_nurses == 0:
            return {
                'success': False,
                'error': '최소 1명의 수간호사가 필요합니다.',
                'validation_results': validation_results
            }
    
    valid_count = sum(1 for result in validation_results if result.get('is_valid', False))
    error_count = len(validation_results) - valid_count
    overwrite_count = sum(1 for result in validation_results if result.get('is_overwrite', False))
    print('error_count', error_count)
    return {
        'success': True,
        'validation_results': validation_results,
        'summary': {
            'total': len(validation_results),
            'valid': valid_count,
            'error': error_count,
            'overwrite': overwrite_count
        }
    }


def save_excel_data(data: List[Dict[str, Any]], user: UserSchema, db: Session) -> Dict[str, Any]:
    """검증된 데이터 DB 저장"""
    
    try:
        saved_count = 0
        updated_count = 0
        for nurse_data in data:
            if not nurse_data:
                continue
                
            # 기존 데이터 확인
            existing_nurse = db.query(NurseModel).filter(
                NurseModel.nurse_id == nurse_data['nurse_id']
            ).first()
            
            if existing_nurse:
                # 업데이트
                for key, value in nurse_data.items():
                    if hasattr(existing_nurse, key):
                        setattr(existing_nurse, key, value)
                existing_nurse.updated_at = datetime.now()
                updated_count += 1
            else:
                # 신규 생성
                new_nurse = NurseModel(**nurse_data)
                db.add(new_nurse)
                saved_count += 1
        
        db.commit()
        
        return {
            'success': True,
            'message': f'저장 완료: 신규 {saved_count}건, 업데이트 {updated_count}건',
            'saved_count': saved_count,
            'updated_count': updated_count
        }
        
    except Exception as e:
        db.rollback()
        return {
            'success': False,
            'error': f'저장 실패: {str(e)}'
        } 


def create_groups_and_save_data(data: List[Dict[str, Any]], new_groups_to_create: List[str], user: UserSchema, db: Session) -> Dict[str, Any]:
    """새 그룹 생성 후 데이터 저장"""
    
    try:
        # 새 그룹 생성
        created_groups = {}
        for group_name in new_groups_to_create:
            # 새 그룹 ID 생성
            new_group_id = generate_new_group_id(user.office_id, db)
            # print('new_group_id', new_group_id)
            # 그룹 생성
            create_new_group(group_name, new_group_id, user, db)
            created_groups[group_name] = new_group_id

        # 데이터의 group_id 업데이트 (이미 생성된 그룹 ID 사용)
        for nurse_data in data:
            if nurse_data:
                group_name = nurse_data.get('group_name')
                if group_name in created_groups:
                    nurse_data['group_id'] = created_groups[group_name]
        
        # 데이터 저장
        result = save_excel_data(data, user, db)
        if result['success']:
            result['created_groups'] = created_groups
            result['message'] = f"{result['message']} (새 병동 {len(created_groups)}개 생성)"
        
        return result
        
    except Exception as e:
        db.rollback()
        return {
            'success': False,
            'error': f'그룹 생성 및 저장 실패: {str(e)}'
        } 


def _merge_team_cell(ws, col, first_row, last_row, label, center, border_all, gray_fill):
    """팀별보기 전용: [first_row, last_row] 의 팀 컬럼(col)을 세로 병합하고 팀명 1개를
    가운데 정렬로 기입한다. 병합 범위 전 행에 테두리·배경을 둔다."""
    if last_row > first_row:
        ws.merge_cells(start_row=first_row, start_column=col, end_row=last_row, end_column=col)
    cell = ws.cell(row=first_row, column=col, value=label)
    cell.alignment = center
    cell.fill = gray_fill
    cell.border = border_all
    for r in range(first_row, last_row + 1):
        cc = ws.cell(row=r, column=col)
        cc.border = border_all
        cc.fill = gray_fill


def _sort_team_ids_by_name(team_name_map: dict) -> list:
    """팀 id 들을 근무표 만들기 화면(team-sort-util.sortTeamGroupsByTeamName)과 동일
    순서로 정렬한다. 1순위 그룹: 한글(0) < 영문(1) < 숫자(2) < 기타(3). 2순위: 자연
    정렬(숫자 런은 int 로 비교)+소문자. 3순위: team_id.
    (export 엔 활성 실팀만 오므로 미배정/임시팀 분기는 불필요.)
    """
    import re as _re

    def _group_rank(name: str) -> int:
        first = (name or "").strip()[:1]
        if not first:
            return 3
        if "가" <= first <= "힣":
            return 0
        if first.isascii() and first.isalpha():
            return 1
        if first.isdigit():
            return 2
        return 3

    def _natural_key(name: str) -> list:
        # Intl.Collator({numeric:true}) 근사: 숫자 런은 (0,int), 그 외는 (1,소문자).
        # 토큰 타입을 앞에 둬 int/str 직접 비교(TypeError)를 원천 차단.
        out = []
        for tok in _re.findall(r"\d+|\D+", name or ""):
            out.append((0, int(tok)) if tok.isdigit() else (1, tok.casefold()))
        return out

    def _key(tid):
        name = team_name_map.get(tid) or ""
        return (_group_rank(name), _natural_key(name), tid)

    return sorted(team_name_map.keys(), key=_key)


def export_schedule_excel_bytes(
    schedule_id: str, current_user, db, target_group_id: str, group_by_team: bool = False
) -> tuple[bytes, bool]:
    """지정된 schedule_id의 근무표를 엑셀(xlsx) 바이트로 생성하여 (bytes, team_view) 로 반환.

    group_by_team=True 이고 실제 팀 배정이 있으면 '팀별보기' 레이아웃으로 그린다:
    기존 레이아웃과 동일하되 맨 앞(컬럼1)에 팀 세로병합 컬럼 1개를 추가한다.

    팀 구분은 **근무표 만들기 화면의 '팀별보기'(filterTeam)와 완전 동일한 규칙**으로 한다:
      teamId = is_night_dedicated ? None : (as_of_team ?? nurses.team_id 캐시)
    - 소속·N전담·시점팀(as_of_team)은 화면과 같은 group_members_in_month 를 재사용한다.
      → N전담은 팀 로테이션 비참여 → '미등록'.
    - 활성 팀(Team.active==1, /teams 와 동일 소스)에 없는 team_id 는 '미등록'으로 둔다.
    - 팀 블록 순서는 화면(sortTeamGroupsByTeamName)과 동일한 팀명 기준 + 미등록 맨 끝.
    같은 팀 간호사 행들의 컬럼1을 세로 병합해 팀명 1개만 표시한다(팀 내부는 기존 순서 유지).

    반환 team_view 는 실제로 팀별보기 레이아웃이 적용됐는지 여부(파일명/헤더 판정용).
    팀 배정이 전무(전원 미등록)하거나 group_by_team=False 이면 기존 레이아웃과
    완전히 동일하게 그리고 team_view=False 를 반환한다(회귀 0).
    """
    from io import BytesIO
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from db.models import Schedule, ScheduleEntry, Nurse, Shift, RosterConfig, ShiftManage
    import calendar

    # ───────── 1) 데이터 로드 ─────────
    schedule = db.query(Schedule).filter(
        Schedule.schedule_id == schedule_id,
        Schedule.group_id == target_group_id,
        Schedule.dropped == False
    ).first()
    if not schedule:
        raise ValueError("스케줄을 찾을 수 없습니다.")

    year, month = schedule.year, schedule.month
    days_in_month = calendar.monthrange(year, month)[1]

    nurses = db.query(Nurse.nurse_id, Nurse.name, Nurse.experience, Nurse.sequence, Nurse.role).filter(
        Nurse.group_id == target_group_id
    ).order_by(Nurse.sequence.asc(), Nurse.nurse_id.asc()).all()

    # ───────── 전입자(인바운드) union — 팀별보기·기본 모두 동일하게 적용 ─────────
    # 이 schedule의 ScheduleEntry에는 있으나 현재 group 멤버가 아닌 간호사(이동 발효 전
    # nurses.group_id 미반영 등)를 추가한다. 화면(get_roster_by_schedule_id)과 동일 기준.
    _home_nurse_ids = {n.nurse_id for n in nurses}
    _entry_nurse_ids = {
        row.nurse_id
        for row in db.query(ScheduleEntry.nurse_id)
        .filter(ScheduleEntry.schedule_id == schedule_id)
        .distinct()
        .all()
    }
    _inbound_ids = _entry_nurse_ids - _home_nurse_ids
    if _inbound_ids:
        inbound_nurses = db.query(
            Nurse.nurse_id, Nurse.name, Nurse.experience, Nurse.sequence, Nurse.role
        ).filter(
            Nurse.nurse_id.in_(_inbound_ids)
        ).order_by(Nurse.sequence.asc(), Nurse.nurse_id.asc()).all()
        nurses = list(nurses) + list(inbound_nurses)

    # ───────── 팀별보기 분기 결정 ─────────
    # 근무표 만들기 화면 '팀별보기'(filterTeam)와 **완전 동일 규칙**으로 팀을 구분한다:
    #   teamId = is_night_dedicated ? None : (as_of_team ?? nurses.team_id 캐시)
    #   gate: teamId 가 활성 팀(Team.active==1) 목록에 없으면 미등록(None).
    # 소속·N전담·시점팀(as_of_team)은 화면과 같은 group_members_in_month 를 그대로 재사용한다.
    # 전원 미등록(팀 배정 전무)이거나 group_by_team=False 면 기존 레이아웃으로 폴백.
    team_view = False
    team_of: dict[str, Optional[int]] = {}     # nurse_id(str) -> gated team_id(int)|None
    nurse_team_label: dict[str, str] = {}      # nurse_id(str) -> 팀명/"미등록"
    if group_by_team:
        from services.assignment_service import group_members_in_month
        from services.team_period import _coerce_team_int
        from db.models import Team

        # (1) 화면 memberStatusMap 동치: as_of_team(시점 팀, N전담=None) + is_night_dedicated.
        _member_map = {
            m["nurse_id"]: m
            for m in group_members_in_month(db, target_group_id, year, month)["members"]
        }
        # (2) knownTeamIds·팀명 = 활성 팀(useTeamsQuery/list_teams_with_members 와 동일 소스).
        _team_name_map = {
            t.team_id: t.team_name
            for t in db.query(Team.team_id, Team.team_name).filter(
                Team.office_id == current_user.office_id,
                Team.group_id == target_group_id,
                Team.active == 1,
            ).all()
        }
        _known_team_ids = set(_team_name_map.keys())
        # (3) as_of_team 이 없을 때의 캐시 폴백(nurses.team_id) — filterTeam 의 `?? nurse.team_id`.
        _cache_team = {
            str(nid): _coerce_team_int(tid)
            for nid, tid in db.query(Nurse.nurse_id, Nurse.team_id).filter(
                Nurse.nurse_id.in_([str(n.nurse_id) for n in nurses])
            ).all()
        }

        for n in nurses:
            nid = str(n.nurse_id)
            m = _member_map.get(nid)
            # N전담도 직접 배정한 팀(as_of_team=period)을 그대로 표시 — 직접 저장한 팀이
            #   '미지정'으로 가려지던 문제 해소. 프론트도 동일하게
            #   `is_night_dedicated ? None : as_of_team` → `as_of_team` 으로 맞춰야 한다.
            as_of = m.get("as_of_team") if m else None
            tid = as_of if as_of is not None else _cache_team.get(nid)
            # 활성 팀 메타에 없으면(미존재/비활성/타그룹 캐시) 미등록으로 안전 배치(누락 방지).
            team_of[nid] = tid if (tid is not None and tid in _known_team_ids) else None

        if any(tid is not None for tid in team_of.values()):
            team_view = True
            # 팀 블록 순서 = 화면 sortTeamGroupsByTeamName(팀명 기준) + 미등록 맨 끝.
            #   stable sort → 같은 팀(또는 미등록) 내부는 기존 sequence 순서 유지.
            _ordered = _sort_team_ids_by_name(_team_name_map)
            _rank = {tid: i for i, tid in enumerate(_ordered)}
            _unassigned_rank = len(_ordered)
            nurses = sorted(
                nurses,
                key=lambda n: _rank.get(team_of.get(str(n.nurse_id)), _unassigned_rank),
            )
            for n in nurses:
                tid = team_of.get(str(n.nurse_id))
                nurse_team_label[str(n.nurse_id)] = (
                    "미등록" if tid is None else (_team_name_map.get(tid) or f"팀 {tid}")
                )

    entries = db.query(ScheduleEntry).filter(ScheduleEntry.schedule_id == schedule_id).all()

    # alias_map 생성
    alias_map: dict[str, str] = {}
    sm_rows = db.query(ShiftManage).filter(
        ShiftManage.office_id == current_user.office_id,
        ShiftManage.group_id == target_group_id,
    ).all()
    for row in sm_rows:
        if not row.main_code:
            continue
        base = row.main_code.upper()
        alias_map[base] = base
        if row.codes:
            for c in row.codes:
                alias_map[str(c).upper()] = base
    alias_map.setdefault('OFF', 'O')
    alias_map.setdefault('O', 'O')

    def to_base(code: str) -> str:
        if not code:
            return '-'
        u = code.upper()
        if u in alias_map:
            return alias_map[u]
        if u.startswith('D'): return 'D'
        if u.startswith('E'): return 'E'
        if u.startswith('N'): return 'N'
        if u in ('O', 'OFF'): return 'O'
        return u

    # ───────── 2) 실제 사용된 모든 base 코드 수집 ─────────
    used_codes = set()
    for e in entries:
        if e.shift_id:
            base = to_base(e.shift_id)
            if base != '-':
                used_codes.add(base)

    core_codes = ['D', 'E', 'N', 'O']
    extra_codes = sorted(used_codes - set(core_codes))
    tail_labels = core_codes + extra_codes
    summary_cols = len(tail_labels)

    # nurse별 매핑
    by_nurse: dict[str, dict[int, str]] = {}
    for e in entries:
        by_nurse.setdefault(e.nurse_id, {})[e.work_date.day] = e.shift_id

    # ───────── 3) 워크북/시트 ─────────
    wb = Workbook()
    ws = wb.active
    ws.title = "근무표"

    # 스타일
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=20)
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    gray_fill = PatternFill("solid", fgColor="DEE2E6")
    highlight_fill = PatternFill("solid", fgColor="FFFACD")

    # ───────── 4) 제목 영역 ─────────
    title = f"{year}년 {month}월 근무표"
    # 팀별보기면 맨 앞(컬럼1)에 팀 컬럼 1개 추가 → static_cols 4→5.
    # 배치: 팀(1) · 번호(2) · 이름(3) · 구분(4) · 경력(5). 기본: 번호(1) · 이름(2) · 구분(3) · 경력(4).
    # 날짜 시작열·spacer·요약열은 static_cols 에서 파생되므로 자동 보정.
    static_cols = 5 if team_view else 4
    idx_col  = 2 if team_view else 1  # 번호 컬럼
    name_col = 3 if team_view else 2  # 이름 컬럼
    role_col = name_col + 1           # 구분
    exp_col  = name_col + 2           # 경력
    spacer_cols = 2
    total_cols = static_cols + days_in_month + spacer_cols + summary_cols

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=total_cols)
    ws.cell(row=2, column=1, value=title).font = title_font
    ws.cell(row=2, column=1).alignment = center

    # ───────── 5) 헤더 행 ─────────
    header_row = 7
    if team_view:
        ws.cell(row=header_row, column=1, value="팀").font = header_font
        ws.cell(row=header_row, column=2, value="번호").font = header_font
    else:
        ws.cell(row=header_row, column=1, value="번호").font = header_font
    ws.cell(row=header_row, column=name_col, value="이름").font = header_font
    ws.cell(row=header_row, column=role_col, value="구분").font = header_font
    ws.cell(row=header_row, column=exp_col, value="경력").font = header_font

    for c in range(1, static_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.alignment = center
        cell.border = border_all
        cell.fill = gray_fill

    for d in range(1, days_in_month + 1):
        col = static_cols + d
        cell = ws.cell(row=header_row, column=col, value=d)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_all
        cell.fill = gray_fill

    tail_start_col = static_cols + days_in_month + spacer_cols + 1
    for i, lab in enumerate(tail_labels):
        col = tail_start_col + i
        cell = ws.cell(row=header_row, column=col, value=lab)
        cell.font = header_font
        cell.alignment = center
        cell.border = border_all
        cell.fill = gray_fill

    # 열 너비 (팀별보기: A=팀8, B=번호5 / 기본: A=번호5)
    if team_view:
        ws.column_dimensions['A'].width = 8  # 팀
        ws.column_dimensions['B'].width = 5  # 번호
    else:
        ws.column_dimensions['A'].width = 5  # 번호
    ws.column_dimensions[get_column_letter(name_col)].width = 12  # 이름
    ws.column_dimensions[get_column_letter(role_col)].width = 6   # 구분
    ws.column_dimensions[get_column_letter(exp_col)].width = 6    # 경력
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(static_cols + d)].width = 4
    for s in range(spacer_cols):
        ws.column_dimensions[get_column_letter(static_cols + days_in_month + 1 + s)].width = 3
    for i, lab in enumerate(tail_labels):
        col_letter = get_column_letter(tail_start_col + i)
        ws.column_dimensions[col_letter].width = 6 if len(lab) > 1 else 5

    # ───────── 6) 본문 ─────────
    start_row = header_row + 1
    daily_counts = {d: {code: 0 for code in tail_labels} for d in range(1, days_in_month + 1)}

    def write_nurse_row(n, r: int, idx: int):
        """간호사 1명을 r 행에 작성하고 daily_counts 를 누적한다."""
        is_current_user = (str(n.nurse_id) == str(current_user.nurse_id))

        ws.cell(row=r, column=idx_col, value=idx)
        ws.cell(row=r, column=name_col, value=n.name)
        ws.cell(row=r, column=role_col, value=n.role)
        ws.cell(row=r, column=exp_col, value=n.experience)
        # 팀 컬럼(col1)은 본문 작성 후 팀 그룹 경계로 세로병합하므로 여기선 비워둔다.

        for c in range(1, static_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border_all
            if is_current_user:
                cell.fill = highlight_fill

        row_counts = {code: 0 for code in tail_labels}
        schedule_map = by_nurse.get(n.nurse_id, {})

        for d in range(1, days_in_month + 1):
            shift_code = schedule_map.get(d, '-')
            cell = ws.cell(row=r, column=static_cols + d, value=shift_code)
            cell.alignment = center
            cell.border = border_all
            if is_current_user:
                cell.fill = highlight_fill

            base = to_base(shift_code)
            if base in row_counts:
                row_counts[base] += 1
                daily_counts[d][base] += 1

        # 요약 열: 0도 그대로 표시 (빈칸 → 0)
        for i, lab in enumerate(tail_labels):
            col = tail_start_col + i
            value = row_counts[lab]
            cell = ws.cell(row=r, column=col, value=value)  # ← 변경: value 그대로
            cell.alignment = center
            cell.border = border_all
            if is_current_user:
                cell.fill = highlight_fill

    # 본문은 팀별보기·기본 모두 동일한 연속 행으로 작성한다(레이아웃 동일). 팀별보기면
    # 작성하면서 같은 팀 라벨의 연속 행 구간을 모아 컬럼2를 세로 병합한다.
    for idx, n in enumerate(nurses, start=1):
        write_nurse_row(n, start_row + idx - 1, idx)
    last_row = start_row + len(nurses) - 1

    if team_view and nurses:
        # 정렬된 nurses 를 순회하며 같은 team_id 의 연속 구간(첫행~끝행)을 모아 컬럼1을
        # 세로 병합한다. 그룹 키는 team_of(gated team_id, None=미등록) — 팀명이 우연히 같아도
        # 다른 팀은 합치지 않는다. 표시 라벨은 nurse_team_label(팀명/"미등록") 사용.
        col_team = 1
        _UNASSIGNED = object()  # None 과 "구간 미시작" 을 구분하기 위한 센티넬
        run_key = _UNASSIGNED
        run_start = start_row
        run_label = "미등록"
        for idx, n in enumerate(nurses):
            r = start_row + idx
            tid = team_of.get(str(n.nurse_id))  # int|None
            label = nurse_team_label.get(str(n.nurse_id), "미등록")
            if run_key is _UNASSIGNED:
                run_key, run_start, run_label = tid, r, label
            elif tid != run_key:
                _merge_team_cell(ws, col_team, run_start, r - 1, run_label,
                                 center, border_all, gray_fill)
                run_key, run_start, run_label = tid, r, label
        _merge_team_cell(ws, col_team, run_start, last_row, run_label,
                         center, border_all, gray_fill)

    # ───────── 7) 풋터 ─────────
    # 라벨 위치는 이름/구분 컬럼에 맞춰 팀별보기 시 한 칸씩 밀린다(레이아웃 동일).
    footer_start = last_row + 2
    ws.cell(row=footer_start, column=name_col, value="일일 근무 현황").font = header_font

    def write_footer_row(label: str, values: list[int], row_idx: int):
        ws.cell(row=row_idx, column=role_col, value=label).font = header_font
        for c in range(1, static_cols):
            ws.cell(row=row_idx, column=c).border = border_all

        for d in range(1, days_in_month + 1):
            val = values[d - 1]
            cell = ws.cell(row=row_idx, column=static_cols + d, value=val)  # ← 변경: val 그대로 (0도 표시)
            cell.alignment = center
            cell.border = border_all

        for i in range(spacer_cols + summary_cols):
            col = static_cols + days_in_month + 1 + i
            ws.cell(row=row_idx, column=col).border = border_all

    for i, lab in enumerate(tail_labels):
        row_idx = footer_start + 1 + i
        vals = [daily_counts[d][lab] for d in range(1, days_in_month + 1)]
        write_footer_row(lab, vals, row_idx)

    # ───────── 8) 테두리 보정 ─────────
    max_col = tail_start_col + len(tail_labels) - 1
    for row in ws.iter_rows(min_row=header_row, max_row=footer_start + len(tail_labels) + 1,
                            min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None and (cell.border is None or cell.border.left.style is None):
                cell.border = border_all

    # ───────── 저장 ─────────
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue(), team_view


def export_members_excel_bytes(office_id: str) -> bytes:
    """ADM용 멤버 목록 엑셀 생성.

    - 입력: office_id
    - 컬럼: 대분류, 중분류, 소분류, 부서명, 사번, 직원명, 계정 ID, 직무, 경력, 수간호사여부, 입사일, 생년월일, 연락처
    - 반환: 생성된 xlsx 바이트
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    rows = msdb_manager.fetch_all(Member.member_export_by_office(), params=(str(office_id),)) or []

    headers = [
        ("big_kind_name", "대분류"),
        ("middle_kind_name", "중분류"),
        ("small_kind_name", "소분류"),
        ("mb_part_name", "부서명"),
        ("OfficeEmpNum", "사번"),
        ("MemberID", "계정 ID"),
        ("EmployeeName", "직원명"),
        ("duty", "직무"),
        ("career", "경력"),
        ("headnurse", "수간호사여부"),
        ("joindate", "입사일"),
        ("DateOfBirth", "생년월일"),
        ("PortableTel", "연락처"),
        ("Gender", "성별"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "구성원"

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    header_font = Font(bold=True)
    gray = PatternFill("solid", fgColor="DEE2E6")
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 안내 문구 (맨 윗줄, 노란색 배경)
    guide_text = '사번 컬럼부터 우측으로 필요한 정보를 복사해 템플릿에 그룹(병동)별로 추가하여 업로드 하시면 됩니다'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    gcell = ws.cell(row=1, column=1, value=guide_text)
    gcell.alignment = left
    gcell.fill = PatternFill("solid", fgColor="FFF59D")  # 연노랑
    gcell.border = border_all

    # 헤더 (2행부터)
    for i, (_, label) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=label)
        cell.font = header_font
        cell.alignment = center
        cell.fill = gray
        cell.border = border_all

    # 데이터 (3행부터)
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, (key, _) in enumerate(headers, start=1):
            # pyodbc.Row 또는 dict 형태 지원
            try:
                val = row[key]
            except Exception:
                val = row.get(key) if hasattr(row, 'get') else None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = left
            cell.border = border_all

    # 약간의 폭 조정
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def upload2_confirm(rows: List[Dict[str, Any]], user: UserSchema, db: Session, target_group_id: str) -> Dict[str, Any]:
    """업로드2: 검증된 행을 nurses 테이블에 저장 (신규/업데이트)"""
    print("[upload2_confirm] 함수 시작")
    print(f"  • target_group_id: {target_group_id}")
    print(f"  • rows 개수: {len(rows)}")

    try:
        if not target_group_id:
            print("[ERROR] target_group_id 누락")
            return {"success": 0, "errors": [{"row": 0, "reason": "group_id가 필요합니다."}]}

        saved = 0
        updated = 0
        errors = []  # ★★★ 이 한 줄이 핵심! NameError 방지 ★★★

        for idx, item in enumerate(rows, 1):
            print(f"[행 {idx}] 처리: account_id={item.get('account_id')}")

            account_id = str(item.get('account_id', '')).strip()
            if not account_id:
                errors.append({"row": item.get('row', 0), "reason": "account_id 누락"})
                continue

            name = str(item.get('name', '')).strip() or account_id
            emp_num = str(item.get('emp_num', '')).strip() or None
            role = str(item.get('role', 'RN')).strip()
            experience = item.get('experience')
            if not isinstance(experience, int):
                experience = 1

            is_head_nurse = bool(item.get('is_head_nurse', False))

            joining_dt = None
            if item.get('joining_date'):
                try:
                    joining_dt = pd.to_datetime(item['joining_date']).to_pydatetime()
                except:
                    pass

            birth_date = str(item.get('birth_date', '')).strip()[:10] or None
            phone_number = str(item.get('phone_number', '')).strip()[:20] or None
            gender = str(item.get('gender', '')).strip()[:3] or None
            email = str(item.get('email', '')).strip()[:100] or None

            is_night_nurse = item.get('is_night_nurse', []) or []
            work_shifts = item.get('work_shifts', []) or []

            nurse_id = item.get('nurse_id')
            # if not nurse_id or not isinstance(nurse_id, str) or len(str(nurse_id).strip()) < 8:
            #     nurse_id = str(uuid.uuid4())
            print(f"   → nurse_id 자동 생성: {nurse_id}")

            existing = db.query(NurseModel).filter(
                NurseModel.account_id == account_id
            ).first()

            if existing:
                print(f"   → 기존 레코드 업데이트 (nurse_id={existing.nurse_id})")
                if name and existing.name != name:
                    existing.name = name
                existing.emp_num = emp_num
                existing.role = role
                existing.experience = experience
                existing.is_head_nurse = is_head_nurse
                existing.office_id = user.office_id
                if joining_dt:
                    existing.joining_date = joining_dt
                existing.birth_date = birth_date
                existing.phone_number = phone_number
                existing.gender = gender
                existing.email = email
                existing.is_night_nurse = is_night_nurse
                existing.work_shifts = work_shifts
                updated += 1
            else:
                print("   → 신규 등록")
                try:
                    seq_next = get_next_sequence(target_group_id, 1, db, role=role)
                    new_nurse = NurseModel(
                        nurse_id=nurse_id,
                        group_id=target_group_id,
                        office_id=user.office_id,
                        account_id=account_id,
                        emp_num=emp_num,
                        name=name,
                        experience=experience,
                        role=role,
                        level_='일반',
                        is_head_nurse=is_head_nurse,
                        is_night_nurse=is_night_nurse,
                        personal_off_adjustment=0,
                        preceptor_id=None,
                        joining_date=joining_dt,
                        sequence=seq_next,
                        active=1,
                        birth_date=birth_date,
                        phone_number=phone_number,
                        gender=gender,
                        email=email,
                        work_shifts=work_shifts,
                    )
                    db.add(new_nurse)
                    saved += 1
                except Exception as inner_e:
                    errors.append({"row": item.get('row', 0), "reason": str(inner_e)})
                    print(f"   → 신규 등록 실패: {str(inner_e)}")

        print(f"[COMMIT 직전] saved={saved}, updated={updated}, errors={len(errors)}")
        db.commit()
        print("[COMMIT 성공]")
        return {"success": saved + updated, "saved": saved, "updated": updated, "errors": errors}

    except Exception as e:
        print(f"[전체 예외] {type(e).__name__}: {str(e)}")
        db.rollback()
        return {"success": 0, "errors": [{"row": 0, "reason": str(e)}]}


# def direct_validate_and_confirm(rows: List[Dict[str, Any]], user: UserSchema, db: Session, group_id: str) -> Dict[str, Any]:
#     """
#     직접 입력된 간호사 데이터 검증 + 저장 (엑셀 없이 사용)
#     """
#     try:
#         office_id = user.office_id

#         # 허용 계정 목록 조회
#         rows_allowed = msdb_manager.fetch_all(Member.member_accounts_by_office(), params=(str(office_id),))
#         allowed = {}
#         for r in rows_allowed or []:
#             acc = str(r.get('account_id', '')).strip()
#             if acc:
#                 allowed[acc] = (
#                     str(r.get('name', '')).strip(),
#                     r.get('EmpAuthGbn'),
#                     str(r.get('nurse_id', uuid.uuid4())).strip()
#                 )

#         normalized = []
#         errors = []
#         head_count = 0
#         acc_set = set()

#         for i, row in enumerate(rows):
#             ridx = i + 1
#             row_errs = []

#             account_id = str(row.get('account_id', '')).strip()
#             name = str(row.get('name', '')).strip()
#             role = str(row.get('role', 'RN')).strip()
#             exp_val = row.get('experience')
#             try:
#                 exp_val = int(exp_val) if exp_val is not None else None
#             except:
#                 row_errs.append("경력은 숫자여야 합니다.")

#             is_head = str(row.get('is_head_nurse', 'N')).upper() == 'Y'
#             if is_head:
#                 head_count += 1

#             joining_dt = row.get('joining_date')

#             if not account_id:
#                 row_errs.append("계정 ID 누락")
#             elif not re.match(r'^\S{1,50}$', account_id):
#                 row_errs.append("계정 ID 형식 오류")
#             elif account_id not in allowed:
#                 row_errs.append(f"허용되지 않은 계정: {account_id}")

#             if not name:
#                 name = allowed.get(account_id, ('', None, ''))[0]
#                 if not name:
#                     row_errs.append("직원명 누락")

#             if not role:
#                 row_errs.append("직무 누락")

#             if account_id in acc_set:
#                 row_errs.append("중복 계정 ID")
#             else:
#                 acc_set.add(account_id)

#             normalized.append({
#                 'row': ridx,
#                 'emp_num': row.get('emp_num'),
#                 'account_id': account_id,
#                 'name': name,
#                 'role': role,
#                 'experience': exp_val,
#                 'is_head_nurse': is_head,
#                 'joining_date': joining_dt,
#                 'nurse_id': allowed.get(account_id, ('', '', str(uuid.uuid4())))[2],
#                 'birth_date': row.get('birth_date'),
#                 'phone_number': row.get('phone_number'),
#                 'is_night_nurse': row.get('is_night_nurse', []),
#                 'gender': row.get('gender')
#             })

#             if row_errs:
#                 errors.append({'row': ridx, 'reason': '; '.join(row_errs)})

#         # 글로벌 검증
#         existing_heads = db.query(NurseModel).filter(
#             NurseModel.office_id == office_id,
#             NurseModel.group_id == group_id,
#             NurseModel.is_head_nurse == 1
#         ).count()

#         if head_count == 0 and existing_heads == 0:
#             errors.append({'row': 0, 'reason': '수간호사는 최소 1명 이상이어야 합니다.'})

#         existing_accs = db.query(NurseModel.account_id).filter(NurseModel.account_id.in_(list(acc_set))).all()
#         if existing_accs:
#             for (acc,) in existing_accs:
#                 errors.append({'row': 0, 'reason': f'이미 등록된 계정 ID: {acc}'})

#         if errors:
#             return {
#                 'success': 0,
#                 'errors': errors,
#                 'rows': normalized,
#                 'summary': {'total': len(normalized), 'error_count': len(errors)}
#             }

#         # group_id 강제 체크 (ADM 대비)
#         if not group_id:
#             raise ValueError("group_id가 누락되었습니다. 관리자 계정은 반드시 group_id를 지정해야 합니다.")

#         # 검증 통과 → 저장
#         return upload2_confirm(normalized, user, db, group_id)

#     except Exception as e:
#         return {"success": 0, "errors": [{"row": 0, "reason": str(e)}]}


# def integrated_member_and_nurse_register(
#     members: List[Dict[str, Any]],
#     user: UserSchema,
#     db: Session,
#     group_id: str
# ) -> Dict[str, Any]:
#     """
#     직접 입력으로 신규 직원 계정 생성 + 근무자 등록 통합 처리
#     """
#     try:
#         office_code = user.office_id
#         emp_seq_no = getattr(user, 'EmpSeqNo', user.nurse_id or '')
#         reg_date = datetime.now()

#         if not members:
#             return {"success": False, "errors": [{"reason": "입력 데이터가 없습니다."}]}

#         # group_id 필수 체크 (ADM 대비)
#         if not group_id:
#             return {"success": False, "errors": [{"reason": "group_id가 누락되었습니다. 대상 병동을 선택해주세요."}]}

#         # 1~4 단계: member 임시 저장 → 외부 API 호출 (기존 그대로)
#         df = pd.DataFrame(members)
#         df['num'] = range(1, len(df) + 1)

#         rename_map = {
#             '사번': 'EmpNum',
#             '회원 아이디': 'MemberID',
#             '이름': 'EmployeeName',
#             '성별': 'Gender',
#             '생년월일': 'Birthday',
#             '입사년월': 'JoinDate',
#             '전화번호': 'Tel',
#             '휴대폰 번호': 'PortableTel',
#             '이메일': 'Email',
#             '주소': 'Address',
#             '부서장': 'Manager',
#             '상위부서': 'Depth1',
#             '하위부서1': 'Depth2',
#             '하위부서2': 'Depth3',
#             '직위': 'Posin',
#             '경력': 'career',
#             '직무': 'duty',
#             '수간호사여부': 'headnurse',
#             '킵여부': 'nightkeep'
#         }
#         df = df.rename(columns=rename_map)

#         # 타입 변환 및 검증 (기존 로직 그대로 유지)
#         # ... (중략: 필수값 체크, 이메일/생년월일/성별/경력/부서 검증 등) ...

#         # 임시 테이블 저장 및 외부 API 호출 (기존 그대로)
#         # ... (중략) ...

#         # 5. nurses 등록
#         nurse_rows = []
#         for row in members:
#             nurse_rows.append({
#                 'emp_num': row.get('사번'),
#                 'account_id': row.get('회원 아이디'),
#                 'name': row.get('이름'),
#                 'role': row.get('직무', 'RN'),
#                 'experience': row.get('경력'),
#                 'is_head_nurse': str(row.get('수간호사여부', 'N')).upper() == 'Y',
#                 'joining_date': row.get('입사년월'),
#                 'birth_date': row.get('생년월일'),
#                 'phone_number': row.get('휴대폰 번호'),
#                 'gender': row.get('성별')
#             })

#         # 검증 + 저장
#         nurse_result = direct_validate_and_confirm(nurse_rows, user, db, group_id)

#         return {
#             "success": nurse_result.get('success', 0) > 0,
#             "member_count": len(members),
#             "nurse_result": nurse_result,
#             "message": "신규 직원 계정 생성 및 근무자 등록 완료" if nurse_result.get('success', 0) > 0 else "nurses 등록 실패"
#         }

#     except Exception as e:
#         return {"success": False, "errors": [{"reason": f"처리 중 오류 발생: {str(e)}"}]}


# def integrated_member_and_nurse_register(
#     members: List[Dict[str, Any]],
#     user: UserSchema,
#     db: Session,
#     group_id: str
# ) -> Dict[str, Any]:
#     """
#     직접 입력으로 신규 직원 계정 생성 + 근무자 등록 통합 처리
#     - member 테이블 생성 → 외부 API 동기화 → nurses 테이블 등록
#     """
#     try:
#         office_code = user.office_id
#         emp_seq_no = getattr(user, 'EmpSeqNo', user.nurse_id or '')
#         reg_date = datetime.now()

#         if not members:
#             return {"success": False, "errors": [{"reason": "입력 데이터가 없습니다."}]}

#         # group_id 필수 체크 (ADM 대비)
#         if not group_id:
#             return {"success": False, "errors": [{"reason": "group_id가 누락되었습니다. 대상 병동을 선택해주세요."}]}

#         # 1. DataFrame 변환 및 컬럼 매핑
#         df = pd.DataFrame(members)
#         df['num'] = range(1, len(df) + 1)

#         rename_map = {
#             '사번': 'EmpNum',
#             '회원 아이디': 'MemberID',
#             '이름': 'EmployeeName',
#             '성별': 'Gender',
#             '생년월일': 'Birthday',
#             '입사년월': 'JoinDate',
#             '전화번호': 'Tel',
#             '휴대폰 번호': 'PortableTel',
#             '이메일': 'Email',
#             '주소': 'Address',
#             '부서장': 'Manager',
#             '상위부서': 'Depth1',
#             '하위부서1': 'Depth2',
#             '하위부서2': 'Depth3',
#             '직위': 'Posin',
#             '경력': 'career',
#             '직무': 'duty',
#             '수간호사여부': 'headnurse',
#             '킵여부': 'nightkeep'
#         }
#         df = df.rename(columns=rename_map)

#         # 타입 변환
#         df['Birthday'] = pd.to_numeric(df.get('Birthday'), errors='coerce').astype('Int64')
#         df['career'] = pd.to_numeric(df.get('career'), errors='coerce').astype('Int64')

#         # 2. 검증 (기존 로직 그대로 유지 - 생략 가능 시 주석 처리)
#         error_rows = []
#         # ... (중략: 중복, 필수값, 이메일, 생년월일, 성별, 수간호사, 경력, 부서, MemberID 중복 체크 등) ...

#         if error_rows:
#             error_df = pd.concat(error_rows, ignore_index=True).replace({np.nan: ''})
#             error_df.drop_duplicates(inplace=True)
#             return {"success": False, "errors": error_df.to_dict(orient='records')}

#         # 3. 임시 테이블 저장
#         df['OfficeCode'] = office_code
#         df['EmpSeqNo'] = emp_seq_no
#         df['RegDate'] = reg_date
#         df = df.replace({np.nan: ''})

#         insert_cols = [
#             'num', 'OfficeCode', 'EmpSeqNo', 'EmpNum', 'MemberID', 'EmployeeName', 'Gender',
#             'Birthday', 'JoinDate', 'Tel', 'PortableTel', 'Email', 'Address', 'Manager',
#             'Depth1', 'Depth2', 'Depth3', 'Posin', 'RegDate', 'career', 'duty', 'headnurse', 'nightkeep'
#         ]
#         df_insert = df[insert_cols]
#         data_to_insert = [tuple(row) for row in df_insert.itertuples(index=False)]

#         msdb_manager.execute(Setting.delete_member(), params=(office_code, emp_seq_no))
#         msdb_manager.bulk_execute(Setting.insert_member(), data_to_insert)

#         # 모바일 설정
#         member_ids = [mid for mid in df["MemberID"].tolist() if mid]
#         mobile_params = [(mid, reg_date) for mid in member_ids]
#         if mobile_params:
#             msdb_manager.bulk_execute(Setting.insert_mobile_user_setting_list(), mobile_params)

#         # 4. 외부 API 호출 (Member 테이블 실제 생성)
#         token = create_access_token(data={"clientSecret": os.getenv("CLIENT_SECRET"), "clientId": os.getenv("CLIENT_ID")})
#         response = requests.post(
#             "https://gw.meditong.com/bizadmin/setting/member_excel_ai_ok.asp",
#             data=f"officeCode={office_code}&EmpSeqNo={emp_seq_no}&Token={token}",
#             headers={'Content-Type': 'application/x-www-form-urlencoded'}
#         )
#         if response.status_code != 200:
#             return {"success": False, "errors": [{"reason": f"외부 동기화 API 실패: {response.text}"}]}

#         # 5. nurses 테이블 등록
#         nurse_rows = []
#         for row in members:
#             nurse_rows.append({
#                 'emp_num': row.get('사번'),
#                 'account_id': row.get('회원 아이디'),
#                 'name': row.get('이름'),
#                 'role': row.get('직무', 'RN'),
#                 'experience': row.get('경력'),
#                 'is_head_nurse': str(row.get('수간호사여부', 'N')).upper() == 'Y',
#                 'joining_date': row.get('입사년월'),
#                 'birth_date': row.get('생년월일'),
#                 'phone_number': row.get('휴대폰 번호'),
#                 'gender': row.get('성별')
#             })

#         # 검증 + 저장 (upload2_confirm 호출)
#         nurse_result = direct_validate_and_confirm(nurse_rows, user, db, group_id)

#         return {
#             "success": nurse_result.get('success', 0) > 0,
#             "member_count": len(members),
#             "nurse_result": nurse_result,
#             "message": "신규 직원 계정 생성 및 근무자 등록 완료" if nurse_result.get('success', 0) > 0 else "nurses 등록 실패"
#         }

#     except Exception as e:
#         return {"success": False, "errors": [{"reason": f"처리 중 오류 발생: {str(e)}"}]}