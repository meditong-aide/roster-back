"""
시프트(근무코드) 관리 관련 서비스 로직 모듈
- DB 쿼리, 데이터 가공 등 라우터에서 분리
- 엑셀 일괄 업로드(템플릿/검증/확정) 포함
- 모든 함수는 한글 docstring, 한글 print/logging, PEP8 스타일 적용
"""
import math
import re
import tempfile
from datetime import time as dt_time
from typing import Any, Dict, List, Optional, Set

import pandas as pd
from fastapi import HTTPException
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session
from sqlalchemy import func

from db.models import Shift, Nurse, Group, ShiftManage, ScheduleEntry
from schemas.auth_schema import User as UserSchema
from services.group_access import (
    caller_is_head_nurse,
    resolve_effective_group,
    resolve_home_group_id,
)


def _assert_off_swap_target_valid(
    db: Session,
    *,
    group_id: str,
    target_value: bool,
    shift_type: Optional[str],
    self_id: Optional[int] = None,
) -> None:
    """off_swap_target=True 저장 전 정책 검증.

    1) 변환 대상 shift 의 type 은 '근무' 가 아니어야 한다.
       (근무 코드를 변환 target 으로 두면 OFF→근무 치환 시 일별 coverage oversupply 발생)
    2) 동일 group_id 내 off_swap_target=True 인 row 는 단 1건만 허용.
       (다수면 _resolve_target_shift 가 sequence ASC 첫 번째만 채택해 의도와 다른 결과)

    target_value=False 면 검증 skip. self_id 는 update 시 본인 row 를 비교 대상에서 제외.
    """
    if not target_value:
        return
    if str(shift_type or "").strip() == "근무":
        raise HTTPException(
            status_code=400,
            detail="off_swap_target=True 는 type='근무' 인 근무코드에는 설정할 수 없습니다.",
        )
    q = db.query(Shift).filter(Shift.group_id == group_id, Shift.off_swap_target == True)  # noqa: E712
    if self_id is not None:
        q = q.filter(Shift.id != self_id)
    other = q.first()
    if other is not None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"off_swap_target=True 는 그룹당 1건만 설정 가능합니다. "
                f"기존 설정: shift_id={other.shift_id}, name={other.name}"
            ),
        )


# ──────────────────────────────────────────────────────────────
# 색상 풀 (brightness ≤ 186 사전 필터링, 흰색 텍스트 가독성 보장)
# ──────────────────────────────────────────────────────────────
COLOR_POOLS: Dict[str, List[str]] = {
    "데이": ["#42AF3A", "#4AA36E", "#0A9999"],
    "이브닝": ["#5777F3", "#7589B8", "#6698CB", "#1C80F1"],
    "나이트": ["#A184E5", "#7C38C0", "#C082DD", "#976BA7", "#C69EC4"],
    "미드": ["#E6A817", "#D4942A", "#C98B2A", "#B8860B", "#DAA520"],
    "holiday": [
        "#EC71A2", "#F45BA8", "#F5A4F0", "#F29B87",
        "#FFA24A", "#FF7D3B", "#F54851",
    ],
    "etc": ["#CF847A", "#B5704F", "#A67B5B", "#8B6E5A", "#9C6644", "#7F5539"],
}

# shift_gb 한글 → ShiftManage slot_map 영문 코드 변환
SHIFT_GB_TO_SLOT_CODE: Dict[str, str] = {
    "데이": "D",
    "이브닝": "E",
    "나이트": "N",
    "미드": "M",
}

# shift_gb 허용 값
VALID_SHIFT_GB = {"데이", "이브닝", "나이트", "미드", "고정"}



def _pick_pool_key(shift_gb: Optional[str], shift_type: str) -> str:
    """shift_gb / type 으로 색상 풀 키 결정."""
    if shift_gb in ("데이", "이브닝", "나이트", "미드"):
        return shift_gb
    if shift_type == "휴무":
        return "holiday"
    return "etc"


def assign_color(
    shift_gb: Optional[str],
    shift_type: str,
    existing_colors: Set[str],
    counter: Dict[str, int],
) -> str:
    """적절한 풀에서 색상을 순차 배정. 미사용 색상 우선, 소진 시 순환."""
    pool_key = _pick_pool_key(shift_gb, shift_type)
    pool = COLOR_POOLS[pool_key]
    start = counter.get(pool_key, 0)

    for i in range(len(pool)):
        candidate = pool[(start + i) % len(pool)]
        if candidate not in existing_colors:
            counter[pool_key] = (start + i + 1) % len(pool)
            return candidate

    color = pool[start % len(pool)]
    counter[pool_key] = (start + 1) % len(pool)
    return color


def _parse_time_str(val: Optional[str]) -> Optional[dt_time]:
    """'HH:MM' 문자열 → datetime.time 변환."""
    if not val:
        return None
    parts = val.strip().split(":")
    if len(parts) >= 2:
        return dt_time(int(parts[0]), int(parts[1]))
    return None


def _parse_boolean(value: Any) -> bool:
    """Y/N/YES/TRUE/1/T/참/예 → bool."""
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    s = str(value).strip().upper()
    return s in ("Y", "YES", "TRUE", "1", "T", "참", "예")


def _safe_str(value: Any) -> str:
    """NaN-safe 문자열 변환."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _is_guide_row(row: pd.Series) -> bool:
    """입력 가이드 행 판별."""
    for v in row.values:
        s = _safe_str(v)
        if "입력 가이드" in s or (s.startswith("(") and s.endswith(")")):
            return True
    return False


# shift_gb(영문 D/E/N/M 또는 한글) → ShiftManage.shift_slot
_SHIFT_GB_TO_SLOT = {"D": 1, "E": 2, "N": 3, "M": 5, "데이": 1, "이브닝": 2, "나이트": 3, "미드": 5}
# shift_gb(한글) → main_code(영문). 영문 키는 그대로 통과한다.
_SHIFT_GB_TO_MAIN_CODE = {"데이": "D", "이브닝": "E", "나이트": "N", "미드": "M"}


def _resolve_slot_main(shift_gb: str | None) -> tuple[int | None, str | None]:
    """shift_gb → (shift_slot, main_code). 매핑이 없으면 (None, None)."""
    if shift_gb not in _SHIFT_GB_TO_SLOT:
        return None, None
    return _SHIFT_GB_TO_SLOT[shift_gb], _SHIFT_GB_TO_MAIN_CODE.get(shift_gb, shift_gb)


def _remove_shift_manage_code(
    db: Session,
    office_id: str | None,
    group_id: str,
    shift_id: str | None,
    shift_gb: str | None,
) -> bool:
    """
    shift_manage.codes 에서 shift_id 를 제거합니다(슬롯=shift_gb 매핑).

    - 근무코드 삭제/변경 시 orphan code 가 codes 에 남지 않도록 정리하는 공용 헬퍼.
    - 커밋은 호출자 책임(원자적 트랜잭션 보존).
    """
    if not office_id or not shift_id:
        return False
    target_slot, main_code = _resolve_slot_main(shift_gb)
    if target_slot is None:
        return False
    rows = (
        db.query(ShiftManage)
        .filter(
            ShiftManage.office_id == office_id,
            ShiftManage.group_id == group_id,
            ShiftManage.shift_slot == target_slot,
            ShiftManage.main_code == main_code,
        )
        .all()
    )
    removed = False
    for shift_manage in rows:
        codes = shift_manage.codes or []
        if shift_id in codes:
            shift_manage.codes = [code for code in codes if code != shift_id]
            removed = True
    return removed


def _append_shift_manage_code(
    db: Session,
    office_id: str | None,
    group_id: str,
    shift_id: str,
    shift_gb: str | None,
    old_shift_id: str | None = None,
    old_shift_gb: str | None = None,
) -> None:
    """
    shift_manage.codes에 근무코드를 중복 없이 추가합니다.

    - shift_gb가 D/E/N/M(또는 한글)일 때만 동작하며, slot은 1/2/3/5에 매핑됩니다.
    - 기존에 코드가 있으면 추가하지 않습니다.
    - update 시 shift_id 또는 shift_gb가 바뀌면 이전 슬롯에서 제거한 뒤 새 슬롯에 추가합니다.
    """
    if not office_id:
        return

    def _append(target_gb: str | None, target_id: str) -> bool:
        target_slot, main_code = _resolve_slot_main(target_gb)
        if target_slot is None:
            return False
        shift_manages = (
            db.query(ShiftManage)
            .filter(
                ShiftManage.office_id == office_id,
                ShiftManage.group_id == group_id,
                ShiftManage.shift_slot == target_slot,
                ShiftManage.main_code == main_code,
            )
            .all()
        )
        added = False
        for shift_manage in shift_manages:
            codes = shift_manage.codes or []
            if shift_id not in codes:
                shift_manage.codes = codes + [shift_id]
                added = True
        return added

    removed_any = False
    if old_shift_id and (old_shift_id != shift_id or old_shift_gb != shift_gb):
        removed_any = _remove_shift_manage_code(db, office_id, group_id, old_shift_id, old_shift_gb)

    added_any = _append(shift_gb, shift_id)

    if removed_any or added_any:
        db.commit()


def add_shift_service(req, current_user, db, override_group_id: str | None = None):
    """
    시프트 등록 서비스 함수
    """
    print('----------------------------------[add_shift_service] group_id', override_group_id)
    if not current_user or not (caller_is_head_nurse(db, current_user) or getattr(current_user, 'is_master_admin', False)):
        raise Exception("Permission denied")
    # 그룹 스코프: 조회(LIST)와 동일하게 토큰 group_id 기준으로 group_access 모듈에서 해석+권한검증.
    target_group_id = resolve_effective_group(db, current_user, override_group_id or current_user.group_id)
    if override_group_id:
        g = db.query(Group).filter(Group.group_id == target_group_id).first()
        if not g:
            raise Exception("Group not found")
        office_id = g.office_id
    else:
        nurse = db.query(Nurse).filter(Nurse.nurse_id == current_user.nurse_id).first()
        if not nurse or not nurse.group:
            raise Exception("User group information not found")
        office_id = nurse.group.office_id
    existing_shift = db.query(Shift).filter(
        Shift.shift_id == req.shift_id,
        Shift.group_id == target_group_id
    ).first()
    if existing_shift:
        raise Exception("이미 존재하는 근무코드입니다.")
    _assert_off_swap_target_valid(
        db,
        group_id=target_group_id,
        target_value=bool(getattr(req, "off_swap_target", False) or False),
        shift_type=getattr(req, "type", None),
        self_id=None,
    )
    max_sequence = db.query(func.max(Shift.sequence)).filter(Shift.group_id == target_group_id).scalar() or 0
    new_shift = Shift(
        shift_id=req.shift_id,
        office_id=office_id,
        group_id=target_group_id,
        name=req.name,
        color=req.color,
        start_time=req.start_time,
        end_time=req.end_time,
        type=req.type,
        duration=req.duration,
        allday=req.allday,
        auto_schedule=req.auto_schedule,
        sequence=max_sequence + 1,
        shift_gb=req.shift_gb,
        # 추가
        show_in_preference=getattr(req, "show_in_preference", False), # 프론트 미 전송 시 False
        off_swap_target=bool(getattr(req, "off_swap_target", False) or False),
        description=getattr(req, "description", None),
    )
    db.add(new_shift)
    db.commit()
    db.refresh(new_shift)
    _append_shift_manage_code(
        db=db,
        office_id=office_id,
        group_id=target_group_id,
        shift_id=new_shift.shift_id,
        shift_gb=req.shift_gb,
    )
    return {
        "message": "근무코드가 성공적으로 추가되었습니다.",
        "shift": {
            "shift_id": new_shift.shift_id,
            "name": new_shift.name,
            "color": new_shift.color,
            "sequence": new_shift.sequence,
            "shift_gb": new_shift.shift_gb,
            "description": new_shift.description,
        }
    }


def update_shift_service(req, current_user, db, override_group_id: str | None = None):
    """
    시프트 수정 서비스 함수
    """
    if not current_user or not (caller_is_head_nurse(db, current_user) or getattr(current_user, 'is_master_admin', False)):
        raise Exception("Permission denied")

    # 그룹 스코프: 조회(LIST)와 동일하게 토큰 group_id 기준으로 group_access 모듈에서 해석+권한검증.
    # (HN 그룹 전환 시 home 강제 필터로 근무코드 미발견 → None 접근 크래시 나던 버그 회귀 방지)
    target_group_id = resolve_effective_group(db, current_user, override_group_id or current_user.group_id)
    existing_shift = db.query(Shift).filter(Shift.id == req.id, Shift.group_id == target_group_id).first()
    if not existing_shift:
        raise Exception("해당 근무코드를 찾을 수 없습니다.")
    # off_swap_target 정책 검증 — 변경 후 (type, off_swap_target) 조합 기준.
    _new_target = (
        bool(req.off_swap_target)
        if (hasattr(req, "off_swap_target") and req.off_swap_target is not None)
        else bool(existing_shift.off_swap_target or False)
    )
    _assert_off_swap_target_valid(
        db,
        group_id=target_group_id,
        target_value=_new_target,
        shift_type=req.type,
        self_id=existing_shift.id,
    )
    old_shift_id = existing_shift.shift_id
    old_shift_gb = getattr(existing_shift, "shift_gb", None)
    existing_shift.shift_id = req.shift_id
    existing_shift.name = req.name
    existing_shift.color = req.color
    existing_shift.start_time = req.start_time
    existing_shift.end_time = req.end_time
    existing_shift.type = req.type
    existing_shift.duration = req.duration
    existing_shift.allday = req.allday
    existing_shift.auto_schedule = req.auto_schedule
    existing_shift.shift_gb = req.shift_gb
    # 원티드 페이지 노출 여부 업데이트
    if hasattr(req, "show_in_preference") and req.show_in_preference is not None:
        existing_shift.show_in_preference = req.show_in_preference
    # 초과 OFF 변환 타깃 업데이트 (None 이면 기존 값 유지)
    if hasattr(req, "off_swap_target") and req.off_swap_target is not None:
        existing_shift.off_swap_target = bool(req.off_swap_target)
    # 근무코드 설명 업데이트 — 프론트가 빈 값을 null 로 전송하므로 클리어 허용(항상 반영).
    existing_shift.description = getattr(req, "description", None)
    db.commit()
    db.refresh(existing_shift)
    _append_shift_manage_code(
        db=db,
        office_id=existing_shift.office_id,
        group_id=target_group_id,
        shift_id=existing_shift.shift_id,
        shift_gb=req.shift_gb,
        old_shift_id=old_shift_id,
        old_shift_gb=old_shift_gb,
    )
    return {
        "message": "근무코드가 성공적으로 수정되었습니다.",
        "shift": {
            "shift_id": existing_shift.shift_id,
            "name": existing_shift.name,
            "color": existing_shift.color,
            "sequence": existing_shift.sequence,
            "shift_gb": existing_shift.shift_gb,
            "description": existing_shift.description,
        }
    }


def remove_shift_service(req, current_user, db, override_group_id: str | None = None):
    """
    시프트 삭제 서비스 함수
    """
    if not current_user or not (caller_is_head_nurse(db, current_user) or getattr(current_user, 'is_master_admin', False)):
        raise Exception("Permission denied")
    # 그룹 스코프: 조회(LIST)와 동일하게 토큰 group_id 기준으로 group_access 모듈에서 해석+권한검증.
    target_group_id = resolve_effective_group(db, current_user, override_group_id or current_user.group_id)
    existing_shift = db.query(Shift).filter(Shift.shift_id == req.shift_id, Shift.group_id == target_group_id).first()
    if not existing_shift:
        raise Exception("해당 근무코드를 찾을 수 없습니다.")
    schedule_entries_count = db.query(ScheduleEntry).filter(
        ScheduleEntry.shift_id == req.shift_id
    ).count()
    if schedule_entries_count > 0:
        raise Exception("해당 근무코드는 현재 사용 중이므로 삭제할 수 없습니다.")
    deleted_sequence = existing_shift.sequence
    # 근무코드 삭제 시 shift_manage.codes 에 남는 orphan code 를 정리한다(같은 트랜잭션 커밋).
    _remove_shift_manage_code(
        db,
        existing_shift.office_id,
        target_group_id,
        existing_shift.shift_id,
        getattr(existing_shift, "shift_gb", None),
    )
    db.delete(existing_shift)
    db.query(Shift).filter(Shift.group_id == target_group_id, Shift.sequence > deleted_sequence).update({"sequence": Shift.sequence - 1})
    db.commit()
    return {"message": "근무코드가 성공적으로 삭제되었습니다."}


def move_shift_service(req, current_user, db, override_group_id: str | None = None):
    """
    시프트 순서 이동 서비스 함수
    """
    if not current_user or not (caller_is_head_nurse(db, current_user) or getattr(current_user, 'is_master_admin', False)):
        raise Exception("Permission denied")
    # 그룹 스코프: 조회(LIST)와 동일하게 토큰 group_id 기준으로 group_access 모듈에서 해석+권한검증.
    target_group_id = resolve_effective_group(db, current_user, override_group_id or current_user.group_id)
    shift_to_move = db.query(Shift).filter(Shift.shift_id == req.shift_id, Shift.group_id == target_group_id).first()
    if not shift_to_move:
        raise Exception("해당 근무코드를 찾을 수 없습니다.")
    old_sequence = shift_to_move.sequence
    new_sequence = req.new_sequence
    if old_sequence == new_sequence:
        return {"message": "변경사항이 없습니다."}
    if old_sequence < new_sequence:
        db.query(Shift).filter(Shift.group_id == target_group_id, Shift.sequence > old_sequence, Shift.sequence <= new_sequence).update({"sequence": Shift.sequence - 1})
    else:
        db.query(Shift).filter(Shift.group_id == target_group_id, Shift.sequence >= new_sequence, Shift.sequence < old_sequence).update({"sequence": Shift.sequence + 1})
    shift_to_move.sequence = new_sequence
    db.commit()
    return {"message": "근무코드 순서가 성공적으로 변경되었습니다."}


# ──────────────────────────────────────────────────────────────
# 엑셀 일괄 업로드: 템플릿 / 검증 / 확정
# ──────────────────────────────────────────────────────────────

def create_shift_template() -> str:
    """근무코드 엑셀 업로드 템플릿 생성. 임시파일 경로 반환."""
    template_data = {
        "근무코드(필수)":     ["D1", "E1", "N1", "OFF"],
        "근무코드명(필수)":   ["Day근무", "Evening근무", "Night근무", "휴무"],
        "근무구분(선택)":     ["데이", "이브닝", "나이트", ""],
        "근무유형(필수)":     ["근무", "근무", "근무", "휴무"],
        "시작시간(선택)":     ["06:00", "14:00", "22:00", ""],
        "종료시간(선택)":     ["14:00", "22:00", "06:00", ""],
        "종일여부(선택)":     ["N", "N", "N", "Y"],
        "자동스케줄(선택)":   ["Y", "Y", "Y", "Y"],
        "근무시간(선택)":     ["", "", "", ""],
        "주휴여부(선택)":     ["N", "N", "N", "N"],
        "희망근무표시(선택)": ["N", "N", "N", "N"],
    }

    df = pd.DataFrame(template_data)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        path = tmp.name

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="근무코드", index=False)
        ws = writer.sheets["근무코드"]
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="CCCCCC")
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

    return path


def shift_upload_validate(
    file_path: str,
    user: UserSchema,
    db: Session,
    group_id: str,
) -> Dict[str, Any]:
    """
    근무코드 엑셀 업로드 검증.
    Returns: {success, errors, rows, summary}
    """
    df = pd.read_excel(file_path, sheet_name=0)
    df = df.dropna(how="all")
    if len(df) > 500:
        raise ValueError("최대 500행까지만 업로드 가능합니다.")

    # ── 컬럼 매칭 ─────────────────────────────────────────────
    def find_col(candidates: List[str]) -> str:
        for c in df.columns:
            if str(c).strip() in candidates:
                return c
        raise ValueError(f"필수 컬럼 누락: {candidates}")

    def find_col_optional(candidates: List[str]) -> Optional[str]:
        for c in df.columns:
            if str(c).strip() in candidates:
                return c
        return None

    col_shift_id = find_col(["근무코드(필수)", "근무코드", "shift_id"])
    col_name     = find_col(["근무코드명(필수)", "근무코드명", "이름", "name"])
    col_type     = find_col(["근무유형(필수)", "근무유형", "type"])
    col_shift_gb = find_col_optional(["근무구분(선택)", "근무구분", "shift_gb"])
    col_start    = find_col_optional(["시작시간(선택)", "시작시간", "start_time"])
    col_end      = find_col_optional(["종료시간(선택)", "종료시간", "end_time"])
    col_allday   = find_col_optional(["종일여부(선택)", "종일여부", "allday"])
    col_auto     = find_col_optional(["자동스케줄(선택)", "자동스케줄", "auto_schedule"])
    col_duration = find_col_optional(["근무시간(선택)", "근무시간", "duration"])
    col_weekly   = find_col_optional(["주휴여부(선택)", "주휴여부", "is_weekly_off"])
    col_pref     = find_col_optional(["희망근무표시(선택)", "희망근무표시", "show_in_preference"])

    # ── 기존 데이터 조회 ──────────────────────────────────────
    existing_shift_ids: Set[str] = set(
        row[0] for row in db.query(Shift.shift_id).filter(Shift.group_id == group_id).all()
    )
    existing_colors: Set[str] = set(
        row[0] for row in db.query(Shift.color).filter(Shift.group_id == group_id).all()
    )

    normalized: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    shift_ids_in_file: Set[str] = set()
    color_counter: Dict[str, int] = {}
    TIME_RE = re.compile(r"^\d{2}:\d{2}$")

    for ridx, row in df.iterrows():
        if _is_guide_row(row):
            continue

        row_errs: List[str] = []

        # --- shift_id ---
        shift_id = _safe_str(row.get(col_shift_id))
        if not shift_id:
            row_errs.append("근무코드 누락")
        elif len(shift_id) > 10:
            row_errs.append("근무코드는 최대 10자까지 가능합니다.")
        elif shift_id in shift_ids_in_file:
            row_errs.append("엑셀 내 중복 근무코드")
        elif shift_id in existing_shift_ids:
            row_errs.append(f"이미 존재하는 근무코드: {shift_id}")

        # --- name ---
        name = _safe_str(row.get(col_name))
        if not name:
            row_errs.append("근무코드명 누락")
        elif len(name) > 20:
            row_errs.append("근무코드명은 최대 20자까지 가능합니다.")

        # --- shift_gb (데이/이브닝/나이트/고정 또는 빈값) ---
        shift_gb = _safe_str(row.get(col_shift_gb)) if col_shift_gb else ""
        shift_gb = shift_gb if shift_gb else None
        if shift_gb and shift_gb not in VALID_SHIFT_GB:
            row_errs.append("근무구분은 데이, 이브닝, 나이트, 미드, 고정 중 하나여야 합니다.")

        # --- type ---
        shift_type = _safe_str(row.get(col_type))
        if not shift_type:
            row_errs.append("근무유형 누락")
        elif shift_type not in ("근무", "휴무"):
            row_errs.append("근무유형은 '근무' 또는 '휴무'여야 합니다.")

        # --- start_time / end_time ---
        start_time_str: Optional[str] = None
        end_time_str: Optional[str] = None
        for label, col_t, target in [("시작시간", col_start, "start"), ("종료시간", col_end, "end")]:
            if col_t is None:
                continue
            val = _safe_str(row.get(col_t))
            if not val:
                continue
            if not TIME_RE.match(val):
                row_errs.append(f"{label} 형식 오류 (HH:MM)")
            else:
                h, m = int(val[:2]), int(val[3:5])
                if h > 23 or m > 59:
                    row_errs.append(f"{label} 범위 오류")
                elif target == "start":
                    start_time_str = val
                else:
                    end_time_str = val

        # --- start_time / end_time 기본값 (미입력 시 00:00 / 23:59) ---
        if not start_time_str:
            start_time_str = "00:00"
        if not end_time_str:
            end_time_str = "23:59"

        # --- Y/N 필드 ---
        allday_val = 1 if (_parse_boolean(row.get(col_allday)) if col_allday else False) else 0

        if col_auto:
            raw_auto = _safe_str(row.get(col_auto))
            auto_val = (0 if raw_auto and not _parse_boolean(raw_auto) else 1)
        else:
            auto_val = 1

        weekly_val = 1 if (_parse_boolean(row.get(col_weekly)) if col_weekly else False) else 0
        pref_val = 1 if (_parse_boolean(row.get(col_pref)) if col_pref else False) else 0

        # --- duration ---
        duration_val: Optional[int] = None
        if col_duration:
            raw_dur = row.get(col_duration)
            if raw_dur is not None and not (isinstance(raw_dur, float) and math.isnan(raw_dur)):
                s = str(raw_dur).strip()
                if s:
                    try:
                        duration_val = int(float(s))
                    except (ValueError, TypeError):
                        row_errs.append("근무시간은 숫자여야 합니다.")

        # --- 색상 자동 배정 ---
        color = assign_color(shift_gb, shift_type or "근무", existing_colors, color_counter)
        existing_colors.add(color)

        # --- 결과 수집 ---
        if shift_id and shift_id not in shift_ids_in_file:
            shift_ids_in_file.add(shift_id)

        if not row_errs:
            normalized.append({
                "row": int(ridx) + 2,
                "shift_id": shift_id,
                "name": name,
                "shift_gb": shift_gb,
                "type": shift_type,
                "start_time": start_time_str,
                "end_time": end_time_str,
                "allday": allday_val,
                "auto_schedule": auto_val,
                "duration": duration_val,
                "is_weekly_off": weekly_val,
                "show_in_preference": pref_val,
                "color": color,
            })

        if row_errs:
            errors.append({"row": int(ridx) + 2, "reason": " | ".join(row_errs)})

    return {
        "success": 0 if errors else len(normalized),
        "errors": errors,
        "rows": normalized,
        "summary": {
            "total": len(normalized),
            "error_count": len(errors),
        },
    }


def shift_upload_confirm(
    rows: List[Dict[str, Any]],
    user: UserSchema,
    db: Session,
    target_group_id: str,
) -> Dict[str, Any]:
    """검증 통과된 근무코드를 DB 에 저장."""
    if not target_group_id:
        return {"success": 0, "errors": [{"row": 0, "reason": "group_id가 필요합니다."}]}

    # office_id 결정
    if getattr(user, "is_master_admin", False):
        g = db.query(Group).filter(Group.group_id == target_group_id).first()
        if not g:
            return {"success": 0, "errors": [{"row": 0, "reason": "그룹을 찾을 수 없습니다."}]}
        office_id = g.office_id
    else:
        nurse = db.query(Nurse).filter(Nurse.nurse_id == user.nurse_id).first()
        if nurse and nurse.group:
            office_id = nurse.group.office_id
        else:
            office_id = getattr(user, "office_id", None)

    if not office_id:
        return {"success": 0, "errors": [{"row": 0, "reason": "office_id를 결정할 수 없습니다."}]}

    max_seq = (
        db.query(func.max(Shift.sequence))
        .filter(Shift.group_id == target_group_id)
        .scalar()
        or 0
    )

    saved = 0
    errors: List[Dict[str, Any]] = []
    saved_items: List[Dict[str, Any]] = []

    for idx, item in enumerate(rows, 1):
        try:
            shift_id = str(item.get("shift_id", "")).strip()
            if not shift_id:
                errors.append({"row": item.get("row", 0), "reason": "shift_id 누락"})
                continue

            max_seq += 1

            new_shift = Shift(
                shift_id=shift_id,
                office_id=office_id,
                group_id=target_group_id,
                name=str(item.get("name", "")).strip(),
                color=str(item.get("color", "#CF847A")).strip(),
                shift_gb=item.get("shift_gb") or None,
                start_time=_parse_time_str(item.get("start_time")),
                end_time=_parse_time_str(item.get("end_time")),
                type=str(item.get("type", "근무")).strip(),
                allday=item.get("allday", 0),
                auto_schedule=item.get("auto_schedule", 1),
                duration=item.get("duration"),
                sequence=max_seq,
                default_shift=None,
                is_weekly_off=item.get("is_weekly_off", 0),
                show_in_preference=item.get("show_in_preference", False),
            )
            db.add(new_shift)
            saved += 1
            saved_items.append(item)

        except Exception as e:
            errors.append({"row": item.get("row", 0), "reason": str(e)})

    db.commit()

    # ShiftManage codes 업데이트 (데이/이브닝/나이트 → D/E/N 변환)
    for item in saved_items:
        sg = item.get("shift_gb")
        slot_code = SHIFT_GB_TO_SLOT_CODE.get(sg)
        if slot_code:
            _append_shift_manage_code(
                db=db,
                office_id=office_id,
                group_id=target_group_id,
                shift_id=item["shift_id"],
                shift_gb=slot_code,
            )

    return {"success": saved, "saved": saved, "errors": errors}


# ──────────────────────────────────────────────────────────────
# 타 병동 근무코드 가져오기
# ──────────────────────────────────────────────────────────────

def get_available_shifts_for_import(
    office_id: str,
    target_group_id: str,
    db: Session,
) -> List[Dict[str, Any]]:
    """동일 오피스 내 다른 그룹의 근무코드 중 현재 그룹에 없는 것만 반환."""
    other_group_ids = [
        g.group_id for g in
        db.query(Group.group_id)
        .filter(Group.office_id == office_id, Group.group_id != target_group_id)
        .all()
    ]
    if not other_group_ids:
        return []

    existing_shift_ids: Set[str] = set(
        row[0] for row in
        db.query(Shift.shift_id).filter(Shift.group_id == target_group_id).all()
    )

    candidates = (
        db.query(Shift, Group.group_name)
        .join(Group, Shift.group_id == Group.group_id)
        .filter(
            Shift.group_id.in_(other_group_ids),
            Shift.office_id == office_id,
        )
        .order_by(Group.group_name, Shift.sequence)
        .all()
    )

    seen: Set[str] = set()
    result: List[Dict[str, Any]] = []
    for shift, group_name in candidates:
        if shift.shift_id in seen or shift.shift_id in existing_shift_ids:
            continue
        seen.add(shift.shift_id)
        result.append({
            "shift_id": shift.shift_id,
            "name": shift.name,
            "color": shift.color,
            "shift_gb": shift.shift_gb,
            "type": shift.type,
            "start_time": shift.start_time.strftime("%H:%M") if shift.start_time else None,
            "end_time": shift.end_time.strftime("%H:%M") if shift.end_time else None,
            "allday": shift.allday,
            "auto_schedule": shift.auto_schedule,
            "duration": shift.duration,
            "is_weekly_off": shift.is_weekly_off,
            "show_in_preference": shift.show_in_preference,
            "source_group_id": shift.group_id,
            "source_group_name": group_name,
        })
    return result


def import_shifts_to_group(
    shift_ids: List[str],
    target_group_id: str,
    office_id: str,
    db: Session,
) -> Dict[str, Any]:
    """선택된 근무코드를 다른 그룹에서 현재 그룹으로 복사."""
    if not shift_ids:
        return {"imported": 0, "skipped": 0, "errors": []}

    existing_ids: Set[str] = set(
        row[0] for row in
        db.query(Shift.shift_id).filter(Shift.group_id == target_group_id).all()
    )
    max_seq = (
        db.query(func.max(Shift.sequence))
        .filter(Shift.group_id == target_group_id)
        .scalar() or 0
    )
    other_group_ids = [
        g.group_id for g in
        db.query(Group.group_id)
        .filter(Group.office_id == office_id, Group.group_id != target_group_id)
        .all()
    ]

    imported = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    saved_items: List[Dict[str, str]] = []

    for sid in shift_ids:
        if sid in existing_ids:
            skipped += 1
            continue

        source = (
            db.query(Shift)
            .filter(
                Shift.shift_id == sid,
                Shift.group_id.in_(other_group_ids),
                Shift.office_id == office_id,
            )
            .first()
        )
        if not source:
            errors.append({"shift_id": sid, "reason": "원본 근무코드를 찾을 수 없습니다."})
            continue

        max_seq += 1

        new_shift = Shift(
            shift_id=sid,
            office_id=office_id,
            group_id=target_group_id,
            name=source.name,
            color=source.color,
            shift_gb=source.shift_gb,
            start_time=source.start_time,
            end_time=source.end_time,
            type=source.type,
            allday=source.allday,
            auto_schedule=source.auto_schedule,
            duration=source.duration,
            sequence=max_seq,
            default_shift=None,
            is_weekly_off=source.is_weekly_off,
            show_in_preference=source.show_in_preference,
        )
        db.add(new_shift)
        existing_ids.add(sid)
        imported += 1
        saved_items.append({"shift_id": sid, "shift_gb": source.shift_gb})

    db.commit()

    for item in saved_items:
        slot_code = SHIFT_GB_TO_SLOT_CODE.get(item["shift_gb"] or "")
        if slot_code:
            _append_shift_manage_code(
                db=db,
                office_id=office_id,
                group_id=target_group_id,
                shift_id=item["shift_id"],
                shift_gb=slot_code,
            )

    return {"imported": imported, "skipped": skipped, "errors": errors}
