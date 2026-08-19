"""icmc 근무표 엑셀 전량을 훑어 보건/감정/수면 휴가 패턴을 뽑는다.

시트 레이아웃이 파일마다 달라 자동 감지한다:
  - 날짜 헤더 행 = 1..28~31 이 연속으로 늘어선 행
  - 이름 열      = 날짜 시작열 바로 왼쪽에서, 근무코드가 5칸 이상 찬 행의 텍스트
DB 를 쓰지 않으므로 N 연번(N1..N15)이 그대로 살아 있다.
"""
from __future__ import annotations

import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date

import openpyxl

SRC_DIR = "/Users/hjj/Downloads/roster_sample/icmc"
WD = "월화수목금토일"

WORK_RE = re.compile(r"^(DE|D|E|N|M)\d*$")
NSEQ_RE = re.compile(r"^N(\d+)$")
LEAVE_KEYS = {
    "보건": "보건",
    "감정": "감정",
    "수면": "수면",
}


def norm(v) -> str:
    """셀 값 정규화. ★ 괄호(타 팀 지원 표기)를 제거한다.

    `N(B)5` → `N5` · `D(B)` → `D`. 괄호를 남기면 연번 정규식·근무코드 정규식이
    전부 빗나가 그 셀이 통째로 누락된다.

    실측(2026-08-03): 이 처리가 없어 41병동 8월 5명의 N 개수가 DB 와 어긋났고,
    권현은은 **N15 도달 자체를 놓쳤다**(N11·N12·N13·N14·N(B)15·N(B)1 → N14 로 오판).
    단순 개수 오차가 아니라 수면OFF 판정이 뒤집히는 문제다.
    """
    if v is None:
        return ""
    s = str(v).strip().replace("\n", "").replace(" ", "")
    return re.sub(r"\([^)]*\)", "", s)


def find_date_header(ws, max_scan: int = 12) -> tuple[int, int, int] | None:
    """(header_row, first_day_col, num_days) 반환. 못 찾으면 None."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            run = 0
            cc = c
            while cc <= ws.max_column:
                v = norm(ws.cell(r, cc).value)
                if not v.isdigit() or int(v) != run + 1:
                    break
                run += 1
                cc += 1
            if run >= 28:
                return r, c, run
    return None


def scan_sheet(path: str, sheet: str, ym: tuple[int, int]) -> list[dict]:
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    hdr = find_date_header(ws)
    if not hdr:
        return []
    _, hdr_col, ndays = hdr
    year, month = ym
    out = []
    for r in range(1, ws.max_row + 1):
        # ★ 헤더의 날짜열과 데이터의 날짜열이 한 칸 어긋난 시트가 있다(41병동 7월).
        #    헤더를 믿지 말고, 이 행에서 근무코드가 가장 많이 잡히는 시작열을 고른다.
        best_dcol, best_works, best_cells = None, -1, []
        for cand in (hdr_col, hdr_col + 1, hdr_col - 1):
            if cand < 1:
                continue
            cc = [norm(ws.cell(r, cand + i).value) for i in range(ndays)]
            w = sum(1 for v in cc if WORK_RE.match(v) or v in ("OFF", "O", "주"))
            if w > best_works:
                best_dcol, best_works, best_cells = cand, w, cc
        dcol, works, cells = best_dcol, best_works, best_cells
        if works < 5:
            continue
        name = ""
        for c in range(dcol - 1, 0, -1):
            v = str(ws.cell(r, c).value or "").strip()
            if v and not v.isdigit() and len(v) <= 12:
                name = v
                break
        if not name or "2026" in name or name in ("8월", "7월", "D", "E", "N", "M", "DE"):
            continue
        rec = {"name": name, "row": r, "cells": cells, "ndays": ndays}
        # 휴가 위치
        for key, label in LEAVE_KEYS.items():
            rec[label] = [i + 1 for i, v in enumerate(cells) if v.startswith(key)]
        # N 연번 위치
        rec["nseq"] = {}
        for i, v in enumerate(cells):
            m = NSEQ_RE.match(v)
            if m:
                rec["nseq"][i + 1] = int(m.group(1))
        rec["n_total"] = sum(1 for v in cells if v == "N" or NSEQ_RE.match(v))
        rec["work_days"] = sum(1 for v in cells if WORK_RE.match(v))
        rec["off_days"] = sum(1 for v in cells if v in ("OFF", "O", "주"))
        rec["is_n_only"] = rec["n_total"] >= 8 and not any(
            v.startswith(("D", "E", "M")) and WORK_RE.match(v) for v in cells
        )
        rec["has_sanjeon"] = any(v.startswith("산전") for v in cells)
        rec["fixed_like"] = _fixed_like(cells)
        rec["year"], rec["month"] = year, month
        out.append(rec)
    return out


def _fixed_like(cells: list[str]) -> str | None:
    """평일 한 코드로 고정된 패턴이면 그 코드를 돌려준다."""
    codes = [v for v in cells if WORK_RE.match(v)]
    if len(codes) < 10:
        return None
    c = Counter(re.sub(r"\d+$", "", v) for v in codes)
    top, n = c.most_common(1)[0]
    return top if n / len(codes) >= 0.9 else None


def main() -> None:
    files = []
    for f in sorted(os.listdir(SRC_DIR)):
        n = unicodedata.normalize("NFC", f)
        if f.startswith("~$") or not f.endswith(".xlsx"):
            continue
        files.append((n, os.path.join(SRC_DIR, f)))

    recs = []
    print("=== 스캔 대상 시트 ===")
    for name, path in files:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sn in wb.sheetnames:
            low = sn.replace(" ", "")
            if "원티드" in low or "WANTED" in low.upper() or "신청" in low:
                continue
            ws = wb[sn]
            if not find_date_header(ws):
                continue
            # 월 추정: 시트명 > 파일명
            m = re.search(r"(\d{1,2})월", sn) or re.search(r"(\d{1,2})월", name)
            month = int(m.group(1)) if m else 0
            got = scan_sheet(path, sn, (2026, month))
            if not got:
                continue
            for g in got:
                g["file"] = name[:34]
                g["sheet"] = sn[:26]
            recs.extend(got)
            print(f"  {name[:34]:<36}{sn[:26]:<28}{month:>2}월 · {len(got)}명")
    print(f"\n총 레코드 {len(recs)}개 (인원-월)")

    import json

    out = "/private/tmp/claude-501/-Users-hjj-Downloads-meditong/073f7752-4e9d-4ef0-b3f6-29a4f022cfa0/scratchpad/rs/leave_scan.json"
    json.dump(recs, open(out, "w"), ensure_ascii=False)
    print(f"저장 → {out}")


if __name__ == "__main__":
    main()
