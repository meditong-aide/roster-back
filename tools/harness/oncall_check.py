"""수술실 온콜 검증 하네스 — 생성 결과를 실측 패턴 규칙에 대조한다.

규칙은 `tools/harness/rules/oncall_core.yaml` 에 있고, 근거는 전부
2026-07/08 인천의료원 수술실 근무일정표 실측이다.

    uv run python tools/harness/oncall_check.py --group 102243f1d943 --year 2026 --month 9
    uv run python tools/harness/oncall_check.py ... --reference-xlsx <8월 엑셀> --reference-month 8

DB 는 **읽기만** 한다. 생성은 하지 않으므로 돌리기 전에 근무표가 있어야 한다.
결과는 `tools/harness/reports/oncall-<group>-<YYYYMM>-<stamp>.json` 에 남는다.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "app"))

from db.client2 import SessionLocal  # noqa: E402
from db.models import (  # noqa: E402
    FixedWantedEntry, Nurse, NurseAssignment, Schedule, ScheduleEntry,
)
from services.oncall_assign import (  # noqa: E402
    MAX_CALL_RUN, MAX_SUB_RUN, load_call_code_map, team_of_week, week_start,
)
from services.oncall_postprocess import (  # noqa: E501
    _call_banned_eves,  # noqa: E402
    _learn_call_quota, _members_asof, _pick_schedule, _resolve_start_team,
)

try:
    import yaml
except ImportError:
    yaml = None


def _runs(days: list[int]) -> list[list[int]]:
    """연속 구간으로 묶는다. `[1,2,4]` → `[[1,2],[4]]`"""
    days = sorted(days)
    if not days:
        return []
    out, cur = [], [days[0]]
    for d in days[1:]:
        if d == cur[-1] + 1:
            cur.append(d)
        else:
            out.append(cur)
            cur = [d]
    out.append(cur)
    return out


class Report:
    def __init__(self, rules: dict):
        self.rules = {c["id"]: c for c in (rules.get("checks") or [])}
        self.rows: list[dict] = []

    def add(self, cid: str, ok: bool, detail: Any = "", note: str = "") -> None:
        meta = self.rules.get(cid, {})
        self.rows.append({
            "id": cid,
            "title": meta.get("title", cid) + (f" — {note}" if note else ""),
            "severity": meta.get("severity", "warning"),
            "status": "PASS" if ok else "FAIL",
            "detail": detail if isinstance(detail, (str, int, float)) else str(detail),
        })

    @property
    def failed(self) -> list[dict]:
        return [r for r in self.rows if r["status"] == "FAIL"]

    @property
    def blocking_failed(self) -> list[dict]:
        return [r for r in self.failed if r["severity"] == "blocking"]


def _load_month(db, gid: str, year: int, month: int, calls: set[str]):
    """그 달 최신 근무표와 파생 지표. 없으면 None."""
    sch = db.query(Schedule).filter(
        Schedule.group_id == gid, Schedule.year == year, Schedule.month == month
    ).order_by(Schedule.created_at.desc()).first()
    if sch is None:
        return None
    es = db.query(ScheduleEntry).filter(
        ScheduleEntry.schedule_id == sch.schedule_id).all()
    if not es:
        return None
    got = {(str(e.nurse_id), e.work_date.day): str(e.shift_id) for e in es}
    byday: dict[int, list[str]] = defaultdict(list)
    for e in es:
        if str(e.shift_id) in calls:
            byday[e.work_date.day].append(str(e.nurse_id))
    return sch, es, got, byday


def run(gid: str, year: int, month: int, rules: dict,
        ref_xlsx: str | None, ref_month: int | None) -> Report:
    rep = Report(rules)
    db = SessionLocal()
    try:
        cm = load_call_code_map(db, gid)
        calls, base = set(cm.values()), set(cm)
        nm = {str(n.nurse_id): n.name for n in
              db.query(Nurse).filter(Nurse.group_id == gid).all()}
        act = {str(n.nurse_id) for n in db.query(Nurse).filter(
            Nurse.group_id == gid, Nurse.active == 1).all()}

        loaded = _load_month(db, gid, year, month, calls)
        if loaded is None:
            rep.add("headcount", False, f"{year}-{month:02d} 근무표 없음")
            return rep
        sch, es, got, byday = loaded
        import calendar as _cal
        days = _cal.monthrange(year, month)[1]

        # ── 구조 ──
        ppl = {str(e.nurse_id) for e in es}
        rep.add("headcount", ppl == act, f"{len(ppl)}/{len(act)}명")
        rep.add("cell_count", len(es) == len(ppl) * days, f"{len(es)}셀")
        cnt = Counter(len(v) for v in byday.values())
        want = (rules.get("checks") and
                next((c.get("expect") for c in rules["checks"]
                      if c["id"] == "daily_headcount"), 4)) or 4
        rep.add("daily_headcount", list(cnt) == [want] and len(byday) == days,
                f"{dict(cnt)} · {len(byday)}/{days}일")

        # ── 윈도우 ──
        start = _resolve_start_team(db, gid, year, month, calls)
        if start is None:
            rep.add("chain_from_prev", False, "직전 달 콜 없음 — 역산 불가")
            rep.add("team_rotation", False, "판정 불가")
            tm_rank = {}
            tm, exp = {}, {}
        else:
            anchor, st = start
            members = _members_asof(db, gid, max(anchor, date(year, month, 1)))
            teams = sorted({m.team_id for m in members})
            order = teams[teams.index(st):] + teams[:teams.index(st)]
            tm = {m.nurse_id: m.team_id for m in members}
            tm_rank = {m.nurse_id: m.rank for m in members}
            exp = {d: team_of_week(week_start(date(year, month, d)), anchor, order)
                   for d in range(1, days + 1)}
            # ★ 하루 단위로 최빈 팀을 보면 대체가 2명 든 날에서 2:2 동률이 나고,
            #   Counter 가 삽입 순서로 타 팀을 골라 오탐한다(실측 2026-09/10).
            #   규칙 자체가 "주 단위 로테이션" 이므로 주로 묶어 판정한다.
            byweek: dict = {}
            for d, v in byday.items():
                byweek.setdefault(week_start(date(year, month, d)), Counter()).update(
                    tm.get(n) for n in v)
            ok = all(c[exp[w.day if w.month == month else 1]] == max(c.values())
                     for w, c in byweek.items())
            rep.add("team_rotation", ok, " · ".join(
                f"{w.strftime('%m/%d')}→t{t}" for w, t in
                sorted({week_start(date(year, month, d)): exp[d] for d in exp}.items())))
            rep.add("chain_from_prev", True, f"anchor={anchor} start=t{st}")

        # ── 코드 ──
        # ★ 공휴일은 평일이어도 OFF 계열이다(fixed_holiday_off_yn 이 켜진 그룹).
        #   그날은 `오프콜` 이 맞으므로 요일만으로 판정하면 오탐이 난다
        #   (2026-09 추석 24~26 에서 실측).
        from services.roster_create_service import (
            _holiday_off_enabled, _kr_holidays_in_month,
        )
        hol = _kr_holidays_in_month(year, month) if _holiday_off_enabled(db, gid) else set()
        work_code, off_code = cm.get("D1", "D1콜"), cm.get("O", "오프콜")
        bad_code = [
            (nm.get(n, n), d, got[(n, d)])
            for d, v in byday.items() for n in v
            if got[(n, d)] != (off_code
                               if (date(year, month, d).weekday() >= 5 or d in hol)
                               else work_code)
        ]
        rep.add("code_switch", not bad_code,
                f"{dict(Counter(v for v in got.values() if v in calls))}"
                + (f" · 공휴일 {sorted(hol)}" if hol else "")
                + (f" · 불일치 {bad_code[:4]}" if bad_code else ""))

        # ── 대체 ──
        subs: dict[str, list[int]] = defaultdict(list)
        allc: dict[str, list[int]] = defaultdict(list)
        for d, v in byday.items():
            for n in v:
                allc[n].append(d)
                if exp and tm.get(n) != exp[d]:
                    subs[n].append(d)
        # ★ 3+1 로 갈라 고립 1일을 만드는 대신 통째로 주므로 상한은 MAX_SUB_RUN+1.
        over = [(nm.get(n, n), len(r)) for n, ds in subs.items()
                for r in _runs(ds) if len(r) > MAX_SUB_RUN + 1]
        rep.add("sub_run_limit", not over, " · ".join(
            f"{nm.get(n, n)}{[f'{r[0]}~{r[-1]}' if len(r) > 1 else r[0] for r in _runs(ds)]}"
            for n, ds in subs.items()) or "대체 없음")

        # ── 블록 분할로 생긴 고립 1일 금지 ──
        #   실측 2026-08 의 1일 블록은 김영민 8/10 하나이고 담당주에 이어 붙는다.
        #   ★ 다만 **결원 자체가 1일**이면 대체도 1일일 수밖에 없다(실측 2026-09
        #     윤보라 9/27 — 한승윤이 휴가 직전일 제약으로 하루만 빠진 자리).
        #     그건 불가피하므로, **같은 역할 자리가 이틀 이상 비었는데 1일로 쪼개진
        #     경우**만 잡는다 — 그게 분할 실패다.
        own_days_by = {n: {d for d in range(1, days + 1) if exp and exp[d] == t}
                       for n, t in tm.items()} if exp else {}
        # 역할별로 "대체가 들어간 날"을 모아 연속 구간을 만든다 = 그 자리의 결원 구간
        by_rank: dict = {}
        for n, ds in subs.items():
            by_rank.setdefault(tm_rank.get(n), set()).update(ds)
        seat_runs = {rk_: _runs(sorted(v)) for rk_, v in by_rank.items()}
        isolated = []
        for n, ds in subs.items():
            for r in _runs(ds):
                if len(r) != 1:
                    continue
                if own_days_by.get(n, set()) & {r[0] - 1, r[0] + 1}:
                    continue          # 담당주에 이어 붙는 꼬리물기 — 정상
                seat = next((sr for sr in seat_runs.get(tm_rank.get(n), [])
                             if r[0] in sr), [r[0]])
                if len(seat) <= 1:
                    continue          # 그 자리의 결원 자체가 1일 — 불가피
                isolated.append((nm.get(n, n), r[0], f"자리 결원 {len(seat)}일"))
        rep.add("no_isolated_single", not isolated,
                f"분할 고립 1일 {len(isolated)}건 {isolated or ''}")
        longest = max((max(len(r) for r in _runs(ds)) for ds in allc.values()), default=0)
        rep.add("total_run_limit", longest <= MAX_CALL_RUN, f"최장 {longest}일")

        # ── 역할(rank) 1·2·3·4 각 1명 ──
        dup = []
        for d, v in sorted(byday.items()):
            rs = sorted(tm_rank.get(n, 0) for n in v)
            if rs != [1, 2, 3, 4]:
                dup.append((d, rs))
        rep.add("role_coverage", not dup,
                f"{len(byday) - len(dup)}/{len(byday)}일 역할 1·2·3·4 각 1명"
                + (f" · 중복일 {dup}" if dup else ""))

        # ── 확정 원티드 ──
        fw = db.query(FixedWantedEntry).filter(
            FixedWantedEntry.group_id == gid, FixedWantedEntry.year == year,
            FixedWantedEntry.month == month,
            FixedWantedEntry.is_applied == True).all()  # noqa: E712
        mism = [(nm.get(str(f.nurse_id)), f.shift_date.day) for f in fw
                if got.get((str(f.nurse_id), f.shift_date.day)) != str(f.shift_id)]
        rep.add("fixed_wanted_not_overwritten", not mism,
                f"{len(fw)}건 중 불일치 {len(mism)} {mism or ''}")
        rep.add("fixed_wanted_applied", not mism, f"{len(fw) - len(mism)}/{len(fw)}")

        # ── 콜 = 야간 대기 → 휴가/공가 직전일 금지(CP-SAT ban_night_before_fixed_off 와 동일) ──
        eves = _call_banned_eves(db, gid, year, month)
        viol = [(nm.get(n, n), d) for n, ds in eves.items() for d in ds
                if got.get((n, d)) in calls]
        rep.add("no_call_before_wanted_off", not viol,
                f"금지 {sum(len(v) for v in eves.values())}일 중 위반 {len(viol)} {viol or ''}")

        # ── 제외·보호 ──
        noteam = {n for n in act if n not in tm} if tm else set()
        rep.add("unassigned_excluded", not (noteam & set(allc)),
                f"미배정 {sorted(nm.get(n, n) for n in noteam)}")
        lv = [a for a in db.query(NurseAssignment).filter(
            NurseAssignment.nurse_id.in_(list(act) or [""]),
            NurseAssignment.kind == "leave").all()
            if str(a.status or "") != "cancelled"
            and str(a.source_group_id or "") == gid
            and a.start_date and a.start_date <= date(year, month, days)
            and (a.end_date or a.expected_end_date or date(year, month, days))
            >= date(year, month, 1)]
        lv_ok = all(got.get((str(a.nurse_id), d)) not in (base | calls)
                    for a in lv for d in range(1, days + 1)
                    if a.start_date <= date(year, month, d))
        rep.add("leave_masked", lv_ok, " · ".join(
            f"{nm.get(str(a.nurse_id))}"
            f"{dict(Counter(got.get((str(a.nurse_id), d)) for d in range(1, days + 1)))}"
            for a in lv) or "휴직자 없음")

        # ── 학습 ──
        q = _learn_call_quota(db, gid, year, month, cm)
        rep.add("quota_learned", True,
                ", ".join(f"{nm.get(k, k)} {v}회" for k, v in q.items()) or "없음")
        bad_q = [(nm.get(k), len(allc.get(k, [])), v) for k, v in q.items()
                 if len(allc.get(k, [])) > v]
        rep.add("quota_respected", not bad_q,
                " · ".join(f"{n} {a}회/상한 {c}" for n, a, c in
                           [(nm.get(k), len(allc.get(k, [])), v) for k, v in q.items()])
                or "학습 없음")
        prev_y, prev_m = (year - 1, 12) if month == 1 else (year, month - 1)
        picked = _pick_schedule(db, gid, prev_y, prev_m)
        n_sched = db.query(Schedule).filter(
            Schedule.group_id == gid, Schedule.year == prev_y,
            Schedule.month == prev_m).count()
        rep.add("single_schedule_per_month", picked is not None,
                f"직전달 {n_sched}건 중 대표 {picked}")

        # ── 미등록 그룹 무영향 ──
        others = db.query(Nurse.group_id).filter(Nurse.group_id != gid).distinct().limit(5).all()
        clean = all(not load_call_code_map(db, g[0]) for g in others)
        rep.add("unregistered_group_untouched", clean,
                f"표본 {len(others)}개 그룹 call_base_id 미등록")

        # ── 실측 대조 ──
        if ref_xlsx and ref_month:
            _compare_reference(db, rep, gid, year, ref_month, ref_xlsx, nm, calls)
    finally:
        db.close()
    return rep


def _compare_reference(db, rep: Report, gid: str, year: int, month: int,
                       xlsx: str, nm: dict, calls: set[str]) -> None:
    """엑셀 실측과 생성분을 대조한다."""
    import openpyxl

    sch = db.query(Schedule).filter(
        Schedule.group_id == gid, Schedule.year == year, Schedule.month == month,
        Schedule.status == "draft").order_by(Schedule.created_at.desc()).first()
    ref_id = _pick_schedule(db, gid, year, month)
    if sch is None or ref_id is None:
        rep.add("reproduce_calls", False, f"{year}-{month:02d} 비교 대상 없음")
        return
    got = {(str(e.nurse_id), e.work_date.day): str(e.shift_id)
           for e in db.query(ScheduleEntry).filter(
               ScheduleEntry.schedule_id == sch.schedule_id).all()}
    byname = {v: k for k, v in nm.items()}
    ws = openpyxl.load_workbook(xlsx, data_only=True).worksheets[0]
    dcol = {int(ws.cell(2, c).value): c for c in range(1, ws.max_column + 1)
            if isinstance(ws.cell(2, c).value, (int, float))
            and 1 <= int(ws.cell(2, c).value) <= 31}
    truth = {}
    for r in range(3, ws.max_row + 1):
        v = ws.cell(r, 2).value
        n = " ".join(str(v).split()) if v else ""
        nid = byname.get(n)
        if not nid:
            continue
        for d, c in dcol.items():
            x = ws.cell(r, c).value
            if x is not None and "콜" in str(x):
                truth[(nid, d)] = "D1콜" if str(x).strip().startswith("D1") else "오프콜"
    mine = {k: v for k, v in got.items() if v in calls}
    exact = [k for k in (set(truth) & set(mine)) if truth[k] == mine[k]]
    rate = len(exact) / max(1, len(truth)) * 100
    th = next((c.get("threshold", 0) for c in rep.rules.values()
               if c.get("title", "").startswith("실측 대비 콜")), 0)
    rep.add("reproduce_calls", rate >= (th or 0), f"{rate:.1f}% ({len(exact)}/{len(truth)})",
            note=f"{year}-{month:02d}")
    ref = {(str(e.nurse_id), e.work_date.day): str(e.shift_id)
           for e in db.query(ScheduleEntry).filter(
               ScheduleEntry.schedule_id == ref_id).all() if str(e.nurse_id) in nm}
    both = set(ref) & set(got)
    same = [k for k in both if ref[k] == got[k]]
    cr = len(same) / max(1, len(both)) * 100
    rep.add("reproduce_cells", cr >= 90.0, f"{cr:.1f}% ({len(same)}/{len(both)})",
            note=f"{year}-{month:02d}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--group", required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--rules", default="tools/harness/rules/oncall_core.yaml")
    ap.add_argument("--reference-xlsx", default=None, help="실측 대조용 근무일정표")
    ap.add_argument("--reference-month", type=int, default=None)
    ap.add_argument("--out-dir", default="tools/harness/reports")
    ap.add_argument("--strict", action="store_true",
                    help="blocking FAIL 이 있으면 종료코드 1")
    a = ap.parse_args()

    if yaml is None:
        raise SystemExit("pyyaml 이 필요합니다")
    rules = yaml.safe_load(Path(a.rules).read_text(encoding="utf-8"))

    rep = run(a.group, a.year, a.month, rules, a.reference_xlsx, a.reference_month)

    print(f"■ {rules.get('meta', {}).get('target', '')} "
          f"{a.group} {a.year}-{a.month:02d}\n")
    print(f"{'항목':<34}{'판정':>6}  {'등급':<9}상세")
    print("─" * 104)
    for r in rep.rows:
        print(f"{r['title'][:33]:<34}{r['status']:>6}  {r['severity']:<9}{r['detail'][:52]}")
    print("─" * 104)
    n_ok = len(rep.rows) - len(rep.failed)
    tail = f"  ★ blocking FAIL: {[r['id'] for r in rep.blocking_failed]}" \
        if rep.blocking_failed else "  ✓ blocking 전항목 통과"
    print(f"{n_ok}/{len(rep.rows)} PASS{tail}")

    os.makedirs(a.out_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = Path(a.out_dir) / f"oncall-{a.group}-{a.year}{a.month:02d}-{stamp}.json"
    out.write_text(json.dumps({
        "meta": {"group": a.group, "year": a.year, "month": a.month,
                 "rules": a.rules, "generated_at": stamp},
        "summary": {"total": len(rep.rows), "passed": n_ok,
                    "failed": len(rep.failed),
                    "blocking_failed": len(rep.blocking_failed)},
        "checks": rep.rows,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n리포트: {out}")
    return 1 if (a.strict and rep.blocking_failed) else 0


if __name__ == "__main__":
    raise SystemExit(main())
