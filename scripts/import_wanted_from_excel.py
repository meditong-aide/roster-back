"""병원 원티드(신청) 엑셀 → roster DB import. **dry-run 기본**, `--apply` 로만 write.

## 무엇을 넣는가
간호사 **제출분**으로 넣는다 — `wanted_requests`(is_submitted=1) + `nurse_shift_requests`.
조정판(원티드 관리보드)은 확정본이 없으면 이 원본을 `source_type='original'` 로 보여주고,
수간호사가 검토 후 저장하면 `fixed_wanted_entries` 로 확정된다. 엑셀이 "신청서"라는
성격과 실제 업무 흐름이 일치한다.

## 알림 미발송 (필수 요구)
푸시는 `POST /roster/publish` **라우터 안에서만** 호출된다. 이 스크립트는
라우터를 타지 않고 ORM 으로 직접 write 하며 **푸시 모듈을 import 조차 하지 않는다.**
★ 그룹웨어 push 는 `EUN_DB_NAME=='eun_roster'` 게이트라 운영 실행 시에만 열린다 —
  그래서 이 파일에 푸시 경로를 **절대 넣지 말 것**.

## 함정
- 한글 리터럴을 raw SQL 로 넣으면 `'??'` 로 조용히 깨진다(pymssql charset).
  → 반드시 **ORM 속성 대입**으로 쓴다.
- `nurse_shift_requests.shift` 는 CHAR(1) — 'O','D','E','N' 만 들어간다.
- `wanted_requests.request_id` 는 IDENTITY 가 아니라 복합 PK → (간호사, 월) 기준 채번.

사용:
  uv run python scripts/import_wanted_from_excel.py --file <xlsx> \
      --sheet-index 0 --group-id 102243131010 --year 2026 --month 9 [--apply]
"""
from __future__ import annotations

import argparse
import calendar
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))

import openpyxl  # noqa: E402

from db.client2 import SessionLocal  # noqa: E402
from db.models import (  # noqa: E402
    Nurse, NurseAssignment, NurseShiftRequest, Shift, WantedRequest,
)

#: 엑셀 표기 → shifts.shift_id
CODE_MAP = {"OFF": "O", "O": "O", "D": "D", "E": "E", "N": "N"}
#: 사용자 직접 입력과 같은 우선순위(앱의 `_replace_shift_requests` 와 동일)
SCORE = 10.0
#: 요일 표기 — 이름 열 탐지에서 제외해야 한다(날짜열 아래 행에 들어온다).
_WEEKDAYS = set("월화수목금토일")


def _find_date_row(ws) -> tuple[int, dict[int, int]]:
    """날짜(1~31)가 가장 많이 늘어선 행을 찾는다.

    ★ 병동마다 서식이 다르다 — 51병동은 2행·C열, 41병동은 3행·D열이었다.
      고정값을 쓰면 병동이 바뀔 때마다 스크립트를 고쳐야 하므로 매번 탐지한다.
    """
    best_row, best = 0, {}
    for r in range(1, min(15, ws.max_row) + 1):
        daycol = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= int(v) <= 31:
                daycol[int(v)] = c
        if len(daycol) > len(best):
            best_row, best = r, daycol
    if len(best) < 20:  # 한 달치가 안 되면 날짜 행이 아니다
        raise SystemExit(f"날짜 행을 찾지 못했습니다(최대 {len(best)}개만 발견)")
    return best_row, best


def _find_name_col(ws, date_row: int, first_day_col: int) -> int:
    """날짜열 왼쪽에서 '사람 이름이 가장 많이 들어 있는 열'을 고른다.

    팀 표기('A')·번호·요일이 섞여 있으므로 **개수**로 판별한다.
    """
    scores = {}
    for c in range(1, first_day_col):
        n = 0
        for r in range(date_row + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            # 숫자·요일·한 글자 표기는 이름이 아니다
            if not s or s in _WEEKDAYS or len(s) < 2:
                continue
            try:
                float(s)
                continue
            except ValueError:
                pass
            n += 1
        scores[c] = n
    col = max(scores, key=lambda k: scores[k]) if scores else 0
    if not col or scores[col] < 2:
        raise SystemExit(f"이름 열을 찾지 못했습니다(후보별 개수: {scores})")
    return col


def parse_sheet(path: str, sheet_index: int) -> tuple[str, list[dict], dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet_index]

    date_row, daycol = _find_date_row(ws)
    name_col = _find_name_col(ws, date_row, min(daycol.values()))
    layout = {"date_row": date_row, "name_col": name_col, "days": len(daycol)}

    rows = []
    for r in range(date_row + 1, ws.max_row + 1):
        raw = ws.cell(r, name_col).value
        name = str(raw).strip() if raw else ""
        if not name or name in _WEEKDAYS or len(name) < 2:
            continue
        cells = {}
        for d, c in daycol.items():
            v = ws.cell(r, c).value
            s = str(v).strip() if v is not None else ""
            if s:
                cells[d] = s
        rows.append({"row": r, "name": name, "cells": cells})
    return ws.title, rows, layout


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet-index", type=int, default=0)
    ap.add_argument("--group-id", required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--apply", action="store_true", help="지정해야만 DB 에 쓴다")
    ap.add_argument(
        "--skip-unknown", default="",
        help="명단에 없지만 근무코드가 있는 사람을 **이름을 명시해서만** 건너뛴다(콤마 구분). "
             "미등록 인원의 신청이 통째로 누락되는 일이라, 조용히 빠지지 않도록 개별 지정을 요구한다.")
    ap.add_argument(
        "--alias", action="append", default=[], metavar="엑셀이름=DB이름",
        help="엑셀 표기와 DB 이름이 다를 때 잇는다. 여러 번 지정 가능. "
             "실측(52병동 9월): 엑셀 '김민영b' ↔ DB '김민영'. 병동마다 동명이인 접미사 표기가 "
             "제각각이라 자동 정규화하면 엉뚱한 사람에게 붙을 수 있어 **명시 지정만** 받는다.")
    a = ap.parse_args()
    skip_names = {s.strip().lower() for s in a.skip_unknown.split(",") if s.strip()}
    alias_map: dict[str, str] = {}
    for item in a.alias:
        if "=" not in item:
            raise SystemExit(f"--alias 형식 오류(엑셀이름=DB이름): {item!r}")
        src, dst = item.split("=", 1)
        alias_map[src.strip().lower()] = dst.strip().lower()

    sheet_title, rows, layout = parse_sheet(a.file, a.sheet_index)
    month_str = f"{a.year}-{a.month:02d}"
    db = SessionLocal()
    try:
        print(f"DB={db.get_bind().url.database}  시트={sheet_title!r}  "
              f"group={a.group_id}  {month_str}")
        print(f"[서식] 날짜행={layout['date_row']}  이름열={layout['name_col']}  "
              f"날짜칸={layout['days']}개  (자동 탐지)")

        # ── 같은 이름 여러 행 병합 ─────────────────────────────────────
        # 실측(41병동 9월): 최수진B 가 r28(OFF 5칸) 과 r37(비고 1칸) 두 행에 나뉘어 있었다.
        # 사람 단위로 합쳐야 한쪽이 통째로 버려지지 않는다.
        merged: dict[str, dict] = {}
        for r in rows:
            key = r["name"].lower()
            if key in merged:
                merged[key]["cells"].update(r["cells"])
                merged[key]["rows"].append(r["row"])
            else:
                merged[key] = {"name": r["name"], "cells": dict(r["cells"]),
                               "row": r["row"], "rows": [r["row"]]}
        multi = [v for v in merged.values() if len(v["rows"]) > 1]
        if multi:
            print("\n[병합] 같은 이름이 여러 행에 있어 합칩니다")
            for v in multi:
                print(f"  {v['name']:<9} r{v['rows']} → {len(v['cells'])}칸")
        rows = list(merged.values())

        # ── 셀을 '근무 신청' 과 '메모' 로 가른다 ────────────────────────
        # 근무표 엑셀에는 근무코드 말고 자유 텍스트가 섞인다.
        #   실측: 'E or N' · 'N가능' · '보수'(보수교육) · '자녀 돌봄' · '특별 휴가(모환갑)'
        # ★ 메모를 버리면 수간호사가 그 사정을 못 본다 → wanted_requests.request 에 남긴다.
        # ★ 메모 때문에 사람을 통째로 빼면 조정판에서 **미제출로 잡혀 독촉 대상**이 된다.
        for r in rows:
            r["shift_cells"] = {}
            notes = []
            for d, raw in sorted(r["cells"].items()):
                if CODE_MAP.get(str(raw).strip().upper()):
                    r["shift_cells"][d] = str(raw).strip()
                else:
                    notes.append(f"{d}일: {' '.join(str(raw).split())}")
            r["note"] = " / ".join(notes)

        # ── 명단 매칭 (대소문자 무시) ─────────────────────────────────
        nurses = db.query(Nurse).filter(
            Nurse.group_id == a.group_id, Nurse.active == True  # noqa: E712
        ).all()
        by_lower = {}
        for n in nurses:
            by_lower.setdefault(str(n.name or "").strip().lower(), []).append(n)

        # ── 별칭 검증 ────────────────────────────────────────────────
        # ★ 조용히 붙이면 안 된다. 대상이 없거나 동명이인이면 엉뚱한 사람에게
        #   신청이 통째로 옮겨 붙는다 → 그 자리에서 중단한다.
        if alias_map:
            print("\n[별칭] 엑셀 표기 ↔ DB 이름을 잇습니다")
            excel_lower = {r["name"].lower() for r in rows}
            for src, dst in alias_map.items():
                cand = by_lower.get(dst, [])
                if len(cand) != 1:
                    raise SystemExit(
                        f"별칭 중단: DB '{dst}' 가 {len(cand)}명입니다 (1명이어야 함)")
                if src not in excel_lower:
                    raise SystemExit(f"별칭 중단: 엑셀에 '{src}' 가 없습니다")
                if src in by_lower:
                    raise SystemExit(
                        f"별칭 중단: '{src}' 는 DB 에도 있는 이름입니다 (별칭 불필요·오적용 위험)")
                print(f"  {src} → {dst} (nurse_id={cand[0].nurse_id})")

        matched, unknown, dup, skipped = [], [], [], []
        for r in rows:
            cand = by_lower.get(alias_map.get(r["name"].lower(), r["name"].lower()), [])
            if len(cand) == 1:
                matched.append((r, cand[0]))
            elif len(cand) > 1:
                dup.append(r["name"])
            elif not r["shift_cells"]:
                # 명단에 없고 근무코드도 없다 = 사람 행이 아니다.
                #   실측: '김예담' 육아휴직 30칸 · 'DE' 전 칸 '0'(집계 행)
                skipped.append(r)
            else:
                unknown.append(r)

        if skipped:
            print("\n[제외] 명단에 없고 근무코드도 없는 행 — 휴직 표시·집계 행")
            for r in skipped:
                sample = " ".join(str(next(iter(r["cells"].values()), "")).split())[:30]
                print(f"  r{r['row']:<3} {r['name']:<9} {len(r['cells'])}칸  예: {sample!r}")

        print(f"\n[명단] 엑셀 {len(rows) - len(skipped)}명 / DB active {len(nurses)}명 "
              f"→ 매칭 {len(matched)} · 미매칭 {len(unknown)} · 동명이인 {len(dup)}")
        if dup:
            raise SystemExit(f"동명이인 중단: {dup}")
        # ★ 근무코드를 낸 사람이 명단에 없으면 근무가 통째로 누락되므로 **중단**한다.
        #   --skip-unknown 으로 이름을 명시한 경우만 예외로 넘어간다.
        if unknown:
            waived = [u for u in unknown if u["name"].lower() in skip_names]
            blocking = [u for u in unknown if u["name"].lower() not in skip_names]
            for u in waived:
                cells = ", ".join(f"{d}일 {v}" for d, v in sorted(u["shift_cells"].items()))
                print(f"  · 제외(지정) r{u['row']} {u['name']} — 신청 {len(u['shift_cells'])}건 누락: {cells}")
            for u in blocking:
                print(f"  ! DB 없음: r{u['row']} {u['name']} (근무 {len(u['shift_cells'])}칸)")
            if blocking:
                raise SystemExit(
                    "엑셀에만 있는 근무자가 있어 중단합니다 "
                    "(의도한 것이면 --skip-unknown 에 이름을 지정하세요)")

        # ── 그 달 휴직자 제외 ─────────────────────────────────────────
        # 휴직자는 근무 대상이 아니라 원티드 행을 만들 이유가 없다. 그런데 병원 엑셀은
        # 휴직자 칸에 **누적 일수 카운터**를 적어 둔다 —
        #   실측(52병동 9월): 박민혜 '육아 489'~'육아 518' · 박지수 '육아 275'~'육아 304'
        # 근무코드가 아니라 메모로 잡히므로 "신청 0건 제출 처리 + 메모 30줄" 로 들어가
        # 조정판에 쓰레기가 남는다.
        # ★ 41병동 때 '김예담' 이 걸러진 건 DB 명단에 **없었기** 때문이고, 여기 둘은
        #   active=1 로 살아 있어 그 경로로는 안 걸린다(가림막은 nurse_assignment 쪽).
        # ★ 근무코드를 **한 칸이라도 낸** 휴직자는 제외하지 않는다 — 복직·부분근무일 수 있고,
        #   그 경우 조용히 빼면 실제 신청이 통째로 사라진다.
        m_start = date(a.year, a.month, 1)
        m_end = date(a.year, a.month, calendar.monthrange(a.year, a.month)[1])
        on_leave = {
            lv.nurse_id for lv in db.query(NurseAssignment).filter(
                NurseAssignment.nurse_id.in_([n.nurse_id for _r, n in matched] or [""]),
                NurseAssignment.kind == "leave",
                NurseAssignment.start_date <= m_end,
            ).all()
            if lv.end_date is None or lv.end_date >= m_start
        }
        leave_rows = [(r, n) for r, n in matched
                      if n.nurse_id in on_leave and not r["shift_cells"]]
        if leave_rows:
            print(f"\n[제외] {month_str} 휴직자 — 원티드 대상 아님")
            for r, n in leave_rows:
                print(f"  {n.name:<9} nurse_id={n.nurse_id}  메모 {len(r['cells'])}칸 "
                      f"→ 원티드·메모 모두 쓰지 않음")
            drop = {n.nurse_id for _r, n in leave_rows}
            matched = [(r, n) for r, n in matched if n.nurse_id not in drop]

        notes_rows = [(r, n) for r, n in matched if r["note"]]
        if notes_rows:
            print("\n[메모] 근무코드가 아닌 입력 — 요청이 아니라 메모로 저장합니다")
            for r, _n in notes_rows:
                print(f"  {r['name']:<9} {r['note'][:60]}")

        # ★ 별칭을 여기에도 적용해야 한다. 안 하면 별칭 대상이 '엑셀에 없는 사람' 으로
        #   잘못 뜬다(매칭은 됐는데 미제출자로 보고되는 모순).
        in_excel = {alias_map.get(r["name"].lower(), r["name"].lower()) for r in rows}
        missing = sorted(set(by_lower) - in_excel)
        if missing:
            print(f"  · DB 에만 있는 인원(신청 없음): "
                  f"{[by_lower[m][0].name for m in missing]}")

        # ── shifts 매핑 ─────────────────────────────────────────────
        shifts = {
            s.shift_id: s for s in db.query(Shift).filter(
                Shift.group_id == a.group_id).all()
        }
        bad_codes = set()
        days_in_month = (date(a.year + (a.month == 12), (a.month % 12) + 1, 1)
                         - date(a.year, a.month, 1)).days

        plan, code_count = [], {}
        for r, n in matched:
            for d, raw in sorted(r["shift_cells"].items()):
                code = CODE_MAP.get(raw.upper())
                # 여기까지 온 값은 CODE_MAP 을 통과했다. shifts 에 없다면 그 그룹이
                # 해당 근무코드를 안 쓰는 것이므로 설정 문제 — 중단시킨다.
                if code is None or code not in shifts:
                    bad_codes.add(raw)
                    continue
                if not (1 <= d <= days_in_month):
                    continue
                plan.append((n, date(a.year, a.month, d), code, shifts[code].id))
                code_count[code] = code_count.get(code, 0) + 1
        if bad_codes:
            raise SystemExit(
                f"그룹 shifts 에 없는 근무코드 중단: {sorted(bad_codes)} "
                f"(그룹 보유: {sorted(shifts)})")

        print(f"[신청] {len(plan)}셀  {code_count}")
        per = {}
        for n, _sd, _c, _t in plan:
            per[n.name] = per.get(n.name, 0) + 1
        print("  " + " · ".join(f"{k}:{v}" for k, v in sorted(per.items())))

        # ★ 신청 0건도 **제출**로 넣는다. 엑셀에 이름이 있으면 원티드를 받은
        #   대상이고, 실제로 낼 게 없었을 뿐이다. 빼면 조정판 제출 현황에서
        #   미제출로 잡혀 수간호사가 없는 사람을 독촉하게 된다.
        submitters = [n for _r, n in matched]
        #: 간호사별 메모(근무코드가 아닌 입력) — wanted_requests.request 에 저장한다.
        note_by_nurse = {n.nurse_id: r["note"] for r, n in matched if r["note"]}
        zero = sorted(n.name for n in submitters if n.name not in per)
        if zero:
            print(f"  · 신청 0건이지만 제출 처리: {zero}")

        # ── 기존 데이터 확인 (덮어쓰기 방지) ─────────────────────────
        exist_wr = db.query(WantedRequest).filter(
            WantedRequest.group_id == a.group_id,
            WantedRequest.month == month_str).count()
        if exist_wr:
            raise SystemExit(f"이미 {month_str} wanted_requests {exist_wr}행이 있습니다 — 중단")

        if not a.apply:
            print("\n[dry-run] --apply 를 붙여야 실제로 씁니다.")
            return 0

        # ── write (ORM 만 사용 · 푸시 경로 없음) ─────────────────────
        now = datetime.now()
        made = 0
        for n in {n.nurse_id: n for n in submitters}.values():
            last = db.query(WantedRequest.request_id).filter(
                WantedRequest.nurse_id == n.nurse_id,
                WantedRequest.month == month_str
            ).order_by(WantedRequest.request_id.desc()).first()
            wr = WantedRequest(
                nurse_id=n.nurse_id, request_id=(last[0] + 1) if last else 1,
                month=month_str, group_id=a.group_id,
                request=note_by_nurse.get(n.nurse_id, ""),
                is_submitted=True, created_at=now, submitted_at=now,
            )
            db.add(wr)
            db.flush()
            seq = 0
            for pn, sd, code, tid in plan:
                if pn.nurse_id != n.nurse_id:
                    continue
                seq += 1
                db.add(NurseShiftRequest(
                    nurse_id=n.nurse_id, request_id=wr.request_id,
                    detailed_request_id=seq, shift_date=sd, group_id=a.group_id,
                    shift=code, shifts_table_id=tid, score=SCORE,
                    partial_request="", comment=None,
                ))
                made += 1
        db.commit()
        # ★ plan(신청 있는 사람)이 아니라 submitters 를 센다 — 신청 0건 제출자가
        #   빠져 실제보다 적게 찍힌다(실측: 25행 넣고 21 로 표시).
        print(f"\n[적용] wanted_requests {len({n.nurse_id for n in submitters})}행 · "
              f"nurse_shift_requests {made}행 커밋")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
