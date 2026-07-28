#!/usr/bin/env python
"""근무신청(원티드) 엑셀 → fixed_wanted_entries import.

병원이 종이/엑셀로 받아 둔 근무신청을 확정 원티드로 올린다.
근무표 본체는 `import_finalized_roster.py`, 신청 내역은 이 스크립트가 담당한다.

★ 알림 없음 — 원티드 관련 푸시는 라우터에만 있고 이 스크립트는 ORM 직접 write 다.
   fixed_wanted_entries 에 DB 트리거도 없다.

레이아웃 자동 인식
    근무신청 시트는 이름 열과 날짜 열 사이 간격이 시트마다 다르다
    (7월 간호사=이름4/날짜5, 7·8월 조무사=이름3/날짜4,
     8월 간호사=이름3/날짜6 ← 입사년도·숙련도 두 칸이 끼어 있음).
    그래서 헤더에서 1..말일이 연속으로 놓인 행·열을 직접 찾고,
    이름 열은 그 왼쪽에서 DB 명단과 대조해 정한다.

사용 예
    cd roster-back
    EUN_DB_NAME=eun_roster .venv/bin/python scripts/import_fixed_wanted.py \
        --file "/path/2026년 41병동 AI 근무표.xlsx" \
        --group-id 1022438ea001 --year 2026 --month 8 \
        --sheet "8월 간호사(근무신청) 숙련도 & 참고사항" --apply
"""
from __future__ import annotations

import argparse
import calendar
import os
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

_APP = Path(__file__).resolve().parents[1] / "app"
sys.path.insert(0, str(_APP))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl  # noqa: E402

from db.client2 import SessionLocal  # noqa: E402
from db.models import FixedWantedEntry, Group, Nurse, Shift  # noqa: E402
from import_finalized_roster import (  # noqa: E402
    SHIFT_DEFS,
    cell_text,
    name_key,
    normalize_cell,
    seed_shifts,
    set_aliases,
)


def _day_num(v) -> int | None:
    """헤더 셀 → 일(day) 정수. 응급실 원티드는 숫자가 아니라 '1' 문자열로 들어온다."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


def locate_dates(ws, days: int) -> tuple[int, int]:
    """헤더에서 1..days 가 연속으로 놓인 (행, 시작열)을 찾는다."""
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, min(ws.max_column, 24) + 1):
            if all(_day_num(ws.cell(r, c + i).value) == i + 1 for i in range(days)):
                return r, c
    raise SystemExit(f"날짜 헤더(1~{days} 연속)를 찾지 못했습니다. 시트를 확인하세요.")


def locate_names(ws, names: set[str], first_day_col: int,
                 row_from: int = 1, row_to: int | None = None) -> tuple[int, dict[str, int]]:
    """날짜 시작열 왼쪽에서 이름 열을 찾고 {이름: 행} 을 만든다.

    row_from/row_to = 한 시트에 여러 그룹이 실린 표에서 대상 구간만 보게 한다
    (응급실 간호사·응급구조사 양쪽에 '김지영' 이 있다).
    """
    hits: list[tuple[int, int, str]] = []
    last = ws.max_row if row_to is None else min(row_to, ws.max_row)
    for r in range(max(1, row_from), last + 1):
        for c in range(1, first_day_col):
            t = name_key(cell_text(ws, r, c))
            if t in names:
                hits.append((c, r, t))
    if not hits:
        raise SystemExit("시트에서 그룹 간호사 이름을 찾지 못했습니다.")
    name_col = Counter(c for c, _, _ in hits).most_common(1)[0][0]
    rows: dict[str, int] = {}
    for c, r, t in hits:
        if c == name_col and t not in rows:
            rows[t] = r
    return name_col, rows


def main() -> None:
    ap = argparse.ArgumentParser(description="근무신청 엑셀 → fixed_wanted_entries")
    ap.add_argument("--file", required=True)
    ap.add_argument("--group-id", required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--sheet", required=True, help="근무신청 시트명(정확히)")
    ap.add_argument("--created-by", default=None, help="created_by nurse_id (기본: 그룹 HN)")
    ap.add_argument("--source-type", default="original", choices=["original", "added", "modified"])
    ap.add_argument("--no-apply-flag", action="store_true",
                    help="is_applied=False 로 넣는다(참고용 보관)")
    ap.add_argument("--replace", action="store_true",
                    help="같은 (그룹·연월) 기존 엔트리를 지우고 다시 넣는다")
    ap.add_argument("--skip-unmapped", action="store_true",
                    help="단일 근무코드로 확정할 수 없는 신청(예 'D/OFF', 'N제외')을 건너뛴다")
    ap.add_argument("--seed-shifts", action="store_true",
                    help="신청에 있으나 shifts 에 없는 코드를 SHIFT_DEFS 정의대로 자동 등록")
    ap.add_argument("--alias", action="append", default=[], metavar="엑셀이름=DB이름",
                    help="엑셀 표기를 DB 이름에 잇는다. 예 '이한솔b=이한솔' (반복 지정)")
    ap.add_argument("--rows-from", type=int, default=1, metavar="행",
                    help="이름을 찾을 시작 행. 한 시트에 여러 그룹이 실렸을 때 대상 구간만 본다")
    ap.add_argument("--rows-to", type=int, default=None, metavar="행",
                    help="이름을 찾을 끝 행(포함)")
    ap.add_argument("--apply", action="store_true", help="실제 반영 (미지정 시 dry-run)")
    args = ap.parse_args()

    set_aliases(args.alias)

    db_name = os.getenv("EUN_DB_NAME", "(미설정)")
    days = calendar.monthrange(args.year, args.month)[1]

    print("=" * 78)
    print(f"대상 DB : {db_name}{'   ← 운영' if db_name == 'eun_roster' else ''}")
    print(f"그룹    : {args.group_id} / {args.year}-{args.month:02d} ({days}일)")
    print(f"시트    : {args.sheet}")
    print(f"모드    : {'APPLY' if args.apply else 'DRY-RUN (쓰기 없음)'}")
    print("=" * 78)

    db = SessionLocal()
    try:
        group = db.query(Group).filter(Group.group_id == args.group_id).first()
        if not group:
            raise SystemExit(f"그룹을 찾을 수 없습니다: {args.group_id}")
        print(f"병동    : {group.group_name}")

        nurses = db.query(Nurse).filter(
            Nurse.group_id == args.group_id, Nurse.active == 1
        ).all()
        by_name = {name_key(n.name): n for n in nurses}
        if len(by_name) != len(nurses):
            raise SystemExit("그룹 내 동명이인이 있어 이름 매칭이 불가합니다.")

        shift_row_id = {
            s.shift_id: s.id
            for s in db.query(Shift).filter(Shift.group_id == args.group_id).all()
        }

        wb = openpyxl.load_workbook(args.file, data_only=True)
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"시트가 없습니다: {args.sheet}\n가능: {wb.sheetnames}")
        ws = wb[args.sheet]

        hdr_row, first_day_col = locate_dates(ws, days)
        name_col, name_rows = locate_names(ws, set(by_name), first_day_col,
                                           args.rows_from, args.rows_to)
        print(f"레이아웃: 날짜헤더 행{hdr_row} · 1일=열{first_day_col} · 이름=열{name_col}")
        print(f"명단    : {len(name_rows)}/{len(nurses)}명 매칭")
        unmatched = sorted(set(by_name) - set(name_rows))
        if unmatched:
            print(f"  · 시트에 없는 간호사: {', '.join(unmatched)}")

        # 파싱 — 값이 있는 셀만 신청으로 본다
        entries: list[tuple[str, int, str]] = []      # (nurse_id, day, shift_id)
        unmapped: Counter = Counter()
        unmapped_detail: list[str] = []
        per_nurse: Counter = Counter()
        code_count: Counter = Counter()
        for nm, row in name_rows.items():
            for i in range(days):
                raw = cell_text(ws, row, first_day_col + i)
                if not raw:
                    continue
                sid, bad, _fixed = normalize_cell(raw)
                if bad:
                    unmapped[bad] += 1
                    unmapped_detail.append(f"{nm} {args.month}/{i + 1} '{bad}'")
                    continue
                if not sid:
                    continue
                entries.append((by_name[nm].nurse_id, i + 1, sid))
                per_nurse[nm] += 1
                code_count[sid] += 1

        print()
        print(f"신청 셀 : {len(entries)}건")
        print("코드    : " + ", ".join(f"{k}={v}" for k, v in code_count.most_common()))
        print("인원별  : " + ", ".join(f"{k} {v}" for k, v in per_nurse.most_common()))

        if unmapped:
            print()
            print(f"  {'·' if args.skip_unmapped else '✗'} 단일 코드로 확정 불가한 신청 "
                  f"{sum(unmapped.values())}건: {dict(unmapped)}")
            for d in unmapped_detail:
                print(f"      {d}")
            if not args.skip_unmapped:
                raise SystemExit(
                    "중단: 'D/OFF'(택일)·'N제외'(조건) 같은 표기는 확정 원티드 한 칸에 담을 수 없습니다.\n"
                    "      건너뛰고 진행하려면 --skip-unmapped 를 주세요."
                )
            print("    → --skip-unmapped 지정: 위 신청은 제외하고 진행합니다.")

        missing_codes = sorted({c for _, _, c in entries if c not in shift_row_id})
        if missing_codes:
            print()
            if not args.seed_shifts:
                print(f"  ✗ shifts 미등록 코드: {missing_codes}")
                raise SystemExit("중단: --seed-shifts 를 주거나 근무코드를 먼저 등록하세요.")
            for c in missing_codes:
                d = SHIFT_DEFS.get(c)
                if not d:
                    raise SystemExit(f"정의가 없는 코드입니다: {c} (SHIFT_DEFS 에 추가 필요)")
                print(f"  · 미등록 '{c}' → 신규 등록 예정 (name={d['name']} type={d['type']})")
            if args.apply:
                created = seed_shifts(db, group.office_id, args.group_id, missing_codes)
                shift_row_id.update(created)
                print("    shifts 신규 등록: "
                      + ", ".join(f"{k}(id={v})" for k, v in created.items()))

        existing = (
            db.query(FixedWantedEntry)
            .filter(
                FixedWantedEntry.group_id == args.group_id,
                FixedWantedEntry.year == args.year,
                FixedWantedEntry.month == args.month,
            )
            .all()
        )
        print()
        print(f"기존 엔트리: {len(existing)}건"
              + ("  → --replace 지정: 삭제 후 재삽입" if existing and args.replace else ""))
        if existing and not args.replace:
            have = {(e.nurse_id, e.shift_date.day) for e in existing}
            before = len(entries)
            entries = [e for e in entries if (e[0], e[1]) not in have]
            print(f"  · 중복 {before - len(entries)}건 건너뜀 → 신규 {len(entries)}건"
                  " (전량 교체하려면 --replace)")

        created_by = args.created_by
        if not created_by:
            hn_ids = group.hn_id if isinstance(group.hn_id, list) else []
            created_by = str(hn_ids[0]) if hn_ids else None
        print(f"created_by : {created_by} · source_type={args.source_type}"
              f" · is_applied={not args.no_apply_flag}")

        if not args.apply:
            print()
            print("DRY-RUN 종료 — 실제 반영하려면 --apply 를 붙이세요.")
            if entries:
                nid, d, sid = entries[0]
                who = next(n.name for n in nurses if n.nurse_id == nid)
                print(f"[샘플] {who} {args.month}/{d} → {sid}")
            return

        if existing and args.replace:
            for e in existing:
                db.delete(e)
            db.flush()

        _now = datetime.now()
        rows = [
            FixedWantedEntry(
                group_id=args.group_id, year=args.year, month=args.month,
                nurse_id=nid, shift_date=date(args.year, args.month, d),
                shift_id=sid, shifts_table_id=shift_row_id[sid],
                is_applied=not args.no_apply_flag,
                source_type=args.source_type,
                original_shift_id=None, reason="병원 근무신청 엑셀 반영",
                created_by=created_by, created_at=_now, updated_at=_now,
            )
            for nid, d, sid in entries
        ]
        db.bulk_save_objects(rows)
        db.commit()
        print()
        print("=" * 78)
        print(f"완료 — fixed_wanted_entries {len(rows)}건 삽입")
        print("알림: 발송되지 않았습니다(원티드 푸시는 라우터 전용 · DB 트리거 없음)")
        print("=" * 78)
    except SystemExit:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
