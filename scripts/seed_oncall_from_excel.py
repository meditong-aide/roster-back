"""근무일정표 엑셀의 콜 표기를 기존 근무표에 주입한다 — 콜 로테이션 **시작점** 만들기.

**dry-run 기본**, `--apply` 로만 write.

## 왜 필요한가

`postprocess_oncall` 은 앵커를 저장하지 않고 **직전 달 근무표에서 마지막 콜 주의 담당
팀을 역산**해 이어간다(월 경계 꼬리물기). 그래서 최초 도입 시 직전 달에 콜이 하나도
없으면 역산할 게 없어 **콜이 0건으로 생성된다.**

시스템은 그동안 콜을 별도 코드로 저장한 적이 없다(2026-08-18 실측: `D1콜`·`오프콜`
schedule_entries 0건). 병원 엑셀에만 `D1 콜` · `OFF (콜)` 로 남아 있으므로, 그걸 한 번
넣어 사슬을 시작시킨다. 이후로는 매달 저절로 이어진다.

## ★ 알림은 나가지 않는다

라우터·서비스를 타지 않고 ORM 으로 직접 write 하며 **푸시 모듈을 import 조차 하지
않는다.** 근무를 바꾸는 게 아니라 **이미 있는 셀의 코드만** 콜 코드로 갈아끼운다
(D1 → D1콜, O → 오프콜). 셀을 새로 만들지 않으므로 인원·일수는 변하지 않는다.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))

import openpyxl  # noqa: E402

from db.client2 import SessionLocal  # noqa: E402
from db.models import Nurse, Schedule, ScheduleEntry, Shift  # noqa: E402
from services.oncall_assign import load_call_code_map  # noqa: E402

_WEEKDAYS = set("월화수목금토일")


def _find_date_row(ws) -> tuple[int, dict[int, int]]:
    best_row, best = 0, {}
    for r in range(1, min(15, ws.max_row) + 1):
        daycol = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= int(v) <= 31:
                daycol[int(v)] = c
        if len(daycol) > len(best):
            best_row, best = r, daycol
    return best_row, best


def _find_name_col(ws, date_row: int) -> int:
    best_col, best_hits = 1, 0
    for c in range(1, min(8, ws.max_column) + 1):
        hits = sum(
            1 for r in range(date_row + 1, ws.max_row + 1)
            if (lambda s: s and 2 <= len(s) <= 5 and s not in _WEEKDAYS)(
                " ".join(str(ws.cell(r, c).value).split()) if ws.cell(r, c).value else "")
        )
        if hits > best_hits:
            best_col, best_hits = c, hits
    return best_col


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet-index", type=int, default=0)
    ap.add_argument("--schedule-id", required=True)
    ap.add_argument("--apply", action="store_true", help="지정해야만 DB 에 쓴다")
    a = ap.parse_args()

    ws = openpyxl.load_workbook(a.file, data_only=True).worksheets[a.sheet_index]
    date_row, day_col = _find_date_row(ws)
    name_col = _find_name_col(ws, date_row)

    db = SessionLocal()
    try:
        sch = db.query(Schedule).filter(Schedule.schedule_id == a.schedule_id).first()
        if sch is None:
            raise SystemExit(f"schedule 없음: {a.schedule_id}")
        gid = sch.group_id
        print(f"DB={db.get_bind().url.database}  schedule={a.schedule_id} "
              f"{sch.year}-{sch.month:02d} status={sch.status} group={gid}")
        print(f"[서식] 날짜행={date_row} 이름열={name_col} 날짜칸={len(day_col)}개 (자동 탐지)")

        code_map = load_call_code_map(db, gid)
        if not code_map:
            raise SystemExit("shifts.call_base_id 미등록 — 콜 코드 맵이 비어 있습니다. DDL·등록 먼저.")
        print(f"[코드] {code_map}   (기반 근무 → 콜 코드)")

        by_name = {str(n.name).strip(): str(n.nurse_id) for n in
                   db.query(Nurse).filter(Nurse.group_id == gid).all()}
        pk = {s.shift_id: s.id for s in db.query(Shift).filter(Shift.group_id == gid).all()}
        ents = {(str(e.nurse_id), (e.work_date.day if hasattr(e.work_date, "day") else e.work_date)): e
                for e in db.query(ScheduleEntry).filter(
                    ScheduleEntry.schedule_id == a.schedule_id).all()}

        plan, skipped, unknown = [], [], []
        for r in range(date_row + 1, ws.max_row + 1):
            v = ws.cell(r, name_col).value
            nm = " ".join(str(v).split()) if v else ""
            nid = by_name.get(nm)
            if not nid:
                continue
            for d, c in sorted(day_col.items()):
                raw = ws.cell(r, c).value
                if raw is None or "콜" not in str(raw):
                    continue
                e = ents.get((nid, d))
                if e is None:
                    skipped.append((nm, d, "근무표에 셀 없음"))
                    continue
                base = str(e.shift_id or "").strip()
                call = code_map.get(base)
                if call is None:
                    # 이미 콜이거나, 콜을 걸 수 없는 근무(휴가·교육 등)
                    if base in code_map.values():
                        skipped.append((nm, d, f"이미 {base}"))
                    else:
                        unknown.append((nm, d, base, " ".join(str(raw).split())))
                    continue
                plan.append((e, nm, d, base, call))

        print(f"\n[계획] {len(plan)}셀 교체 · 건너뜀 {len(skipped)} · 불가 {len(unknown)}")
        from collections import Counter
        print(f"   {dict(Counter(f'{b}→{c}' for _e, _n, _d, b, c in plan))}")
        if unknown:
            print("\n[불가 — 엑셀은 콜인데 근무표 코드가 콜 대상이 아님]")
            for nm, d, base, raw in unknown[:20]:
                print(f"   {nm:<8} {d:>2}일 근무표={base!r} 엑셀={raw!r}")
        if skipped:
            print(f"\n[건너뜀] {skipped[:10]}{' …' if len(skipped) > 10 else ''}")

        if not a.apply:
            print("\n[dry-run] --apply 를 붙여야 실제로 씁니다.")
            return 0

        for e, _nm, _d, _base, call in plan:
            e.shift_id = call
            e.id = pk.get(call)
        db.commit()
        print(f"\n[적용] schedule_entries {len(plan)}셀 콜 코드로 교체 커밋 (알림 없음)")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
