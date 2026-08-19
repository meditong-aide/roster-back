"""마감된 근무표에서 통째로 빠진 인원의 근무를 엑셀에서 채워 넣는다.

**dry-run 기본**, `--apply` 로만 write.

실측(2026-08 수술실 `84eece700fdf`): 마감본 589셀에 19명이 있는데 그중 하나는
`nurses` 에 없는 삭제 인원(458754)이고, active 20명 중 **김성태·고수연 2명이 통째로
누락**돼 있었다. 두 사람은 엑셀 근무일정표에는 정상으로 들어 있다.

## ★ 알림은 나가지 않는다

라우터·서비스를 타지 않고 ORM 으로 직접 write 하며 **푸시 모듈을 import 조차 하지
않는다.** 마감된 근무표에 행을 더하는 작업이라 재발행(issue) 흐름도 건드리지 않는다.

## 코드 변환

시스템은 콜을 별도 코드로 저장하지 않는다 — 8월 마감본 코드 분포에 `D1콜`·`오프콜` 이
0건이고 전부 `D1`/`O` 다. 엑셀의 `D1 콜`·`OFF (콜)` 도 같은 기준으로 눕힌다.
그래야 기존 19명과 일관된다.
"""
from __future__ import annotations

import argparse
import re
import sys
import uuid
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))

import openpyxl  # noqa: E402

from db.client2 import SessionLocal  # noqa: E402
from db.models import Nurse, Schedule, ScheduleEntry, Shift  # noqa: E402

_WEEKDAYS = set("월화수목금토일")

#: 엑셀 표기 → DB `shifts.shift_id`. 콜은 근무/휴무로 눕힌다(위 docstring 참고).
CODE_MAP: dict[str, str] = {
    "D1": "D1", "D1 콜": "D1", "D1콜": "D1",
    "OFF": "O", "OFF (콜)": "O", "OFF(콜)": "O", "O": "O",
    "M": "M", "주": "주",
    "보건": "보건", "휴": "휴", "감": "감",
    "노조 교육": "노조교육", "노조교육": "노조교육",
    "노조": "노조", "보수 교육": "보수교육", "보수교육": "보수교육",
    "예비군": "예비군", "출장": "출장", "산전": "산전", "육휴": "육휴",
    "공가": "공가", "병가": "병가", "특휴": "특휴",
    "자녀돌봄": "자녀돌봄", "장기재직": "장기재직", "대휴": "대휴",
}


def normalize(raw: str) -> str | None:
    """엑셀 셀 → shift_id. 꼬리 숫자가 붙는 표기를 걷어낸다.

    실측: `휴PM 14-5.5` · `휴 20-14.5` · `육아 휴직 95` · `노조 대휴 (2/27)`
    """
    s = " ".join(str(raw).split())
    if not s:
        return None
    if s in CODE_MAP:
        return CODE_MAP[s]
    head = s.split()[0]
    if head.startswith("휴PM"):
        return "휴PM"
    if head.startswith("휴AM"):
        return "휴AM"
    if head.startswith("휴"):
        return "휴"
    if s.startswith("육아 휴직") or s.startswith("육아휴직"):
        return "육휴"
    if s.startswith("노조 대휴") or s.startswith("대휴"):
        return "대휴"
    if s.startswith("특휴"):
        return "특휴"
    if s.startswith("감정"):
        return "감"
    if head in CODE_MAP:
        return CODE_MAP[head]
    return None


def find_date_row(ws) -> tuple[int, dict[int, int]]:
    best_row, best = 0, {}
    for r in range(1, min(15, ws.max_row) + 1):
        daycol = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= int(v) <= 31:
                daycol[int(v)] = c
        if len(daycol) > len(best):
            best_row, best = r, daycol
    return best_row, best


def find_name_col(ws, date_row: int) -> int:
    best_col, best_hits = 1, 0
    for c in range(1, min(8, ws.max_column) + 1):
        hits = 0
        for r in range(date_row + 1, ws.max_row + 1):
            v = ws.cell(r, c).value
            s = " ".join(str(v).split()) if v else ""
            if s and 2 <= len(s) <= 5 and s not in _WEEKDAYS:
                hits += 1
        if hits > best_hits:
            best_col, best_hits = c, hits
    return best_col


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--sheet-index", type=int, default=0)
    ap.add_argument("--schedule-id", required=True)
    ap.add_argument("--names", required=True, help="채워 넣을 사람(콤마 구분)")
    ap.add_argument("--apply", action="store_true", help="지정해야만 DB 에 쓴다")
    a = ap.parse_args()
    targets = [s.strip() for s in a.names.split(",") if s.strip()]

    ws = openpyxl.load_workbook(a.file, data_only=True).worksheets[a.sheet_index]
    date_row, day_col = find_date_row(ws)
    name_col = find_name_col(ws, date_row)

    db = SessionLocal()
    try:
        sch = db.query(Schedule).filter(Schedule.schedule_id == a.schedule_id).first()
        if sch is None:
            raise SystemExit(f"schedule 없음: {a.schedule_id}")
        print(f"DB={db.get_bind().url.database}  schedule={a.schedule_id} "
              f"{sch.year}-{sch.month:02d} status={sch.status} group={sch.group_id}")
        print(f"[서식] 날짜행={date_row} 이름열={name_col} 날짜칸={len(day_col)}개 (자동 탐지)")

        nurses = {str(n.name or "").strip(): n for n in
                  db.query(Nurse).filter(Nurse.group_id == sch.group_id).all()}
        shift_pk = {s.shift_id: s.id for s in
                    db.query(Shift).filter(Shift.group_id == sch.group_id).all()}

        existing = db.query(ScheduleEntry).filter(
            ScheduleEntry.schedule_id == a.schedule_id).all()
        have = {(str(e.nurse_id), (e.work_date.date() if hasattr(e.work_date, "date")
                                   else e.work_date)) for e in existing}
        print(f"[현황] 기존 {len(existing)}셀 · 등장 {len({n for n, _ in have})}명")

        rows = {}
        for r in range(date_row + 1, ws.max_row + 1):
            v = ws.cell(r, name_col).value
            nm = " ".join(str(v).split()) if v else ""
            if nm in targets:
                rows[nm] = r

        missing_names = [t for t in targets if t not in rows]
        if missing_names:
            raise SystemExit(f"엑셀에 없음: {missing_names}")

        planned, skipped, unknown = [], [], []
        for nm in targets:
            n = nurses.get(nm)
            if n is None:
                raise SystemExit(f"nurses 에 없음: {nm}")
            r = rows[nm]
            for d in sorted(day_col):
                wd = date(sch.year, sch.month, d)
                raw = ws.cell(r, day_col[d]).value
                if raw is None or not str(raw).strip():
                    skipped.append((nm, d, "엑셀 빈칸"))
                    continue
                code = normalize(raw)
                if code is None:
                    unknown.append((nm, d, " ".join(str(raw).split())))
                    continue
                if code not in shift_pk:
                    unknown.append((nm, d, f"{code}(그룹 shifts 에 없음)"))
                    continue
                if (str(n.nurse_id), wd) in have:
                    skipped.append((nm, d, "이미 있음"))
                    continue
                planned.append((n, wd, code, " ".join(str(raw).split())))

        print(f"\n[계획] {len(planned)}셀 추가 · 건너뜀 {len(skipped)} · 미해석 {len(unknown)}")
        for nm in targets:
            cnt = sum(1 for n, _, _, _ in planned if str(n.name).strip() == nm)
            print(f"   {nm:<8} {cnt:>2}셀")
        if skipped:
            print("\n[건너뜀]")
            for nm, d, why in skipped:
                print(f"   {nm:<8} {d:>2}일 — {why}")
        if unknown:
            print("\n[미해석 — 중단 사유]")
            for nm, d, raw in unknown:
                print(f"   {nm:<8} {d:>2}일 {raw!r}")
            raise SystemExit("해석하지 못한 근무코드가 있어 중단합니다.")

        if not a.apply:
            print("\n[dry-run] --apply 를 붙여야 실제로 씁니다.")
            return 0

        for n, wd, code, _raw in planned:
            db.add(ScheduleEntry(
                entry_id=str(uuid.uuid4().hex)[:16],
                schedule_id=a.schedule_id,
                nurse_id=n.nurse_id,
                work_date=datetime(wd.year, wd.month, wd.day),
                shift_id=code,
                id=shift_pk[code],
            ))
        db.commit()
        print(f"\n[적용] schedule_entries {len(planned)}행 커밋 (알림 없음)")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
