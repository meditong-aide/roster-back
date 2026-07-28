#!/usr/bin/env python
"""확정(마감) 근무표 엑셀 → roster DB import.

병원이 이미 확정해 운영 중인 근무표를 시스템에 그대로 올린다. 첫 대상은 인천의료원
41병동-RN 이며, 다른 병동도 같은 스크립트에 인자만 바꿔 처리한다.

★ 알림 없음
    푸시는 `POST /roster/publish` 라우터 안에서만 발송된다
    (`send_roster_publish_push` / `send_roster_republish_push` / 파견자 알림).
    이 스크립트는 그 라우터를 타지 않고 ORM 으로 직접 write 하며,
    푸시 관련 모듈을 **import 하지 않는다**. schedules/schedule_entries/issued_roster
    에는 DB 트리거도 없음을 확인했다. 따라서 알림은 어떤 경로로도 발생하지 않는다.

사용법
    # 1) dry-run (기본) — 아무것도 쓰지 않고 매핑 결과만 출력
    cd roster-back
    EUN_DB_NAME=eun_roster .venv/bin/python scripts/import_finalized_roster.py \
        --file "/path/2026년 41병동 AI 근무표.xlsx" \
        --group-id 1022438ea001 --year 2026 --month 8

    # 2) 실제 적용
    EUN_DB_NAME=eun_roster .venv/bin/python scripts/import_finalized_roster.py \
        --file "..." --group-id 1022438ea001 --year 2026 --month 8 --apply

선행 조건
    엑셀에 있는 근무코드가 모두 그룹 `shifts` 에 등록돼 있어야 한다.
    미등록 코드가 있으면 중단하며, 41병동은 `scripts/ward_shifts_seed.sql` 을 먼저 실행한다.
"""
from __future__ import annotations

import argparse
import calendar
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

# ── app 패키지 경로 등록 (EUN_DB_NAME 은 import 전에 정해져 있어야 engine 이 올바른 DB 를 잡는다)
_APP = Path(__file__).resolve().parents[1] / "app"
sys.path.insert(0, str(_APP))

import openpyxl  # noqa: E402
from sqlalchemy import func, text  # noqa: E402

from db.client2 import SessionLocal  # noqa: E402
from db.models import (  # noqa: E402
    Group,
    IssuedRoster,
    Nurse,
    Schedule,
    ScheduleEntry,
    Shift,
)

# ── 엑셀 표기 → 그룹 shift_id ------------------------------------------------
# 접두 일치로 판정하며, 긴 것부터 확인한다(수면휴가 → 수면 / DE → D 오인 방지).
PREFIX_MAP: list[tuple[str, str]] = [
    ("수면휴가", "수면"),
    ("수면", "수면"),
    ("노조공가", "노조"),
    ("노조", "노조"),
    # 51병동은 축약 표기를 쓴다: '보수'(=보수교육) · '단'(=단축). 긴 것부터 매칭해야 하므로 순서 주의.
    ("보수교육", "보수교육"),
    ("보수", "보수교육"),
    ("교육", "교육"),
    # 응급실: KALS(전문심장소생술) 교육 · 자녀돌봄휴가
    ("KALS", "KALS"), ("자녀돌봄", "자녀돌봄"),
    # 호스피스: 출장 · 병가 · 예비군(SHIFT_DEFS 엔 있었으나 매핑 규칙이 없었다)
    ("출장", "출장"), ("병가", "병가"), ("예비군", "예비군"),
    # 52병동: 대휴(휴일근무 대체휴무) — '대휴(2/27)' · '대휴/(2/27)' 처럼 원래 근무일이 붙는다
    ("대휴", "대휴"),
    ("보건", "보건"),
    ("단축근무", "단축"),
    ("단축", "단축"),
    ("단", "단축"),
    ("상가", "상가"),
    # 근무신청 시트는 근무표와 표기가 다르다 — 근무표 '특휴(모환갑)'/'공가(회계감사)' 가
    # 신청서에는 '특별휴가'/'회계감사' 로만 적힌다. 같은 코드로 수렴시킨다.
    ("특별휴가", "특휴"),
    ("특휴", "특휴"),
    ("회계감사", "공가"),
    ("공가", "공가"),
    ("산전", "산전"),
    ("출산휴가", "출산"),
    ("출산", "출산"),
    # 같은 감정노동휴가를 병동마다 달리 적는다('감정휴가' / 응급실 '감정노동')
    ("감정휴가", "감"),
    ("감정노동", "감"),
    ("휴", "휴"),
    ("주휴", "주"),
    ("주", "주"),
]
# 근무 계열은 뒤에 누적번호가 붙는다(N7, D2, E1). DE 를 D 보다 먼저 본다.
WORK_RE = re.compile(r"^(DE|D|E|N|M)\d*$")
# 지원 근무 표기는 제거하고 기본코드만 남긴다. 병동마다 표기가 제각각이다.
#   타 팀 지원  : D(B) · N15(B) · E(C)4      (41병동)
#   타 병동 지원: D(w41) · N11(w41)          (42병동 → 41병동)
#                 D(52) · E(42) · N(42)      (41병동 → 52·42병동, 병동번호만)
# ※ 지원 근무는 개념상 파견(nurse_assignment reason='파견')이지만, 확정본 import 는
#   원본 셀을 그대로 옮기는 것이 목적이라 여기서는 기본코드로만 정규화한다.
#                 E(출장)                    (호스피스 — 출장을 겸한 E 근무)
#                 D(41w) · E(41W)             (52병동 → 41병동, w 가 뒤에 붙는 표기)
TEAM_SUFFIX_RE = re.compile(r"\((?:[ABCD]|[wW]?\d{2,3}[wW]?|출장)\)")
# 닫히지 않은 여는 괄호부터 문자열 끝까지 — 손으로 단 주석의 흔적
UNCLOSED_PAREN_RE = re.compile(r"\([^)]*$")
# 앞머리 괄호주석 — 코드 앞에 붙은 인사 메모 ('(입사)M' → M, '(퇴사)OFF' → OFF)
LEADING_PAREN_RE = re.compile(r"^\([^)]*\)")
# 타 병동 소속 기간 표시 — 근무가 아니라 "그 달 일부를 다른 병동에서 보냈다"는 메모.
#   호스피스 박지민 7/1 '52뵹동'(52병동 오타) 뒤로 22일이 빈칸이고 7/24 부터 전입 근무가 시작된다.
#   52병동 박지민 7월 말 '호스피스병동(7/24~)' 과 호스피스 쪽 '52뵹동' 이 같은 이동의 양면이다.
WARD_MOVE_RE = re.compile(r"^[가-힣0-9]{2,8}[뵹병]동(\(.*\))?$")
# 그룹 명단 밖(휴직 등) — 행 자체를 건너뛴다.
SKIP_PREFIXES = ("육아",)

# ── --seed-shifts 로 자동 등록할 근무코드 정의 --------------------------------
# 41병동-RN 기존 12종의 관행을 따른다(allday=0 / off_swap_target=0 / deleteless=0
# / is_weekly_off=0 / default_shift=NULL). 다른 병동도 같은 정의를 재사용한다.
# DE = 수간호사 고정근무 → 자동배정·원티드 선택 대상이 아니므로 둘 다 0.
#      근무시간은 병원 확인 전이라 NULL(확인되면 UPDATE).
SHIFT_DEFS: dict[str, dict] = {
    # 41병동-RN 이 이미 갖고 있던 표준 휴가/공가 6종 — 조무사(AN) 그룹 등 코드가
    # D/E/N/M/O/주 6종뿐인 병동에서 필요하다. 색·sequence 를 RN 과 맞춰 통일한다.
    "보건":     {"name": "보건휴가", "type": "공가", "color": "#004cff",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 7},
    "감":       {"name": "감정휴가", "type": "휴가", "color": "#319b86",
                 "auto_schedule": 1, "show_in_preference": False, "sequence": 8},
    "노조":     {"name": "노조공가", "type": "공가", "color": "#dc1ed6",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 9},
    "수면":     {"name": "수면휴가", "type": "휴가", "color": "#d6c400",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 10},
    "휴":       {"name": "휴가", "type": "휴가", "color": "#fb7ee0",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 11},
    "산전":     {"name": "산전휴가", "type": "휴가", "color": "#85d016",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 12},
    "DE":       {"name": "수간호사 고정근무", "type": "근무", "color": "#6b5b95",
                 "auto_schedule": 0, "show_in_preference": False, "sequence": 13},
    "보수교육": {"name": "보수교육", "type": "근무", "color": "#2e8b57",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 14},
    "상가":     {"name": "상가", "type": "휴가", "color": "#808080",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 15},
    "단축":     {"name": "단축근무", "type": "휴가", "color": "#ff8c42",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 16},
    "공가":     {"name": "공가", "type": "공가", "color": "#4a90d9",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 17},
    "특휴":     {"name": "특별휴가", "type": "휴가", "color": "#c94f7c",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 18},
    # 별관1병동에서 처음 등장 — 예비군훈련(법정 공가)
    "예비군":   {"name": "예비군훈련", "type": "공가", "color": "#3f7d3f",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 19},
    # 51병동에서 처음 등장 — 보수교육과 별개인 일반 교육
    "교육":     {"name": "교육", "type": "근무", "color": "#8a6d3b",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 20},
    # 42병동에서 처음 등장 — 출산휴가(연속 장기)
    "출산":     {"name": "출산휴가", "type": "휴가", "color": "#d16ba5",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 21},
    # 응급실에서 처음 등장 — KALS(전문심장소생술) 교육. 보수교육/교육과 별개로
    # 엑셀 표기를 그대로 보존한다(병원이 화면에서 같은 이름으로 찾을 수 있어야 한다).
    "KALS":     {"name": "KALS 교육", "type": "근무", "color": "#5b8c85",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 22},
    "자녀돌봄": {"name": "자녀돌봄휴가", "type": "휴가", "color": "#e0994f",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 23},
    # 호스피스에서 처음 등장
    "출장":     {"name": "출장", "type": "근무", "color": "#4f8ae0",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 24},
    "병가":     {"name": "병가", "type": "휴가", "color": "#a8736b",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 25},
    # 52병동에서 처음 등장 — 휴일근무 대체휴무
    "대휴":     {"name": "대체휴무", "type": "휴가", "color": "#7fa8c9",
                 "auto_schedule": 1, "show_in_preference": True, "sequence": 26},
}


def seed_shifts(db, office_id: str, group_id: str, codes: list[str]) -> dict[str, int]:
    """미등록 근무코드를 SHIFT_DEFS 정의대로 등록하고 {shift_id: shifts.id} 를 돌려준다.

    ORM(Shift) 대신 raw SQL 을 쓰는 이유: `deleteless` 가 DB 에는 있으나 모델에
    매핑돼 있지 않다. 기존 12종이 모두 deleteless=0 이라 같은 값을 넣어 맞춘다.
    shifts.id 는 IDENTITY 이므로 값을 넣지 않고 OUTPUT 으로 채번 결과를 받는다.
    """
    undefined = [c for c in codes if c not in SHIFT_DEFS]
    if undefined:
        raise SystemExit(f"정의가 없는 코드라 자동 등록할 수 없습니다: {undefined}")
    stmt = text(
        "INSERT INTO shifts (shift_id, office_id, group_id, name, color, start_time, end_time,"
        " type, allday, auto_schedule, duration, sequence, deleteless,"
        " default_shift, is_weekly_off, shift_gb, show_in_preference, off_swap_target, description)"
        " OUTPUT INSERTED.id"
        " VALUES (:shift_id, :office_id, :group_id, :name, :color, NULL, NULL,"
        " :type, 0, :auto_schedule, NULL, :sequence, 0,"
        " NULL, 0, NULL, :show_in_pref, 0, NULL)"
    )
    created: dict[str, int] = {}
    for code in sorted(codes, key=lambda c: SHIFT_DEFS[c]["sequence"]):
        d = SHIFT_DEFS[code]
        new_id = db.execute(stmt, {
            "shift_id": code, "office_id": office_id, "group_id": group_id,
            "name": d["name"], "color": d["color"], "type": d["type"],
            "auto_schedule": d["auto_schedule"], "sequence": d["sequence"],
            "show_in_pref": 1 if d["show_in_preference"] else 0,
        }).scalar()
        created[code] = int(new_id)
    db.flush()
    return created


def _match(s: str) -> str | None:
    """정리된 문자열 → shift_id (실패 시 None)."""
    if s.upper() in ("OFF", "O"):
        return "O"
    m = WORK_RE.match(s.upper())
    if m:
        return m.group(1)
    for prefix, shift_id in PREFIX_MAP:    # 상가1(외조부상) / 휴15-9 / 공가(회계감사)
        if s.startswith(prefix):
            return shift_id
    return None


def normalize_cell(raw: str) -> tuple[str | None, str | None, str | None]:
    """엑셀 셀 원문 → (shift_id, 매핑실패원문, 오타보정원문).

    운영자가 손으로 채운 표라 짝이 안 맞는 괄호 같은 오타가 섞인다
    (예: 최지은 2026-07-20 '(DE'). 1차 매칭이 실패하면 괄호만 털어내고
    한 번 더 시도하되, 무엇을 보정했는지 호출측에 알려 로그로 남긴다.
    """
    s = (raw or "").replace("\n", "").replace(" ", "").strip()
    if not s:
        return None, None, None
    if s.startswith(SKIP_PREFIXES):
        return None, None, None
    if WARD_MOVE_RE.match(s):              # '52뵹동' = 타 병동 소속 기간 메모
        return None, None, None
    s = TEAM_SUFFIX_RE.sub("", s)          # D(B) → D,  N15(B) → N15,  E(출장) → E
    s = LEADING_PAREN_RE.sub("", s) or s   # '(입사)M' → M,  '(퇴사)OFF' → OFF
    hit = _match(s)
    if hit:
        return hit, None, None
    stripped = s.strip("()")               # '(DE' → 'DE'
    if stripped and stripped != s:
        hit = _match(stripped)
        if hit:
            return hit, None, s
    # 닫히지 않은 여는 괄호부터 끝까지 = 손으로 단 주석. 잘라내고 다시 맞춰본다
    # (응급실 장승수 2026-07-20 'DE(입사'). 닫힌 괄호는 건드리지 않으므로
    #  '공가(의료지원)'·'상가N(사유)' 같은 정상 표기는 영향이 없다.
    cut = UNCLOSED_PAREN_RE.sub("", s)
    if cut and cut != s:
        hit = _match(cut)
        if hit:
            return hit, None, s
    return None, s, None


def month_days(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def cell_text(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    return "" if v is None else str(v).replace("\n", "").strip()


_ALIAS: dict[str, str] = {}


def _raw_key(s: str) -> str:
    return re.sub(r"\([^)]*\)", "", s).replace(" ", "").strip().upper()


def set_aliases(pairs: list[str]) -> None:
    """`--alias 엑셀이름=DB이름` 목록을 등록한다.

    엑셀에만 붙는 표기를 DB 이름으로 잇는다(51병동 `이한솔b` → DB `이한솔`:
    office 안에 이한솔이 1명뿐이라 b 는 동명이인 구분자가 아니었다).
    DB 데이터를 건드리지 않고 매칭만 보정하는 용도.
    """
    for p in pairs or []:
        if "=" not in p:
            raise SystemExit(f"--alias 는 엑셀이름=DB이름 형식: {p!r}")
        src, dst = p.split("=", 1)
        _ALIAS[_raw_key(src)] = _raw_key(dst)


def name_key(s: str) -> str:
    """이름 비교용 정규화 — 괄호 주석 제거 + 공백 제거 + 영문 접미사 대소문자 통일 + 별칭.

    시트마다 이름 표기가 다르다(별관1병동):
      - 동명이인 구분자가 엑셀은 소문자(`이정희a`), DB는 대문자(`이정희A`)
      - 원티드 시트는 역할·성별 주석이 붙는다(`황상은(파트장)`, `이정희b(남)`)
    한글에는 대소문자가 없어 upper() 부작용이 없다.
    """
    k = _raw_key(s)
    return _ALIAS.get(k, k)


def pick_sheet(wb, month: int, explicit: str | None) -> str:
    if explicit:
        if explicit not in wb.sheetnames:
            raise SystemExit(f"시트를 찾을 수 없습니다: {explicit} (가능: {wb.sheetnames})")
        return explicit
    cands = [s for s in wb.sheetnames if f"{month}월" in s and "근무표" in s and "조무사" not in s]
    if len(cands) != 1:
        raise SystemExit(
            f"시트 자동 선택 실패(후보 {cands}). --sheet 로 지정하세요. 전체: {wb.sheetnames}"
        )
    return cands[0]


def locate_layout(ws, names: set[str], days: int,
                  row_from: int = 1, row_to: int | None = None,
                  day_col: int | None = None) -> tuple[int, dict[str, int]]:
    """이름이 적힌 열과 (이름 → 행)을 찾는다. 날짜 열은 이름 열 바로 다음부터.

    row_from/row_to 는 한 시트에 여러 그룹이 함께 실린 표에서 대상 구간만 보게 한다
    (응급실은 간호사표 아래에 응급구조사표가 이어지고, 양쪽에 '김지영' 이 각각 있다.
     alias 로는 못 가른다 — name_key 는 DB 이름에도 걸려 되레 뒤바뀐다).
    """
    hits: list[tuple[int, int, str]] = []
    last = ws.max_row if row_to is None else min(row_to, ws.max_row)
    for r in range(max(1, row_from), last + 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            t = name_key(cell_text(ws, r, c))
            if t in names:
                hits.append((c, r, t))
    if not hits:
        raise SystemExit("시트에서 그룹 간호사 이름을 하나도 찾지 못했습니다.")
    name_col = Counter(c for c, _, _ in hits).most_common(1)[0][0]
    rows: dict[str, int] = {}
    for c, r, t in hits:
        if c != name_col:
            continue
        # 같은 이름이 여러 행이면 근무 셀이 채워진 행을 택한다(참고사항 행 배제).
        _d0 = day_col or (name_col + 1)
        filled = sum(1 for i in range(days) if cell_text(ws, r, _d0 + i))
        if t not in rows or filled > rows[t][1]:
            rows[t] = (r, filled)
    return name_col, {n: rc[0] for n, rc in rows.items() if rc[1] > 0}


def find_unknown_rows(ws, name_col: int, names: set[str], days: int,
                      row_from: int = 1, row_to: int | None = None,
                      day_col: int | None = None) -> list[tuple[str, int]]:
    """이름 열에 있으나 DB 명단에 없는 근무자 행을 찾는다.

    이걸 잡지 않으면 그 사람의 근무가 통째로 누락된 채 "명단 9/9 매칭" 처럼 보인다
    (실제 41병동-AN 조무사표에서 DE 상시근무 4명이 이 경우였다).
    D/E/N 소계 행은 값이 숫자라 제외한다.
    ※ 절반 미만만 채워진 행은 헤더/메모로 보고 넘긴다 — 중도 입·퇴사자는 여기서
      안 걸릴 수 있으므로(응급실 김은총: N9 뒤 병가11 후 퇴사) 시트를 함께 볼 것.
    """
    out: list[tuple[str, int]] = []
    last = ws.max_row if row_to is None else min(row_to, ws.max_row)
    for r in range(max(1, row_from), last + 1):
        t = cell_text(ws, r, name_col)
        if not t or name_key(t) in names or t in ("이름", "성명"):
            continue
        vals = [cell_text(ws, r, (day_col or name_col + 1) + i) for i in range(days)]
        filled = [v for v in vals if v]
        if len(filled) < days // 2:          # 절반 미만이면 헤더/메모 행으로 본다
            continue
        if all(v.replace(".", "").isdigit() for v in filled):   # 소계 행
            continue
        out.append((t, len(filled)))
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="확정 근무표 엑셀 → roster DB import")
    ap.add_argument("--file", required=True, help="엑셀 파일 경로")
    ap.add_argument("--group-id", required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--sheet", default=None, help="미지정 시 '{month}월 ... 근무표' 자동 선택")
    ap.add_argument("--status", default="issued", choices=["issued", "draft"])
    ap.add_argument("--created-by", default=None, help="schedules.created_by (기본: 그룹 HN account_id)")
    ap.add_argument("--name", default=None, help="schedules.name (기본: 'N월 근무표 (확정본)')")
    ap.add_argument("--config-id", type=int, default=None, help="기본: 그룹의 최신 config")
    ap.add_argument("--no-snapshot", action="store_true", help="issued_roster_snapshot 생성 생략")
    ap.add_argument("--seed-shifts", action="store_true",
                    help="엑셀에 있으나 shifts 에 없는 코드를 SHIFT_DEFS 정의대로 자동 등록")
    ap.add_argument("--allow-unknown", action="store_true",
                    help="엑셀에만 있고 DB 명단에 없는 근무자를 제외하고 진행(기본은 중단)")
    ap.add_argument("--alias", action="append", default=[], metavar="엑셀이름=DB이름",
                    help="엑셀 표기를 DB 이름에 잇는다. 예 '이한솔b=이한솔' (반복 지정)")
    ap.add_argument("--first-day-col", type=int, default=None, metavar="열",
                    help="1일이 놓인 열. 미지정 시 '이름 열 바로 다음'. "
                         "호스피스처럼 이름과 날짜 사이에 입사년도·숙련도가 끼면 지정한다")
    ap.add_argument("--rows-from", type=int, default=1, metavar="행",
                    help="이름을 찾을 시작 행. 한 시트에 여러 그룹이 실렸을 때 대상 구간만 본다")
    ap.add_argument("--rows-to", type=int, default=None, metavar="행",
                    help="이름을 찾을 끝 행(포함). 예 응급실 조무사만: --rows-from 20")
    ap.add_argument("--apply", action="store_true", help="실제 DB 반영 (미지정 시 dry-run)")
    args = ap.parse_args()

    set_aliases(args.alias)

    db_name = os.getenv("EUN_DB_NAME", "(미설정)")
    days = month_days(args.year, args.month)

    print("=" * 78)
    print(f"대상 DB      : {db_name}{'   ← 운영' if db_name == 'eun_roster' else ''}")
    print(f"그룹         : {args.group_id}")
    print(f"연월         : {args.year}-{args.month:02d} ({days}일)")
    print(f"모드         : {'APPLY (실제 반영)' if args.apply else 'DRY-RUN (쓰기 없음)'}")
    print("=" * 78)

    db = SessionLocal()
    try:
        group = db.query(Group).filter(Group.group_id == args.group_id).first()
        if not group:
            raise SystemExit(f"그룹을 찾을 수 없습니다: {args.group_id}")
        office_id = group.office_id
        print(f"병동         : {group.group_name} (office {office_id})")

        nurses = (
            db.query(Nurse)
            .filter(Nurse.group_id == args.group_id, Nurse.active == 1)
            .all()
        )
        by_name: dict[str, list[Nurse]] = defaultdict(list)
        for n in nurses:
            by_name[name_key(n.name)].append(n)
        dup = {k: len(v) for k, v in by_name.items() if len(v) > 1}
        if dup:
            raise SystemExit(f"그룹 내 동명이인이 있어 이름 매칭이 불가합니다: {dup}")
        print(f"간호사       : {len(nurses)}명 (active)")

        shifts = db.query(Shift).filter(Shift.group_id == args.group_id).all()
        shift_row_id = {s.shift_id: s.id for s in shifts}   # schedule_entries.id 용 stable key
        print(f"근무코드     : {len(shifts)}종 — {', '.join(sorted(shift_row_id))}")

        # ── 엑셀 파싱 ──────────────────────────────────────────────────
        wb = openpyxl.load_workbook(args.file, data_only=True)
        sheet = pick_sheet(wb, args.month, args.sheet)
        ws = wb[sheet]
        print(f"시트         : {sheet}")

        name_col, name_rows = locate_layout(ws, set(by_name), days,
                                            args.rows_from, args.rows_to,
                                            args.first_day_col)
        # 대개 이름 바로 옆이 1일이지만, 호스피스는 사이에 입사년도·숙련도 두 칸이 끼어 있다.
        first_day_col = args.first_day_col or (name_col + 1)
        _rng = (f" / 행{args.rows_from}~{args.rows_to or '끝'}"
                if (args.rows_from > 1 or args.rows_to) else "")
        print(f"레이아웃     : 이름=열{name_col} / 1일=열{first_day_col}{_rng}")

        matched = sorted(name_rows)
        missing = sorted(set(by_name) - set(matched))
        print(f"명단 매칭    : {len(matched)}/{len(nurses)}명")
        if missing:
            print(f"  ⚠ 엑셀에 없는 간호사: {', '.join(missing)}")

        unknown = find_unknown_rows(ws, name_col, set(by_name), days,
                                    args.rows_from, args.rows_to, first_day_col)
        if unknown:
            print(f"  ✗ 엑셀에만 있고 DB 명단에 없는 근무자 {len(unknown)}명 — 이 인원의 근무는 누락됩니다:")
            for nm, cnt in unknown:
                print(f"      {nm} ({cnt}일 배정)")
            if not args.allow_unknown:
                raise SystemExit(
                    "중단: 해당 인원을 그룹에 등록하거나, 누락을 감수하려면 --allow-unknown 을 주세요."
                )
            print("    → --allow-unknown 지정됨: 위 인원을 제외하고 진행합니다.")

        parsed: dict[str, list[str | None]] = {}
        unmapped: Counter = Counter()
        code_count: Counter = Counter()
        corrections: list[str] = []
        for nm in matched:
            row = name_rows[nm]
            cells: list[str | None] = []
            for i in range(days):
                raw = cell_text(ws, row, first_day_col + i)
                sid, bad, fixed_from = normalize_cell(raw)
                if bad:
                    unmapped[bad] += 1
                if fixed_from:
                    corrections.append(f"{nm} {args.month}/{i + 1} '{fixed_from}' → {sid}")
                if sid:
                    code_count[sid] += 1
                cells.append(sid)
            parsed[nm] = cells

        print()
        print("코드 분포    : " + ", ".join(f"{k}={v}" for k, v in code_count.most_common()))
        if corrections:
            print(f"오타 보정    : {len(corrections)}건")
            for c in corrections:
                print(f"  · {c}")

        missing_codes = sorted({c for c in code_count if c not in shift_row_id})
        if unmapped:
            print()
            print(f"  ✗ 매핑 실패 셀: {dict(unmapped)}")
            raise SystemExit("중단: 매핑 규칙을 보완한 뒤 다시 실행하세요.")

        if missing_codes:
            print()
            if not args.seed_shifts:
                print(f"  ✗ shifts 미등록 코드: {missing_codes}")
                print("    → --seed-shifts 를 붙이거나 scripts/ward_shifts_seed.sql 을 먼저 실행하세요.")
                raise SystemExit("중단: 근무코드를 먼저 등록해야 합니다.")
            print(f"  · 미등록 코드 {missing_codes} → --seed-shifts 로 신규 등록 예정")
            for c in missing_codes:
                d = SHIFT_DEFS.get(c)
                if not d:
                    raise SystemExit(f"정의가 없는 코드입니다: {c} (SHIFT_DEFS 에 추가 필요)")
                print(f"      {c}: name={d['name']} type={d['type']} "
                      f"auto_schedule={d['auto_schedule']} seq={d['sequence']}")

        total_entries = sum(1 for cs in parsed.values() for c in cs if c)
        blank = sum(1 for cs in parsed.values() for c in cs if not c)
        print(f"생성될 entry : {total_entries}건 (빈 셀 {blank}건은 저장하지 않음)")
        if blank:
            per_nurse = {nm: sum(1 for c in cs if not c) for nm, cs in parsed.items()}
            detail = ", ".join(f"{nm} {v}일" for nm, v in sorted(
                per_nurse.items(), key=lambda x: -x[1]) if v)
            print(f"  빈 셀 분포 : {detail}")

        # ── 기존 근무표 확인 ──────────────────────────────────────────
        existing = (
            db.query(Schedule)
            .filter(
                Schedule.group_id == args.group_id,
                Schedule.year == args.year,
                Schedule.month == args.month,
            )
            .all()
        )
        if existing:
            print(f"기존 근무표  : {len(existing)}건 — "
                  + ", ".join(f"{e.schedule_id}(v{e.version}/{e.status})" for e in existing))

        max_ver = (
            db.query(func.max(Schedule.version))
            .filter(Schedule.group_id == args.group_id)
            .scalar()
            or 0
        )
        version = int(max_ver) + 1

        created_by = args.created_by
        if not created_by:
            hn_ids = group.hn_id if isinstance(group.hn_id, list) else []
            hn = (
                db.query(Nurse).filter(Nurse.nurse_id.in_([str(x) for x in hn_ids])).first()
                if hn_ids else None
            )
            created_by = hn.account_id if hn else None
        if not created_by:
            raise SystemExit("created_by 를 결정할 수 없습니다. --created-by 로 지정하세요.")

        config_id = args.config_id
        if config_id is None:
            from db.models import RosterConfig

            cfg = (
                db.query(RosterConfig)
                .filter(RosterConfig.group_id == args.group_id)
                .order_by(RosterConfig.config_id.desc())
                .first()
            )
            config_id = cfg.config_id if cfg else None

        schedule_id = uuid.uuid4().hex[:12]
        sched_name = args.name or f"{args.month}월 근무표 (확정본)"
        print()
        print(f"신규 schedule: {schedule_id} / v{version} / status={args.status}")
        print(f"  name       : {sched_name}")
        print(f"  created_by : {created_by}")
        print(f"  config_id  : {config_id}")

        if not args.apply:
            print()
            print("DRY-RUN 종료 — 실제 반영하려면 --apply 를 붙이세요.")
            sample = matched[0]
            print(f"\n[샘플] {sample}: " + " ".join(
                f"{i+1}:{parsed[sample][i] or '-'}" for i in range(min(days, 31))
            ))
            return

        # ── 실제 반영 ─────────────────────────────────────────────────
        if missing_codes:
            created = seed_shifts(db, office_id, args.group_id, missing_codes)
            shift_row_id.update(created)
            print(f"shifts 신규 등록 {len(created)}종: "
                  + ", ".join(f"{k}(id={v})" for k, v in created.items()))

        schedule = Schedule(
            schedule_id=schedule_id,
            office_id=office_id,
            group_id=args.group_id,
            year=args.year,
            month=args.month,
            version=version,
            config_id=config_id,
            created_by=created_by,
            status=args.status,
            dropped=False,
            name=sched_name,
            memo=f"병원 확정 근무표 import ({Path(args.file).name} / {sheet})",
        )
        db.add(schedule)

        rows = []
        for nm, cells in parsed.items():
            nurse = by_name[nm][0]
            for i, sid in enumerate(cells):
                if not sid:
                    continue
                rows.append(
                    ScheduleEntry(
                        entry_id=uuid.uuid4().hex[:16],
                        schedule_id=schedule_id,
                        nurse_id=nurse.nurse_id,
                        work_date=datetime(args.year, args.month, i + 1),
                        shift_id=sid,
                        id=shift_row_id[sid],
                    )
                )
        db.bulk_save_objects(rows)
        print(f"schedule_entries {len(rows)}건 준비")

        if args.status == "issued":
            # publish 라우터와 같은 순서: 같은 연월의 기존 issued 를 draft 로 내린다.
            db.query(Schedule).filter(
                Schedule.group_id == args.group_id,
                Schedule.year == args.year,
                Schedule.month == args.month,
                Schedule.status == "issued",
                Schedule.schedule_id != schedule_id,
            ).update({"status": "draft"}, synchronize_session=False)

            max_seq = (
                db.query(func.max(IssuedRoster.seq_no))
                .filter(
                    IssuedRoster.group_id == args.group_id,
                    IssuedRoster.office_id == office_id,
                )
                .scalar()
                or 0
            )
            issuer = db.query(Nurse).filter(Nurse.account_id == created_by).first()
            db.add(
                IssuedRoster(
                    seq_no=int(max_seq) + 1,
                    office_id=office_id,
                    group_id=args.group_id,
                    nurse_id=issuer.nurse_id if issuer else None,
                    version=version,
                    v_name=f"v{version}",
                    issue_cmmt="병원 확정 근무표 반영(알림 미발송)",
                    schedule_id=schedule_id,
                    is_active=True,
                )
            )
            print(f"issued_roster seq_no={int(max_seq) + 1} 준비")

            if not args.no_snapshot:
                try:
                    from types import SimpleNamespace

                    from services.roster_service import create_issued_roster_snapshot

                    db.flush()  # 스냅샷이 entries 를 조회하므로 먼저 flush
                    snap = create_issued_roster_snapshot(
                        schedule=schedule,
                        current_user=SimpleNamespace(
                            nurse_id=issuer.nurse_id if issuer else None,
                            account_id=created_by,
                        ),
                        year=args.year,
                        month=args.month,
                        office_id=office_id,
                        group_id=args.group_id,
                        db=db,
                    )
                    db.add(snap)
                    print("issued_roster_snapshot 준비")
                except Exception as e:  # 스냅샷 실패가 본체를 막지 않도록
                    print(f"  ⚠ 스냅샷 생성 생략(오류): {e}")

        db.commit()
        print()
        print("=" * 78)
        print(f"완료 — schedule_id = {schedule_id}")
        print(f"롤백: scripts/rollback_imported_roster.sql 의 @schedule_id 에 위 값을 넣어 실행")
        print("알림: 발송되지 않았습니다(푸시는 publish 라우터 전용 · DB 트리거 없음)")
        print("=" * 78)
    except SystemExit:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
