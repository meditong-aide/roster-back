"""조 편성 엑셀 → `nurse_team_period` 반영. **dry-run 기본**, `--apply` 로만 write.

## 알림 미발송
푸시는 `POST /roster/publish` **라우터 안에서만** 호출된다. 이 스크립트는 라우터를
타지 않고 `set_team_period` 로 직접 write 하며 **푸시 모듈을 import 조차 하지 않는다.**
(`import_wanted_from_excel.py` 와 같은 규약 — 이 파일에 푸시 경로를 절대 넣지 말 것.)

## 왜 컬럼이 아니라 period 인가
`nurses.team_id` 는 '현재값' 캐시이고 월별 team 의 진실은 `nurse_team_period` 다.
컬럼만 UPDATE 하면 as-of 조회가 옛 값을 돌려줘 **생성 결과가 안 바뀐다.**
`set_team_period` 가 유일한 쓰기 창구다 — close-before-open(옛 구간을 valid_from 으로
닫고 새 구간을 연다). 삭제는 하지 않는다(완전 타임라인).

★ **캐시는 투영하지 않는다.** `upsert_period` 와 달리 `set_team_period` 본문에
  `nurses.team_id` 대입이 없다(실측 2026-08-19). 미래 발효분(`--valid-from` 이 내일
  이후)이면 어차피 투영 대상도 아니다 — `upsert_period` 도 `valid_from <= today` 일
  때만 캐시를 건드린다. 즉 적용 직후 캐시와 9월 as-of 가 다른 것은 **정상**이다.

## 엑셀 서식 가정
시트 하나에 **두 개 조**가 순서대로 들어 있다(실측 2026-09 42병동 '근무신청표.xlsx').

    [A,B조]  r5~r16  12명 → 앞 6명 team1 · 뒤 6명 team2
    [C,D조]  r5~r16  12명 → 앞 6명 team3 · 뒤 6명 team4

    A열 = 순번(1..) · B열 = 이름. 순번이 끊기면 명단 끝으로 본다.
    r17 이후의 D/E/N/DE 는 근무유형 범례라 명단이 아니다.

★ 날짜 칸은 낡은 헤더가 남아 있어도(2024년 5월·8월) 쓰지 않는다 — **이름만** 읽는다.

## 엑셀에 없는 재직자
`--drop-missing` 을 주면 그 그룹의 나머지 인원을 team=None 으로 **해제**한다.
기본은 건드리지 않는다(누락인지 의도인지 엑셀만으로는 못 가른다).
★ 고정근무자는 애초에 조 편성 대상이 아니라 기본에서도 해제 대상이 아니다.

## 실행

    cd roster-back
    uv run python scripts/apply_team_from_excel.py --file <xlsx> \\
        --group-id 102243e39fbf --valid-from 2026-09-01 \\
        --sheet "A,B조=1,2" --sheet "C,D조=3,4" \\
        [--squad-size 6] [--drop-missing] [--apply]

`--sheet <시트명>=<앞조 team>,<뒤조 team>` 을 조 수만큼 준다.
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, "app")

import openpyxl  # noqa: E402
from db.client2 import SessionLocal  # noqa: E402
from db.models import Nurse  # noqa: E402
from services.team_period import resolve_teams_for_month, set_team_period  # noqa: E402

#: 조당 인원 **기본값**. `--squad-size` 로 덮는다. 시트 12명 = 6+6 이 표준.
SQUAD_SIZE = 6


def norm_team(v: object) -> int | None:
    """team 값을 int 로 정규화.

    ★ `nurses.team_id` 는 모델이 INTEGER 인데 **DB 는 varchar(50)** 이다(실측
      2026-08-19: nurses=varchar / nurse_team_period·teams=int). SQLAlchemy 가
      DB 가 준 문자열을 그대로 돌려주므로 `n.team_id == 1` 이 **항상 False** 다.
      비교 전에 반드시 정규화한다.
    """
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def read_squad(path: str, sheet: str) -> list[str]:
    """시트에서 순번(A열)이 이어지는 동안 이름(B열)을 읽는다."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"시트 없음: {sheet!r} (있는 시트: {wb.sheetnames})")
    ws = wb[sheet]
    out: list[str] = []
    for r in range(1, ws.max_row + 1):
        no, name = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not isinstance(no, (int, float)):
            if out:
                break            # 명단이 시작된 뒤 순번이 끊기면 끝
            continue
        if name is None or not str(name).strip():
            break
        out.append(" ".join(str(name).split()))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    ap.add_argument("--group-id", required=True)
    ap.add_argument("--valid-from", required=True, help="적용 시작일 YYYY-MM-DD")
    ap.add_argument("--sheet", action="append", required=True,
                    metavar="시트명=앞조team,뒤조team")
    ap.add_argument("--squad-size", type=int, default=SQUAD_SIZE,
                    help=f"조당 인원(기본 {SQUAD_SIZE})")
    ap.add_argument("--drop-missing", action="store_true",
                    help="엑셀에 없는 재직자를 team=None 으로 해제")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args()

    if not Path(a.file).exists():
        raise SystemExit(f"파일 없음: {a.file}")
    valid_from = date.fromisoformat(a.valid_from)
    squad = int(a.squad_size)

    plan: dict[str, int] = {}          # 이름 → team
    for spec in a.sheet:
        if "=" not in spec:
            raise SystemExit(f"--sheet 형식 오류(시트명=앞,뒤): {spec!r}")
        sheet, teams = spec.rsplit("=", 1)
        try:
            t_head, t_tail = (int(x) for x in teams.split(","))
        except ValueError:
            raise SystemExit(f"--sheet team 형식 오류: {teams!r}")
        names = read_squad(a.file, sheet)
        # ★ 6+6 가정을 어기면 조용히 비대칭 배정된다(10명이면 6/4). 명시적으로 막는다.
        if len(names) != squad * 2:
            raise SystemExit(
                f"{sheet}: {len(names)}명 — 조당 {squad}명 가정과 다릅니다. "
                f"엑셀을 확인하거나 --squad-size 로 명시하십시오.")
        print(f"[명단] {sheet}: {len(names)}명 → 앞 {squad}명 team{t_head} · "
              f"뒤 {squad}명 team{t_tail}")
        for i, nm in enumerate(names):
            plan[nm] = t_head if i < squad else t_tail

    db = SessionLocal()
    try:
        ns = db.query(Nurse).filter(Nurse.group_id == a.group_id,
                                    Nurse.active == 1).all()
        by_name: dict[str, list] = {}
        for n in ns:
            by_name.setdefault(n.name, []).append(n)

        dup = {k for k, v in by_name.items() if len(v) > 1}
        miss = [k for k in plan if k not in by_name]
        print(f"\n[매칭] 엑셀 {len(plan)}명 / DB active {len(ns)}명 · "
              f"미매칭 {len(miss)} · 동명이인 {len(dup)}")
        if miss:
            print(f"   ★ DB 에 없는 이름: {miss}")
        if dup:
            print(f"   ★ 동명이인(수동 확인 필요): {sorted(dup)}")
        if miss or dup:
            raise SystemExit("이름 해소 실패 — 중단합니다.")

        # ★ 판정 기준은 캐시(nurses.team_id)가 아니라 **period as-of** 다.
        #   캐시와 SSOT 가 갈린 사람을 캐시로 보면 "동일" 로 분류해 period 를 안 쓰고,
        #   그러면 생성은 계속 옛 팀으로 돈다(이 파일 docstring 의 전제와도 어긋난다).
        cur = resolve_teams_for_month(db, a.group_id, valid_from) or {}
        changes, keeps = [], []
        for nm, t in plan.items():
            n = by_name[nm][0]
            now = norm_team(cur.get(str(n.nurse_id)))
            (changes if now != t else keeps).append((n, t))

        extra = []
        if a.drop_missing:
            fixed_names = set()
            for n in ns:
                if n.name in plan:
                    continue
                if str(getattr(n, "fixed_shift", "") or "").strip():
                    fixed_names.add(n.name)      # 고정근무자는 조 편성 대상 아님
                    continue
                if norm_team(cur.get(str(n.nurse_id))) is not None:
                    extra.append(n)
            if fixed_names:
                print(f"   (고정근무자 제외: {sorted(fixed_names)})")

        # ★ 캐시(nurses.team_id)와 SSOT 가 실제로 갈리는 인원 — 판정을 왜 period 로
        #   해야 하는지의 근거를 실행마다 남긴다.
        # ★★ 비교는 **같은 기준일(오늘)** 로 해야 한다. valid_from 이 미래면 캐시가
        #   아직 옛 값인 것이 정상이라(set_team_period 는 캐시를 투영하지 않는다),
        #   valid_from as-of 와 캐시를 맞대면 미래 발효분이 전부 오탐으로 잡힌다.
        today_map = resolve_teams_for_month(db, a.group_id, date.today()) or {}
        drift = [n.name for n in ns
                 if norm_team(n.team_id) != norm_team(today_map.get(str(n.nurse_id)))]
        if drift:
            print(f"\n[주의] 캐시 ≠ period(오늘 기준) 인 인원 {len(drift)}명 {drift}"
                  f"  — 판정은 period 기준으로 한다")

        print(f"\n[변경] {len(changes)}건 · 동일 {len(keeps)}건 · 해제 {len(extra)}건")
        for n, t in sorted(changes, key=lambda x: (x[1], x[0].name)):
            print(f"   {n.name:<8} team {norm_team(cur.get(str(n.nurse_id)))} → {t}")
        for n in extra:
            print(f"   {n.name:<8} team {norm_team(cur.get(str(n.nurse_id)))} → None  (엑셀 미기재)")

        if not a.apply:
            print("\n[dry-run] --apply 를 붙여야 실제로 씁니다.")
            return 0

        for n, t in changes:
            set_team_period(db, nurse_id=str(n.nurse_id), group_id=a.group_id,
                            valid_from=valid_from, team_id=t,
                            source="edited", note="조편성 엑셀 반영", commit=False)
        for n in extra:
            set_team_period(db, nurse_id=str(n.nurse_id), group_id=a.group_id,
                            valid_from=valid_from, team_id=None,
                            source="edited", note="조편성 엑셀 미기재", commit=False)
        db.commit()

        # 검증 — as-of 로 다시 읽어 계획과 대조한다(캐시가 아니라 period 가 진실)
        got = resolve_teams_for_month(db, a.group_id, valid_from) or {}
        bad = []
        for nm, t in plan.items():
            nid = str(by_name[nm][0].nurse_id)
            if norm_team(got.get(nid)) != t:
                bad.append((nm, t, got.get(nid)))
        for n in extra:
            if norm_team(got.get(str(n.nurse_id))) is not None:
                bad.append((n.name, None, got.get(str(n.nurse_id))))
        print(f"\n[반영] {len(changes) + len(extra)}건")
        print(f"[검증] as-of {valid_from.year}-{valid_from.month:02d} 불일치 {len(bad)} {bad or ''}")
        if bad:
            raise SystemExit("검증 실패")
        print("OK")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
